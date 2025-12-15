"""
Основной модуль приложения для обработки месячных данных.

Собирает данные из файлов в каталогах OD, RA, PS,
формирует сводный файл с уникальными табельными номерами
и сохраняет результат в форматированном Excel файле.

Все настройки и конфигурация находятся в этом файле.
"""

import logging
import os
import sys
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import partial

import pandas as pd

# Попытка импортировать openpyxl для форматирования (обычно доступен в Anaconda)
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# xlsxwriter удален - используется только openpyxl


# ============================================================================
# НАСТРОЙКИ ПРИЛОЖЕНИЯ
# ============================================================================

# Пути к каталогам
INPUT_DIR = "IN"  # Каталог с входными данными
OUTPUT_DIR = "OUT"  # Каталог для выходных файлов
LOG_DIR = "log"  # Каталог для логов

# Уровень логирования (INFO или DEBUG)
LOG_LEVEL = "DEBUG"  # Уровень логирования: DEBUG (в файлы) - детальное, INFO (в консоль) - верхнеуровневое

# Тема логов (используется в имени файла)
LOG_THEME = "processor"

# Включить/выключить сбор и вывод статистики
ENABLE_STATISTICS = True  # True - собирать статистику и создавать лист "Статистика", False - не собирать

# Параметры оптимизации производительности
ENABLE_PARALLEL_LOADING = True  # True - параллельная загрузка файлов, False - последовательная
MAX_WORKERS = 16  # Количество потоков для параллельной загрузки (рекомендуется 8 по числу виртуальных ядер)
ENABLE_CHUNKING = False  # True - использовать chunking для больших файлов, False - загружать целиком (chunking медленный, отключен)
CHUNK_SIZE = 50000  # Размер chunk для чтения больших файлов (строк)
CHUNKING_THRESHOLD_MB = 200  # Порог размера файла для chunking (МБ) - если файл больше, используем chunking

# Параметры детального логирования
DEBUG_TAB_NUMBER: Optional[List[str]] = ["08346532", "01378623", "00406092", "00755745", "01778882"]  # Список табельных номеров для детального логирования (например, ["12345678", "87654321"] или None для отключения)
# Если указан список, в лог будет записываться подробная информация о всех операциях с этими табельными номерами
# Если список пустой или None, детальное логирование отключено

# Параметр выбора режима данных
DATA_MODE: str = "TEST"  # "TEST" - тестовые данные, "PROM" - пром данные
# Определяет, какие columns использовать из конфигурации (columns_test или columns_prom)

# Триггер для формирования RAW листов
ENABLE_RAW_SHEETS: bool = False  # True - формировать RAW листы, False - не формировать (по умолчанию выключено)

# Триггер для форматирования листов
# "full" - полное форматирование (как сейчас, по умолчанию)
# "off" - форматирование выключено (листы формируются, но не переформатируются, кроме ТН и ИНН - их форматы всегда работают)
# "simple" - упрощенное форматирование (только ТН, ИНН, ФИО, ТБ, ГОСБ и заголовок, не форматируем данные показателей и расчетов)
FORMATTING_MODE: str = "simple"  # "full", "off", "simple"


# ============================================================================
# КОНФИГУРАЦИЯ МАППИНГА ТЕРРИТОРИАЛЬНЫХ БАНКОВ (ТБ)
# ============================================================================

@dataclass
class TBMapping:
    """
    Маппинг для территориального банка (ТБ).
    
    Параметры:
        short_name_en: Короткое имя на английском (например, "BB", "VVB", "DVB")
        short_name: Короткое имя, которое используется в файлах и расчетах (например, "ББ", "ВВБ", "ДВБ")
        aliases: Массив алиасов (короткое и длинное имя, например, ["ББ", "Байкальский банк"])
    
    Примеры:
        TBMapping(short_name_en="BB", short_name="ББ", aliases=["ББ", "Байкальский банк"])
        TBMapping(short_name_en="VVB", short_name="ВВБ", aliases=["ВВБ", "Волго-Вятский банк"])
    """
    short_name_en: str  # Короткое имя на английском
    short_name: str     # Короткое имя, используемое в файлах и расчетах
    aliases: List[str]  # Массив алиасов (короткое и длинное имя)


# Маппинг всех территориальных банков
TB_MAPPINGS: Dict[str, TBMapping] = {
    "BB": TBMapping(
        short_name_en="BB",
        short_name="ББ",
        aliases=["ББ", "Байкальский банк", "BB"]
    ),
    "VVB": TBMapping(
        short_name_en="VVB",
        short_name="ВВБ",
        aliases=["ВВБ", "Волго-Вятский банк", "VVB"]
    ),
    "DVB": TBMapping(
        short_name_en="DVB",
        short_name="ДВБ",
        aliases=["ДВБ", "Дальневосточный банк", "DVB"]
    ),
    "MB": TBMapping(
        short_name_en="MB",
        short_name="МБ",
        aliases=["МБ", "Московский банк", "MB"]
    ),
    "PVB": TBMapping(
        short_name_en="PVB",
        short_name="ПВБ",
        aliases=["ПВБ", "Поволжский банк", "PVB", "ПБ"]
    ),
    "SZB": TBMapping(
        short_name_en="SZB",
        short_name="СЗБ",
        aliases=["СЗБ", "Северо-Западный банк", "SZB"]
    ),
    "SIB": TBMapping(
        short_name_en="SIB",
        short_name="СИБ",
        aliases=["СИБ", "Сибирский банк", "SIB"]
    ),
    "SRB": TBMapping(
        short_name_en="SRB",
        short_name="СРБ",
        aliases=["СРБ", "Среднерусский банк", "SRB"]
    ),
    "UB": TBMapping(
        short_name_en="UB",
        short_name="УБ",
        aliases=["УБ", "Уральский банк", "UB"]
    ),
    "CA": TBMapping(
        short_name_en="CA",
        short_name="ЦА",
        aliases=["ЦА", "Центральный аппарат", "CA"]
    ),
    "CZB": TBMapping(
        short_name_en="CZB",
        short_name="ЦЧБ",
        aliases=["ЦЧБ", "Центрально-Черноземный банк", "CZB"]
    ),
    "SWB": TBMapping(
        short_name_en="SWB",
        short_name="ЮЗБ",
        aliases=["ЮЗБ", "Юго-Западный банк", "SWB"]
    ),
}


def normalize_tb_value(value: Any) -> Optional[str]:
    """
    Нормализует значение ТБ, приводя его к стандартному короткому имени.
    
    Ищет значение в алиасах всех ТБ и возвращает стандартное короткое имя.
    Если значение не найдено, возвращает None.
    
    Args:
        value: Значение ТБ из файла (может быть любым алиасом)
    
    Returns:
        Optional[str]: Стандартное короткое имя ТБ или None, если не найдено
    
    Примеры:
        normalize_tb_value("ББ") -> "ББ"
        normalize_tb_value("Байкальский банк") -> "ББ"
        normalize_tb_value("BB") -> "ББ" (если "BB" есть в алиасах)
        normalize_tb_value("Неизвестный") -> None
    """
    if value is None or pd.isna(value):
        return None
    
    # Преобразуем в строку и очищаем
    value_str = str(value).strip()
    if not value_str or value_str.lower() in ['nan', 'none', '']:
        return None
    
    # Ищем в алиасах всех ТБ
    for tb_mapping in TB_MAPPINGS.values():
        # Проверяем точное совпадение (без учета регистра)
        for alias in tb_mapping.aliases:
            if value_str.lower() == alias.lower():
                return tb_mapping.short_name
    
    # Если не найдено, возвращаем None
    return None


def get_tb_short_name_en(short_name: str) -> Optional[str]:
    """
    Получает короткое имя ТБ на английском по короткому имени на русском.
    
    Args:
        short_name: Короткое имя ТБ на русском (например, "ББ", "ВВБ")
    
    Returns:
        Optional[str]: Короткое имя на английском (например, "BB", "VVB") или None
    """
    for tb_mapping in TB_MAPPINGS.values():
        if tb_mapping.short_name == short_name:
            return tb_mapping.short_name_en
    return None


def get_tb_aliases(short_name: str) -> Optional[List[str]]:
    """
    Получает список всех алиасов для ТБ по короткому имени.
    
    Args:
        short_name: Короткое имя ТБ на русском (например, "ББ", "ВВБ")
    
    Returns:
        Optional[List[str]]: Список алиасов или None
    """
    for tb_mapping in TB_MAPPINGS.values():
        if tb_mapping.short_name == short_name:
            return tb_mapping.aliases.copy()
    return None


# ============================================================================
# КОНФИГУРАЦИЯ ЗАГРУЗКИ ФАЙЛОВ
# ============================================================================

@dataclass
class DropRule:
    """
    Правило удаления строк.
    
    Параметры:
        alias: Имя поля после маппинга (из default_columns, например "tb", "status")
        values: Список запрещенных значений (строки будут удалены, если значение поля совпадает с одним из них)
        remove_unconditionally: True - удалять всегда (по умолчанию), False - не удалять (правило игнорируется)
        check_by_inn: True - не удалять строку, если по этому ИНН (client_id) есть другие строки с незапрещенными значениями
        check_by_tn: True - не удалять строку, если по этому ТН (tab_number) есть другие строки с незапрещенными значениями
    
    Комбинации:
        - remove_unconditionally=True, check_by_inn=False, check_by_tn=False: удаляем ВСЕ строки с запрещенными значениями
        - remove_unconditionally=True, check_by_inn=True: удаляем, НО если по ИНН есть другие значения - не удаляем
        - remove_unconditionally=True, check_by_tn=True: удаляем, НО если по ТН есть другие значения - не удаляем
        - remove_unconditionally=True, check_by_inn=True, check_by_tn=True: удаляем, НО если по ИНН ИЛИ по ТН есть другие значения - не удаляем (логика ИЛИ)
        - remove_unconditionally=False: строки НЕ удаляются, правило игнорируется
    """
    alias: str
    values: List[str]
    remove_unconditionally: bool = True
    check_by_inn: bool = False
    check_by_tn: bool = False


@dataclass
class TBMapping:
    """
    Маппинг для территориального банка (ТБ).
    
    Параметры:
        short_name_en: Короткое имя на английском (например, "BB", "VVB", "DVB")
        short_name: Короткое имя, которое используется в файлах и расчетах (например, "ББ", "ВВБ", "ДВБ")
        aliases: Массив алиасов (короткое и длинное имя, например, ["ББ", "Байкальский банк"])
    
    Примеры:
        TBMapping(short_name_en="BB", short_name="ББ", aliases=["ББ", "Байкальский банк"])
        TBMapping(short_name_en="VVB", short_name="ВВБ", aliases=["ВВБ", "Волго-Вятский банк"])
    """
    short_name_en: str  # Короткое имя на английском
    short_name: str     # Короткое имя, используемое в файлах и расчетах
    aliases: List[str]  # Массив алиасов (короткое и длинное имя)


@dataclass
class IncludeRule:
    """
    Правило включения строк.
    
    Строка попадает в расчет только если она проходит ВСЕ условия из in_rules (логика И).
    И при этом НЕ попадает под drop_rules (исключается из DROP).
    
    Параметры:
        alias: Имя поля после маппинга (из default_columns, например "type", "tb")
        values: Список значений для проверки
        condition: "in" - значение должно быть в списке values, "not_in" - значение НЕ должно быть в списке
    
    Примеры:
        IncludeRule(alias="type", values=["Активен"], condition="in") - только строки с type="Активен"
        IncludeRule(alias="tb", values=["ЦА"], condition="not_in") - только строки где tb НЕ равно "ЦА"
    """
    alias: str
    values: List[str]
    condition: str = "in"  # "in" или "not_in"


@dataclass
class FileItem:
    """
    Элемент конфигурации для одного файла.
    
    Структура соответствует примеру из репозитория:
    - key: ключ файла (для идентификации)
    - label: подпись для логов
    - file_name: имя файла в каталоге IN (если пустое "", файл не используется)
    - sheet: название листа (если None, используется default_sheet из группы)
    - columns: список колонок (если пустой [], используются из defaults.columns)
    - filters: словарь с drop_rules и in_rules (если пустые [], используются из defaults)
    - calculation_type: тип расчета для второго листа (1, 2, 3 или None - использовать default)
    - first_month_value: значение для первого месяца при расчете типа 2 (None - использовать default)
    """
    # Ключ файла (для идентификации, например "OD_01", "OD_02")
    key: str
    
    # Подпись для логов (например "OD Январь", "OD Февраль")
    label: str
    
    # Имя файла в каталоге IN (например "OD_01.xlsx")
    # Если пустое "", файл не используется
    file_name: str
    
    # Название листа для чтения (если None, используется default_sheet из группы)
    sheet: Optional[str] = None
    
    # Колонки для этого файла (если пустой массив [], используются из defaults.columns_test или defaults.columns_prom в зависимости от DATA_MODE)
    # Формат: [{"alias": "tb", "source": "Короткое ТБ"}, ...]
    columns: List[Dict[str, str]] = field(default_factory=list)
    
    # Фильтры для этого файла
    # Формат: {"drop_rules": [...], "in_rules": [...]}
    # drop_rules: список словарей {"alias": "...", "values": [...], "remove_unconditionally": True, ...}
    # in_rules: список словарей {"alias": "...", "values": [...], "condition": "in" или "not_in"}
    # Если drop_rules или in_rules пустые массивы [], используются из defaults
    filters: Dict[str, Any] = field(default_factory=dict)
    
    # Тип расчета для второго листа (1, 2, 3 или None - использовать default из группы)
    # 1: Как есть - просто сумма
    # 2: Прирост по 2 месяцам (текущий - предыдущий)
    # 3: Прирост по трем периодам (М-3 - 2*М-2 + М-1)
    calculation_type: Optional[int] = None
    
    # Значение для первого месяца при расчете типа 2
    # "self" - равен самому себе (сумме в этом месяце)
    # "zero" - равен 0
    # None - использовать default из группы
    first_month_value: Optional[str] = None
    
    # Правила для первого и второго месяца при расчете типа 3
    # "zero_both" - первый и второй месяц оба равны 0
    # "zero_first_diff_second" - первый равен 0, второй равен разнице между вторым и первым
    # "self_first_diff_second" - первый равен самому себе, второй равен разнице между вторым и первым
    # None - использовать default из группы
    three_periods_first_months: Optional[str] = None
    


@dataclass
class DefaultsConfig:
    """
    Настройки по умолчанию для группы файлов.
    
    Все эти настройки используются, если в FileItem не указаны индивидуальные значения.
    """
    # Колонки по умолчанию для тестовых данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
    columns_test: List[Dict[str, str]] = field(default_factory=list)
    
    # Колонки по умолчанию для пром данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
    columns_prom: List[Dict[str, str]] = field(default_factory=list)
    
    # Правила удаления строк по умолчанию (drop_rules)
    drop_rules: List[DropRule] = field(default_factory=list)
    
    # Правила включения строк по умолчанию (in_rules)
    in_rules: List[IncludeRule] = field(default_factory=list)
    
    # Имена колонок после маппинга (используются alias)
    tab_number_column: str = "tab_number"
    tb_column: str = "tb"
    gosb_column: str = "gosb"
    fio_column: str = "fio"
    indicator_column: str = "indicator"
    
    # Параметры нормализации данных
    tab_number_length: int = 8  # Длина табельного номера с лидирующими нулями
    tab_number_fill_char: str = "0"  # Символ для заполнения табельного номера
    inn_length: int = 12  # Длина ИНН с лидирующими нулями
    inn_fill_char: str = "0"  # Символ для заполнения ИНН
    
    # Параметры обработки файлов
    header_row: Optional[int] = 0
    skip_rows: int = 0
    skip_footer: int = 0
    sheet_name: Optional[str] = None
    sheet_index: Optional[int] = None
    
    # Параметры расчета для второго листа
    # Тип расчета: 1 - как есть, 2 - прирост по 2 месяцам, 3 - прирост по трем периодам
    calculation_type: int = 1
    
    # Значение для первого месяца при расчете типа 2: "self" или "zero"
    first_month_value: str = "self"
    
    # Правила для первого и второго месяца при расчете типа 3
    # "zero_both", "zero_first_diff_second", "self_first_diff_second"
    three_periods_first_months: str = "zero_both"
    
    # Направление показателя для расчета лучшего месяца (вариант 3 с нормализацией)
    # "MAX" - большее значение лучше, "MIN" - меньшее значение лучше
    # Используется при нормализации показателей перед расчетом Score
    indicator_direction: str = "MAX"
    
    # Вес для расчета итогового Score (для данной группы)
    # Score (M-X) = OD_norm(M-X) * weight_OD + RA_norm(M-X) * weight_RA + PS_norm(M-X) * weight_PS
    # В каждом разделе (OD, RA, PS) задается только свой вес
    weight: float = 0.33


@dataclass
class GroupConfig:
    """Конфигурация для группы файлов (OD, RA, PS)."""
    # Название группы
    name: str
    
    # Лист по умолчанию (если у конкретного файла другой лист, задайте его в items)
    default_sheet: str = "Sheet1"
    
    # Список файлов (items) - для каждого файла указываем key, label, file_name и параметры
    # Если file_name пустое "", файл не используется
    # Если columns или filters.drop_rules пустые массивы [], используются значения из defaults
    items: List[FileItem] = field(default_factory=list)
    
    # Настройки по умолчанию для этой группы
    defaults: DefaultsConfig = field(default_factory=DefaultsConfig)


class ConfigManager:
    """Менеджер конфигурации для управления настройками загрузки файлов."""
    
    def __init__(self):
        """Инициализация менеджера конфигурации с настройками по умолчанию."""
        self.groups: Dict[str, GroupConfig] = self._create_default_configs()
    
    def _create_default_configs(self) -> Dict[str, GroupConfig]:
        """
        Создает конфигурации по умолчанию для всех групп.
        
        Returns:
            Dict[str, GroupConfig]: Словарь с конфигурациями групп
        """
        configs = {}
        
        # Конфигурация для группы OD (ОперДоход)
        configs["OD"] = GroupConfig(
            name="OD",
            default_sheet="Sheet1",
            items=[
                # Параметры расчета можно задавать для каждого файла индивидуально:
                # - calculation_type: 1, 2, 3 или None (использовать default)
                # - first_month_value: "self", "zero" или None (использовать default)
                # - three_periods_first_months: "zero_both", "zero_first_diff_second", "self_first_diff_second" или None (использовать default)
                # Примеры:
                #   FileItem(..., calculation_type=2, first_month_value="zero")  # Для этого файла тип 2, первый месяц = 0
                #   FileItem(..., calculation_type=3, three_periods_first_months="self_first_diff_second")  # Для этого файла тип 3 с особыми правилами
                # Если параметры не указаны (None), используются значения из defaults
                FileItem(key="OD_01", label="OD Январь", file_name="M-1_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_02", label="OD Февраль", file_name="M-2_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_03", label="OD Март", file_name="M-3_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_04", label="OD Апрель", file_name="M-4_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_05", label="OD Май", file_name="M-5_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_06", label="OD Июнь", file_name="M-6_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_07", label="OD Июль", file_name="M-7_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_08", label="OD Август", file_name="M-8_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_09", label="OD Сентябрь", file_name="M-9_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_10", label="OD Октябрь", file_name="M-10_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_11", label="OD Ноябрь", file_name="M-11_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="OD_12", label="OD Декабрь", file_name="M-12_OD.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
            ],
            defaults=DefaultsConfig(
                # Колонки для тестовых данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
                # Формат: [{"alias": "внутреннее_имя", "source": "Имя в Excel"}, ...]
                # Примеры:
                #   {"alias": "tab_number", "source": "Табельный номер"}
                #   {"alias": "tb", "source": "Короткое ТБ"}
                #   {"alias": "indicator", "source": "Факт"}
                columns_test=[
                    {"alias": "tab_number", "source": "Табельный номер"},
                    {"alias": "tb", "source": "Короткое ТБ"},
                    {"alias": "gosb", "source": "Полное ГОСБ"},
                    {"alias": "client_id", "source": "ИНН"},
                    {"alias": "fio", "source": "ФИО"},
                    {"alias": "indicator", "source": "Факт"}
                ],
                # Колонки для пром данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
                columns_prom=[
                    {"alias": "tab_number", "source": "Таб (8)"},
                    {"alias": "tb", "source": "ТБ"},
                    {"alias": "gosb", "source": "ГОСБ"},
                    {"alias": "client_id", "source": "ИНН"},
                    {"alias": "fio", "source": "КМ"},
                    {"alias": "indicator", "source": "2025, руб."}
                ],
                
                # Правила удаления строк по умолчанию (drop_rules)
                # Формат: [DropRule(alias="...", values=[...], ...), ...]
                # Параметры DropRule:
                #   - alias: имя поля после маппинга (из columns)
                #   - values: список запрещенных значений
                #   - remove_unconditionally: True - удалять всегда, False - не удалять
                #   - check_by_inn: True - не удалять, если по ИНН есть другие значения
                #   - check_by_tn: True - не удалять, если по ТН есть другие значения
                # Примеры:
                #   DropRule(alias="status", values=["Удален", "Архив"], remove_unconditionally=True, check_by_inn=False, check_by_tn=False)
                #   DropRule(alias="tb", values=["ЦА"], remove_unconditionally=True, check_by_inn=True, check_by_tn=False)
                drop_rules=[
                    DropRule(alias="fio", values=["Серая зона"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="client_id", values=["НЕ ОПРЕДЕЛЕН"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                ],
                
                # Правила включения строк по умолчанию (in_rules)
                # Формат: [IncludeRule(alias="...", values=[...], condition="..."), ...]
                # Параметры IncludeRule:
                #   - alias: имя поля после маппинга (из columns)
                #   - values: список разрешенных значений
                #   - condition: "in" - значение должно быть в списке, "not_in" - не должно быть
                # Строка попадает в расчет только если она проходит ВСЕ условия из in_rules (И)
                # Примеры:
                #   IncludeRule(alias="type", values=["Активен"], condition="in")
                #   IncludeRule(alias="tb", values=["ЦА"], condition="not_in")
                in_rules=[
                    # IncludeRule(alias="type", values=["Активен"], condition="in"),
                ],
                
                # Имена колонок после маппинга (используются alias из columns)
                # Эти имена используются для доступа к данным после преобразования
                tab_number_column="tab_number",  # Колонка с табельным номером
                tb_column="tb",                   # Колонка с ТБ (территориальный банк)
                gosb_column="gosb",               # Колонка с ГОСБ (головной офис)
                fio_column="fio",                 # Колонка с ФИО
                indicator_column="indicator",     # Колонка с показателем (факт)
                
                # Параметры обработки файлов
                header_row=0,          # Номер строки с заголовками (0 - первая строка, None - автоматическое определение)
                skip_rows=0,          # Количество строк для пропуска в начале файла
                skip_footer=0,        # Количество строк для пропуска в конце файла
                sheet_name=None,      # Название листа для чтения (None - первый лист)
                sheet_index=None,     # Номер листа для чтения (0 - первый лист, None - использовать sheet_name)
                
                # Параметры расчета для второго листа "Расчеты"
                # Тип расчета (calculation_type):
                #   1 - "Как есть": просто загружаем сумму данных по табельному в указанный месяц (аналог первого листа)
                #   2 - "Прирост по 2 месяцам": текущий месяц - предыдущий месяц
                #      Пример: Февраль М-2 = Февраль М-2 - Январь М-1
                #      Пример: Апрель М-4 = Апрель М-4 - Март М-3
                #   3 - "Прирост по трем периодам": М-N = М-N - 2*М-(N-1) + М-(N-2)
                #      Пример: М-3 = М-3 - 2*М-2 + М-1
                #      Пример: М-4 = М-4 - 2*М-3 + М-2
                calculation_type=2,
                
                # Значение для первого месяца при расчете типа 2 (first_month_value):
                #   "self" - первый месяц равен самому себе (сумме по этому ТН в этом месяце)
                #   "zero" - первый месяц равен 0
                # Пример: если первый месяц = Январь М-1, то:
                #   "self" -> М-1 = сумма по ТН в январе
                #   "zero" -> М-1 = 0
                first_month_value="self",
                
                # Правила для первого и второго месяца при расчете типа 3 (three_periods_first_months):
                #   "zero_both" - первый и второй месяц оба равны 0
                #     Пример: М-1 = 0, М-2 = 0, М-3 = М-3 - 2*М-2 + М-1
                #   "zero_first_diff_second" - первый равен 0, второй равен разнице между вторым и первым
                #     Пример: М-1 = 0, М-2 = М-2 - М-1, М-3 = М-3 - 2*М-2 + М-1
                #   "self_first_diff_second" - первый равен самому себе, второй равен разнице между вторым и первым
                #     Пример: М-1 = М-1 (сумма), М-2 = М-2 - М-1, М-3 = М-3 - 2*М-2 + М-1
                three_periods_first_months="self_first_diff_second",
                
                # Направление показателя для расчета лучшего месяца (вариант 3 с нормализацией)
                # "MAX" - большее значение лучше, "MIN" - меньшее значение лучше
                indicator_direction="MAX",
                
                # Параметры нормализации данных
                tab_number_length=8,      # Длина табельного номера с лидирующими нулями
                tab_number_fill_char="0", # Символ для заполнения табельного номера
                inn_length=12,            # Длина ИНН с лидирующими нулями
                inn_fill_char="0"          # Символ для заполнения ИНН
            )
        )
        
        configs["RA"] = GroupConfig(
            name="RA",
            default_sheet="Sheet1",
            items=[
                FileItem(key="RA_01", label="RA Январь", file_name="M-1_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_02", label="RA Февраль", file_name="M-2_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_03", label="RA Март", file_name="M-3_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_04", label="RA Апрель", file_name="M-4_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_05", label="RA Май", file_name="M-5_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_06", label="RA Июнь", file_name="M-6_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_07", label="RA Июль", file_name="M-7_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_08", label="RA Август", file_name="M-8_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_09", label="RA Сентябрь", file_name="M-9_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_10", label="RA Октябрь", file_name="M-10_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_11", label="RA Ноябрь", file_name="M-11_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="RA_12", label="RA Декабрь", file_name="M-12_RA.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
            ],
            defaults=DefaultsConfig(
                # Колонки для тестовых данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
                # Формат: [{"alias": "внутреннее_имя", "source": "Имя в Excel"}, ...]
                # Примеры:
                #   {"alias": "tab_number", "source": "Табельный номер"}
                #   {"alias": "tb", "source": "Короткое ТБ"}
                #   {"alias": "indicator", "source": "Факт"}
                columns_test=[
                    {"alias": "tab_number", "source": "Табельный номер"},
                    {"alias": "tb", "source": "Короткое ТБ"},
                    {"alias": "gosb", "source": "Полное ГОСБ"},
                    {"alias": "client_id", "source": "ИНН"},
                    {"alias": "fio", "source": "ФИО"},
                    {"alias": "indicator", "source": "Факт"}
                ],
                # Колонки для пром данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
                columns_prom=[
                    {"alias": "tab_number", "source": "Таб. номер ВКО"},
                    {"alias": "tb", "source": "ТБ"},
                    {"alias": "gosb", "source": "ГОСБ"},
                    {"alias": "client_id", "source": "ИНН"},
                    {"alias": "fio", "source": "ВКО"},
                    {"alias": "indicator", "source": "СО РА (M). план курс"}
                ],
                # Правила удаления строк по умолчанию (drop_rules)
                # Формат: [DropRule(alias="...", values=[...], ...), ...]
                # Параметры DropRule:
                #   - alias: имя поля после маппинга (из columns)
                #   - values: список запрещенных значений
                #   - remove_unconditionally: True - удалять всегда, False - не удалять
                #   - check_by_inn: True - не удалять, если по ИНН есть другие значения
                #   - check_by_tn: True - не удалять, если по ТН есть другие значения
                # Примеры:
                #   DropRule(alias="status", values=["Удален", "Архив"], remove_unconditionally=True, check_by_inn=False, check_by_tn=False)
                #   DropRule(alias="tb", values=["ЦА"], remove_unconditionally=True, check_by_inn=True, check_by_tn=False)
                drop_rules=[
                    DropRule(alias="tb", values=["ЦА"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="gosb", values=["9999"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="client_id", values=["0"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="fio", values=["-"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="tab_number", values=["-", "Tech_Sib"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                ],
                in_rules=[],
                tab_number_column="tab_number", tb_column="tb", gosb_column="gosb", fio_column="fio", indicator_column="indicator",
                header_row=0, skip_rows=0, skip_footer=0, sheet_name=None, sheet_index=None,
                calculation_type=1, first_month_value="self", three_periods_first_months="self_first_diff_second",
                indicator_direction="MAX", weight=0.33
            )
        )
        
        # Конфигурация для группы PS (Пассивы)
        configs["PS"] = GroupConfig(
            name="PS",
            default_sheet="Sheet1",
            items=[
                # Параметры расчета можно задавать для каждого файла индивидуально:
                # - calculation_type: 1, 2, 3 или None (использовать default)
                # - first_month_value: "self", "zero" или None (использовать default)
                # - three_periods_first_months: "zero_both", "zero_first_diff_second", "self_first_diff_second" или None (использовать default)
                # Примеры:
                #   FileItem(..., calculation_type=2, first_month_value="zero")  # Для этого файла тип 2, первый месяц = 0
                #   FileItem(..., calculation_type=3, three_periods_first_months="self_first_diff_second")  # Для этого файла тип 3 с особыми правилами
                # Если параметры не указаны (None), используются значения из defaults
                FileItem(key="PS_01", label="PS Январь", file_name="M-1_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_02", label="PS Февраль", file_name="M-2_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_03", label="PS Март", file_name="M-3_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_04", label="PS Апрель", file_name="M-4_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_05", label="PS Май", file_name="M-5_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_06", label="PS Июнь", file_name="M-6_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_07", label="PS Июль", file_name="M-7_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_08", label="PS Август", file_name="M-8_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_09", label="PS Сентябрь", file_name="M-9_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_10", label="PS Октябрь", file_name="M-10_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_11", label="PS Ноябрь", file_name="M-11_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
                FileItem(key="PS_12", label="PS Декабрь", file_name="M-12_PS.xlsx", sheet=None, columns=[], filters={"drop_rules": [], "in_rules": []}, calculation_type=None, first_month_value=None, three_periods_first_months=None),
            ],
            defaults=DefaultsConfig(
                # Колонки для тестовых данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
                # Формат: [{"alias": "внутреннее_имя", "source": "Имя в Excel"}, ...]
                # Примеры:
                #   {"alias": "tab_number", "source": "Табельный номер"}
                #   {"alias": "tb", "source": "Короткое ТБ"}
                #   {"alias": "indicator", "source": "Факт"}
                columns_test=[
                    {"alias": "tab_number", "source": "Табельный номер"},
                    {"alias": "tb", "source": "Короткое ТБ"},
                    {"alias": "gosb", "source": "Полное ГОСБ"},
                    {"alias": "client_id", "source": "ИНН"},
                    {"alias": "fio", "source": "ФИО"},
                    {"alias": "indicator", "source": "Факт"}
                ],
                # Колонки для пром данных: маппинг source (имя в Excel) -> alias (внутреннее имя)
                columns_prom=[
                    {"alias": "tab_number", "source": "Табельный номер ВКО"},
                    {"alias": "tb", "source": "ТБ"},
                    {"alias": "gosb", "source": "ГОСБ"},
                    {"alias": "client_id", "source": "ИНН"},
                    {"alias": "fio", "source": "ВКО"},
                    {"alias": "indicator", "source": "СО за месяц, план курс"}
                ],
                # Правила удаления строк по умолчанию (drop_rules)
                # Формат: [DropRule(alias="...", values=[...], ...), ...]
                # Параметры DropRule:
                #   - alias: имя поля после маппинга (из columns)
                #   - values: список запрещенных значений
                #   - remove_unconditionally: True - удалять всегда, False - не удалять
                #   - check_by_inn: True - не удалять, если по ИНН есть другие значения
                #   - check_by_tn: True - не удалять, если по ТН есть другие значения
                # Примеры:
                #   DropRule(alias="status", values=["Удален", "Архив"], remove_unconditionally=True, check_by_inn=False, check_by_tn=False)
                #   DropRule(alias="tb", values=["ЦА"], remove_unconditionally=True, check_by_inn=True, check_by_tn=False)
                drop_rules=[
                    DropRule(alias="tb", values=["ЦА"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="gosb", values=["9999"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="client_id", values=["0", "-"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="fio", values=["Серая зона", "-"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                    DropRule(alias="tab_number", values=["Серая зона", "-", "0", "00000000", "Tech_UB", "Tech_YZB", "Tech_SRB", "Tech_SRB", "Tech_Sib", "Tech_PB", "TECH_000006", "TECH_000006"], remove_unconditionally=True,
                             check_by_inn=False, check_by_tn=False),
                ],
                in_rules=[],
                tab_number_column="tab_number", tb_column="tb", gosb_column="gosb", fio_column="fio", indicator_column="indicator",
                header_row=0, skip_rows=0, skip_footer=0, sheet_name=None, sheet_index=None,
                calculation_type=1, first_month_value="self", three_periods_first_months="self_first_diff_second",
                indicator_direction="MAX", weight=0.34
            )
        )

        return configs
    
    def get_file_item(self, group_name: str, file_name: str) -> Optional[FileItem]:
        """
        Получает конфигурацию элемента файла (FileItem) по имени файла.
        
        Args:
            group_name: Название группы (OD, RA, PS)
            file_name: Имя файла
            
        Returns:
            Optional[FileItem]: Элемент конфигурации файла или None
        """
        if group_name not in self.groups:
            return None
        
        group_config = self.groups[group_name]
        
        # Ищем файл в items по file_name
        for item in group_config.items:
            if item.file_name == file_name:
                return item
        
        return None
    
    def get_config_for_file(self, group_name: str, file_name: str) -> Dict[str, Any]:
        """
        Получает конфигурацию для конкретного файла.
        
        Args:
            group_name: Название группы (OD, RA, PS)
            file_name: Имя файла
            
        Returns:
            Dict[str, Any]: Конфигурация для файла
        """
        if group_name not in self.groups:
            raise ValueError(f"Неизвестная группа: {group_name}")
        
        group_config = self.groups[group_name]
        
        # Ищем элемент файла в items
        file_item = self.get_file_item(group_name, file_name)
        
        # Получаем defaults из конфигурации группы
        defaults = group_config.defaults
        
        # Формируем итоговую конфигурацию
        # Колонки: если в item есть columns и он не пустой, используем их, иначе defaults
        # Выбираем columns в зависимости от режима DATA_MODE
        if file_item and file_item.columns:
            columns = file_item.columns
        else:
            # Выбираем columns в зависимости от режима DATA_MODE
            if DATA_MODE == "PROM":
                columns = defaults.columns_prom if defaults.columns_prom else defaults.columns_test
            else:
                columns = defaults.columns_test if defaults.columns_test else defaults.columns_prom
        
        # Правила удаления: если в item есть filters.drop_rules и он не пустой, используем их, иначе defaults
        if file_item and file_item.filters.get("drop_rules"):
            # Преобразуем словари в DropRule объекты
            drop_rules = [
                DropRule(
                    alias=rule["alias"],
                    values=rule["values"],
                    remove_unconditionally=rule.get("remove_unconditionally", True),
                    check_by_inn=rule.get("check_by_inn", False),
                    check_by_tn=rule.get("check_by_tn", False)
                ) for rule in file_item.filters["drop_rules"]
            ]
        else:
            drop_rules = defaults.drop_rules
        
        # Правила включения: если в item есть filters.in_rules и он не пустой, используем их, иначе defaults
        if file_item and file_item.filters.get("in_rules"):
            # Преобразуем словари в IncludeRule объекты
            in_rules = [
                IncludeRule(
                    alias=rule["alias"],
                    values=rule["values"],
                    condition=rule.get("condition", "in")
                ) for rule in file_item.filters["in_rules"]
            ]
        else:
            in_rules = defaults.in_rules
        
        # Лист: если в item есть sheet, используем его, иначе default_sheet группы
        sheet_name = file_item.sheet if file_item and file_item.sheet else group_config.default_sheet
        
        # Тип расчета: если в item есть calculation_type, используем его, иначе default
        calculation_type = file_item.calculation_type if file_item and file_item.calculation_type is not None else defaults.calculation_type
        
        # Значение для первого месяца: если в item есть first_month_value, используем его, иначе default
        first_month_value = file_item.first_month_value if file_item and file_item.first_month_value is not None else defaults.first_month_value
        
        # Правила для первого и второго месяца при расчете типа 3: если в item есть three_periods_first_months, используем его, иначе default
        three_periods_first_months = file_item.three_periods_first_months if file_item and file_item.three_periods_first_months is not None else defaults.three_periods_first_months
        
        # Направление показателя для расчета лучшего месяца (вариант 3): используем из defaults
        indicator_direction = defaults.indicator_direction
        
        result = {
            "columns": columns,
            "drop_rules": drop_rules,
            "in_rules": in_rules,
            "tab_number_column": defaults.tab_number_column,
            "tb_column": defaults.tb_column,
            "gosb_column": defaults.gosb_column,
            "fio_column": defaults.fio_column,
            "indicator_column": defaults.indicator_column,
            "header_row": defaults.header_row,
            "skip_rows": defaults.skip_rows,
            "skip_footer": defaults.skip_footer,
            "sheet_name": sheet_name,
            "sheet_index": defaults.sheet_index,
            "calculation_type": calculation_type,
            "first_month_value": first_month_value,
            "three_periods_first_months": three_periods_first_months,
            "indicator_direction": indicator_direction,
            "label": file_item.label if file_item else file_name
        }
        
        return result
    
    def add_file_item(self, group_name: str, file_item: FileItem) -> None:
        """
        Добавляет элемент файла в конфигурацию группы.
        
        Args:
            group_name: Название группы
            file_item: Элемент конфигурации файла
        """
        if group_name not in self.groups:
            raise ValueError(f"Неизвестная группа: {group_name}")
        
        self.groups[group_name].items.append(file_item)
    
    def get_group_config(self, group_name: str) -> GroupConfig:
        """
        Получает конфигурацию группы.
        
        Args:
            group_name: Название группы
            
        Returns:
            GroupConfig: Конфигурация группы
        """
        if group_name not in self.groups:
            raise ValueError(f"Неизвестная группа: {group_name}")
        
        return self.groups[group_name]


# Глобальный экземпляр менеджера конфигурации
config_manager = ConfigManager()


# ============================================================================
# МОДУЛЬ ЛОГИРОВАНИЯ
# ============================================================================

class Logger:
    """Класс для настройки и управления логированием."""
    
    def __init__(self, log_dir: str = LOG_DIR, level: str = LOG_LEVEL, theme: str = LOG_THEME):
        """
        Инициализация логгера.
        
        Args:
            log_dir: Директория для хранения логов
            level: Уровень логирования (INFO или DEBUG)
            theme: Тема логов (используется в имени файла)
        """
        self.log_dir = Path(log_dir)
        self.level = level.upper()
        self.theme = theme
        
        # Создаем директорию для логов, если её нет
        self.log_dir.mkdir(parents=True, exist_ok=True)
        
        # Настраиваем логгер
        self.logger = logging.getLogger("YEAR_SPOD_TOP_Month")
        self.logger.setLevel(getattr(logging, self.level, logging.INFO))
        
        # Очищаем существующие обработчики
        self.logger.handlers.clear()
        
        # Создаем файловый обработчик
        log_file = self._generate_log_filename()
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        
        # Создаем форматтер для DEBUG уровня (с [class: ... | def: ...], но без YEAR_SPOD_TOP_Month и debug)
        debug_formatter = logging.Formatter(
            '%(asctime)s - [%(levelname)s] - %(message)s [class: %(name)s | def: %(funcName)s]',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(debug_formatter)
        
        # Создаем консольный обработчик
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_formatter = logging.Formatter(
            '%(asctime)s - [%(levelname)s] - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        console_handler.setFormatter(console_formatter)
        
        # Добавляем обработчики
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
    
    def _generate_log_filename(self) -> Path:
        """
        Генерирует имя файла лога по шаблону.
        
        Returns:
            Path: Путь к файлу лога
        """
        now = datetime.now()
        filename = f"{self.level}_{self.theme}_{now.strftime('%Y%m%d_%H%M')}.log"
        return self.log_dir / filename
    
    def _mask_tab_number(self, text: str) -> str:
        """
        Маскирует табельные номера в тексте (xxx***xxx - первые 3 и последние 3 символа).
        Маскирует ТОЛЬКО значения полей tab_number, Табельный, ТН и т.д., когда они явно указаны.
        
        Args:
            text: Текст для маскировки
            
        Returns:
            str: Текст с замаскированными табельными номерами
        """
        # Ищем табельные номера только после явных меток полей
        # Паттерны: "tab_number: 12345678", "Табельный: 12345678", "ТН: 12345678", "tab_number='12345678'", и т.д.
        # ВАЖНО: Используем \b для границ слов, чтобы не маскировать числа в других контекстах
        patterns = [
            # С кавычками
            r"\b(tab_number|Табельный|ТН|tab_number_column)\b\s*[:=]\s*['\"](\d{8})['\"]",
            # Без кавычек (требуем пробел или начало строки перед полем)
            r"(?:^|\s)\b(tab_number|Табельный|ТН|tab_number_column)\b\s*[:=]\s*(\d{8})(?=\s|$|,|;|\.|\[|\]|\})",
            # В словарях/структурах
            r"(['\"]tab_number['\"]|['\"]Табельный['\"]|['\"]ТН['\"])\s*:\s*['\"](\d{8})['\"]",
            r"(['\"]tab_number['\"]|['\"]Табельный['\"]|['\"]ТН['\"])\s*:\s*(\d{8})(?=\s|$|,|;|\.|\[|\]|\})",
        ]
        
        def mask_match(match):
            """Маскирует значение табельного номера."""
            # Определяем группу с значением (последняя группа с цифрами)
            groups = match.groups()
            # Ищем группу с 8-значным числом
            tab = None
            for group in reversed(groups):
                if group and re.match(r'^\d{8}$', str(group)):
                    tab = group
                    break
            
            if tab and len(tab) >= 6:
                # Маскируем: первые 3 и последние 3 символа остаются, средние заменяются на ***
                masked = f"{tab[:3]}***{tab[-3:]}"
                return match.group(0).replace(tab, masked)
            return match.group(0)
        
        for pattern in patterns:
            text = re.sub(pattern, mask_match, text, flags=re.IGNORECASE)
        
        return text
    
    def _mask_client_id(self, text: str) -> str:
        """
        Маскирует ИД клиента (ИНН) в тексте (xxx***xxx - первые 3 и последние 3 символа).
        Маскирует ТОЛЬКО значения полей client_id, ИНН и т.д., когда они явно указаны.
        
        Args:
            text: Текст для маскировки
            
        Returns:
            str: Текст с замаскированными ИД клиентов
        """
        # Ищем ИНН только после явных меток полей
        # Паттерны: "client_id: 123456789012", "ИНН: 123456789012", "client_id='123456789012'", и т.д.
        # ВАЖНО: Используем \b для границ слов, чтобы не маскировать числа в других контекстах
        patterns = [
            # С кавычками
            r"\b(client_id|ИНН|client_id_column)\b\s*[:=]\s*['\"](\d{10,12})['\"]",
            # Без кавычек (требуем пробел или начало строки перед полем)
            r"(?:^|\s)\b(client_id|ИНН|client_id_column)\b\s*[:=]\s*(\d{10,12})(?=\s|$|,|;|\.|\[|\]|\})",
            # В словарях/структурах
            r"(['\"]client_id['\"]|['\"]ИНН['\"])\s*:\s*['\"](\d{10,12})['\"]",
            r"(['\"]client_id['\"]|['\"]ИНН['\"])\s*:\s*(\d{10,12})(?=\s|$|,|;|\.|\[|\]|\})",
        ]
        
        def mask_match(match):
            """Маскирует значение ИНН."""
            # Определяем группу с значением (последняя группа с цифрами)
            groups = match.groups()
            # Ищем группу с 10-12-значным числом
            inn = None
            for group in reversed(groups):
                if group and re.match(r'^\d{10,12}$', str(group)):
                    inn = group
                    break
            
            if inn and len(inn) >= 6:
                # Маскируем: первые 3 и последние 3 символа остаются, средние заменяются на ***
                masked = f"{inn[:3]}***{inn[-3:]}"
                return match.group(0).replace(inn, masked)
            return match.group(0)
        
        for pattern in patterns:
            text = re.sub(pattern, mask_match, text, flags=re.IGNORECASE)
        
        return text
    
    def _mask_fio(self, text: str) -> str:
        """
        Маскирует ФИО в тексте: от каждого слова оставляем первые 2 буквы, далее три звездочки и последняя 1 буква.
        Пример: Иванов Иван Сергеевич -> Ив***в Ив***н Се***ч
        
        Маскирует ФИО во всех контекстах: после меток "ФИО:", "ВКО:", "КМ:", "fio:" и т.д.,
        в структурированных данных, в значениях словарей, и просто в тексте (если это похоже на ФИО).
        
        Args:
            text: Текст для маскировки
            
        Returns:
            str: Текст с замаскированными ФИО
        """
        def mask_fio_word(word: str) -> str:
            """Маскирует одно слово ФИО: первые 2 буквы, затем ***, затем последняя буква."""
            if len(word) >= 4:
                return f"{word[:2]}***{word[-1]}"
            elif len(word) >= 3:
                return f"{word[0]}***{word[-1]}"
            else:
                return word
        
        def mask_fio_text(fio_text: str) -> str:
            """Маскирует текст ФИО (может содержать несколько слов)."""
            words = fio_text.split()
            masked_words = [mask_fio_word(word) for word in words]
            return ' '.join(masked_words)
        
        # Паттерн 1: ФИО после меток типа "ФИО:", "ВКО:", "КМ:", "fio:" и т.д. (с двоеточием или равно, с кавычками)
        # Ищем паттерн типа "ФИО='Петров Иван'" или "ВКО: 'Иванов'" или "КМ=\"Сидоров\"" или "fio: 'Иванов'"
        pattern1 = r"(ФИО|ВКО|КМ|fio|FIO)\s*[:=]\s*['\"]([А-ЯЁ][а-яё]{2,}(?:\s+[А-ЯЁ][а-яё]{2,}){0,2})['\"]"
        def replace_fio1(match):
            label = match.group(1)
            fio_text = match.group(2)
            masked = mask_fio_text(fio_text)
            return f"{label}='{masked}'"
        text = re.sub(pattern1, replace_fio1, text, flags=re.IGNORECASE)
        
        # Паттерн 2: ФИО после меток без кавычек (например, "ФИО: Иванов Иван Сергеевич" или "ФИО=Иванов Иван Сергеевич")
        # Ищем паттерн типа "ФИО: Иванов Иван Сергеевич" или "fio: Петров Иван" или "ФИО=Иванов Иван"
        pattern2 = r"(ФИО|ВКО|КМ|fio|FIO)\s*([:=])\s*([А-ЯЁ][а-яё]{2,}(?:\s+[А-ЯЁ][а-яё]{2,}){0,2})(?=\s|$|,|;|\.|\[|\]|\})"
        def replace_fio2(match):
            label = match.group(1)
            separator = match.group(2)
            fio_text = match.group(3)
            masked = mask_fio_text(fio_text)
            return f"{label}{separator} {masked}"
        text = re.sub(pattern2, replace_fio2, text, flags=re.IGNORECASE)
        
        # Паттерн 3: ФИО в контексте "ФИО='...'" или "ФИО=\"...\"" (с кавычками)
        pattern3 = r"(ФИО\s*=\s*['\"])([А-ЯЁ][а-яё]{2,}(?:\s+[А-ЯЁ][а-яё]{2,}){0,2})(['\"])"
        def replace_fio3(match):
            prefix = match.group(1)
            fio_text = match.group(2)
            suffix = match.group(3)
            masked = mask_fio_text(fio_text)
            return prefix + masked + suffix
        text = re.sub(pattern3, replace_fio3, text, flags=re.IGNORECASE)
        
        # Паттерн 4: ФИО в структурированных данных типа "{'ФИО': 'Петров Иван'}" или "{'fio': 'Иванов'}"
        pattern4 = r"(['\"]ФИО['\"]|['\"]fio['\"])\s*:\s*['\"]([А-ЯЁ][а-яё]{2,}(?:\s+[А-ЯЁ][а-яё]{2,}){0,2})['\"]"
        def replace_fio4(match):
            key = match.group(1)
            fio_text = match.group(2)
            masked = mask_fio_text(fio_text)
            return f"{key}: '{masked}'"
        text = re.sub(pattern4, replace_fio4, text, flags=re.IGNORECASE)
        
        # Паттерн 5: ФИО в структурированных данных без кавычек типа "{'ФИО': Петров Иван}" (редкий случай)
        pattern5 = r"(['\"]ФИО['\"]|['\"]fio['\"])\s*:\s+([А-ЯЁ][а-яё]{2,}(?:\s+[А-ЯЁ][а-яё]{2,}){0,2})(?=\s|$|,|;|\.|\[|\})"
        def replace_fio5(match):
            key = match.group(1)
            fio_text = match.group(2)
            masked = mask_fio_text(fio_text)
            return f"{key}: {masked}"
        text = re.sub(pattern5, replace_fio5, text, flags=re.IGNORECASE)
        
        # Удаляем паттерн 6 - он слишком агрессивный и может маскировать значения индикаторов
        # Маскируем только явные поля fio/ФИО/ВКО/КМ
        
        return text
    
    def _mask_sensitive_data(self, text: str) -> str:
        """
        Маскирует все чувствительные данные (табельные номера, ИД клиентов и ФИО).
        
        Args:
            text: Текст для маскировки
            
        Returns:
            str: Текст с замаскированными данными
        """
        text = self._mask_tab_number(text)
        text = self._mask_client_id(text)
        text = self._mask_fio(text)
        return text
    
    def _is_debug_tab_number(self, tab_number: Any) -> bool:
        """
        Проверяет, является ли табельный номер тем, для которого нужно детальное логирование.
        Поддерживает список табельных номеров.
        
        Args:
            tab_number: Табельный номер для проверки
            
        Returns:
            bool: True, если это табельный номер для детального логирования
        """
        if DEBUG_TAB_NUMBER is None or not DEBUG_TAB_NUMBER or len(DEBUG_TAB_NUMBER) == 0:
            return False
        
        if tab_number is None or pd.isna(tab_number):
            return False
        
        # Нормализуем табельный номер для сравнения
        tab_str = str(tab_number).strip().lstrip('0')
        
        # Проверяем, есть ли этот табельный номер в списке
        for debug_tab in DEBUG_TAB_NUMBER:
            if debug_tab is None:
                continue
            debug_tab_str = str(debug_tab).strip().lstrip('0')
            if tab_str == debug_tab_str:
                return True
        
        return False
    
    def debug_tab(self, message: str, tab_number: Any = None, class_name: Optional[str] = None, func_name: Optional[str] = None) -> None:
        """
        Детальное логирование для указанного табельного номера.
        Логирует только если DEBUG_TAB_NUMBER указан (список) и содержит tab_number.
        
        Args:
            message: Сообщение для логирования
            tab_number: Табельный номер (опционально, для проверки)
            class_name: Имя класса (опционально)
            func_name: Имя функции (опционально)
        """
        # Если DEBUG_TAB_NUMBER не указан или пустой список, ничего не делаем
        if DEBUG_TAB_NUMBER is None or not DEBUG_TAB_NUMBER or len(DEBUG_TAB_NUMBER) == 0:
            return
        
        # Если указан tab_number, проверяем совпадение
        if tab_number is not None:
            if not self._is_debug_tab_number(tab_number):
                return
        
        # Маскируем чувствительные данные (включая ФИО)
        masked_message = self._mask_sensitive_data(message)
        
        # Форматируем сообщение
        if class_name and func_name:
            clean_class = class_name.replace("YEAR_SPOD_TOP_Month", "").strip()
            if clean_class:
                formatted_message = f"[ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ ТН] {masked_message} [class: {clean_class} | def: {func_name}]"
            else:
                formatted_message = f"[ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ ТН] {masked_message} [def: {func_name}]"
        else:
            formatted_message = f"[ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ ТН] {masked_message}"
        
        self.logger.debug(formatted_message)
    
    def info(self, message: str, class_name: Optional[str] = None, func_name: Optional[str] = None) -> None:
        """
        Логирует сообщение уровня INFO.
        
        Args:
            message: Сообщение для логирования
            class_name: Имя класса (опционально)
            func_name: Имя функции (опционально)
        """
        # Маскируем чувствительные данные (табельные номера и ИД клиентов)
        masked_message = self._mask_sensitive_data(message)
        # Форматируем сообщение с классом и функцией (если указаны), но убираем только YEAR_SPOD_TOP_Month
        if class_name and func_name:
            # Убираем YEAR_SPOD_TOP_Month из class_name, если есть
            clean_class = class_name.replace("YEAR_SPOD_TOP_Month", "").strip()
            if clean_class:
                formatted_message = f"{masked_message} [class: {clean_class} | def: {func_name}]"
            else:
                formatted_message = f"{masked_message} [def: {func_name}]"
            self.logger.info(formatted_message)
        else:
            self.logger.info(masked_message)
    
    def debug(self, message: str, class_name: Optional[str] = None, func_name: Optional[str] = None) -> None:
        """
        Логирует сообщение уровня DEBUG.
        
        Args:
            message: Сообщение для логирования
            class_name: Имя класса (опционально)
            func_name: Имя функции (опционально)
        """
        # Маскируем чувствительные данные (табельные номера и ИД клиентов)
        masked_message = self._mask_sensitive_data(message)
        # Форматируем сообщение с классом и функцией (если указаны), но убираем только YEAR_SPOD_TOP_Month и "debug"
        if class_name and func_name:
            # Убираем YEAR_SPOD_TOP_Month из class_name, если есть
            clean_class = class_name.replace("YEAR_SPOD_TOP_Month", "").strip()
            # Убираем "debug" из func_name, если есть (но оставляем остальное)
            clean_func = func_name.replace("debug", "").strip() if func_name and func_name == "debug" else func_name
            if clean_class and clean_func:
                formatted_message = f"{masked_message} [class: {clean_class} | def: {clean_func}]"
            elif clean_class:
                formatted_message = f"{masked_message} [class: {clean_class}]"
            elif clean_func:
                formatted_message = f"{masked_message} [def: {clean_func}]"
            else:
                formatted_message = masked_message
            self.logger.debug(formatted_message)
        else:
            self.logger.debug(masked_message)
    
    def warning(self, message: str, class_name: Optional[str] = None, func_name: Optional[str] = None) -> None:
        """
        Логирует сообщение уровня WARNING.
        
        Args:
            message: Сообщение для логирования
            class_name: Имя класса (опционально)
            func_name: Имя функции (опционально)
        """
        # Маскируем чувствительные данные (табельные номера и ИД клиентов)
        masked_message = self._mask_sensitive_data(message)
        if class_name and func_name:
            clean_class = class_name.replace("YEAR_SPOD_TOP_Month", "").strip()
            if clean_class:
                formatted_message = f"{masked_message} [class: {clean_class} | def: {func_name}]"
            else:
                formatted_message = f"{masked_message} [def: {func_name}]"
            self.logger.warning(formatted_message)
        else:
            self.logger.warning(masked_message)
    
    def error(self, message: str, class_name: Optional[str] = None, func_name: Optional[str] = None) -> None:
        """
        Логирует сообщение уровня ERROR.
        
        Args:
            message: Сообщение для логирования
            class_name: Имя класса (опционально)
            func_name: Имя функции (опционально)
        """
        # Маскируем чувствительные данные (табельные номера и ИД клиентов)
        masked_message = self._mask_sensitive_data(message)
        if class_name and func_name:
            clean_class = class_name.replace("YEAR_SPOD_TOP_Month", "").strip()
            if clean_class:
                formatted_message = f"{masked_message} [class: {clean_class} | def: {func_name}]"
            else:
                formatted_message = f"{masked_message} [def: {func_name}]"
            self.logger.error(formatted_message)
        else:
            self.logger.error(masked_message)


# ============================================================================
# МОДУЛЬ ОБРАБОТКИ ФАЙЛОВ
# ============================================================================

# ============================================================================
# КЛАСС ДЛЯ СБОРА ДЕТАЛЬНОЙ СТАТИСТИКИ ПО ТАБЕЛЬНЫМ НОМЕРАМ
# ============================================================================

class DebugTabNumberTracker:
    """Класс для сбора детальной статистики по табельным номерам из DEBUG_TAB_NUMBER."""
    
    def __init__(self, logger_instance: Optional[Logger] = None):
        """
        Инициализация трекера.
        
        Args:
            logger_instance: Экземпляр логгера
        """
        self.logger = logger_instance
        # Структура данных для каждого табельного номера:
        # {
        #   "tab_number": {
        #     "source_files": {  # Данные из исходных файлов
        #       "file_name": {
        #         "group": str,
        #         "month": int,
        #         "clients": [  # Список клиентов (ИНН) с их данными
        #           {
        #             "ИНН": str,
        #             "ТБ": str,
        #             "ФИО": str,
        #             "Показатель": float,
        #             "Выбран": bool  # Был ли выбран этот вариант
        #           }
        #         ],
        #         "tb_variants": {  # Варианты ТБ с суммами
        #           "ТБ": float  # Сумма показателя для этого ТБ
        #         },
        #         "selected_tb": str,  # Выбранный ТБ
        #         "selected_sum": float  # Сумма выбранного варианта
        #       }
        #     },
        #     "raw_data": {  # Данные после схлопывания (RAW)
        #       "ИНН": {
        #         "ТБ": str,
        #         "ФИО": str,
        #         "sums_by_file": {  # Суммы по файлам
        #           "file_name": float
        #         }
        #       }
        #     },
        #     "calculations": {  # Результаты расчетов
        #       "month": {
        #         "fact": float,
        #         "growth_2m": float,
        #         "growth_3m": float
        #       }
        #     },
        #     "normalization": {  # Нормализованные значения
        #       "month": {
        #         "indicator": float
        #       }
        #     },
        #     "scores": {  # Score по месяцам
        #       "month": float
        #     },
        #     "best_month": str,  # Лучший месяц
        #     "unique_inn_count": int  # Количество уникальных ИНН
        #   }
        # }
        self.tab_data: Dict[str, Dict[str, Any]] = {}
        
        # Инициализируем структуру для каждого табельного номера из DEBUG_TAB_NUMBER
        # ВАЖНО: Нормализуем табельные номера используя ту же логику, что и _normalize_tab_number
        # (удаляем лидирующие нули, затем добавляем до 8 знаков)
        if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0:
            for tab_num in DEBUG_TAB_NUMBER:
                # Используем ту же логику нормализации, что и _normalize_tab_number
                tab_num_str = str(tab_num).strip()
                if not tab_num_str or tab_num_str.lower() == 'nan':
                    continue
                # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
                tab_num_clean = tab_num_str.lstrip('0') if tab_num_str.lstrip('0') else '0'
                tab_num_normalized = tab_num_clean.zfill(8)
                
                self.tab_data[tab_num_normalized] = {
                    "source_files": {},
                    "raw_data": {},
                    "calculations": {},
                    "normalization": {},
                    "scores": {},
                    "best_month": None,
                    "unique_inn_count": 0
                }
                # Также сохраняем оригинальный номер (без нормализации) для обратной совместимости
                if tab_num_normalized != tab_num_str:
                    self.tab_data[tab_num_str] = self.tab_data[tab_num_normalized]
    
    def add_source_file_data(self, tab_number: str, file_name: str, group: str, month: int,
                             clients_data: List[Dict[str, Any]], tb_variants: Dict[str, float],
                             selected_tb: str, selected_sum: float) -> None:
        """
        Добавляет данные из исходного файла для табельного номера.
        
        Args:
            tab_number: Табельный номер
            file_name: Имя файла
            group: Группа (OD, RA, PS)
            month: Номер месяца
            clients_data: Список клиентов с их данными
            tb_variants: Варианты ТБ с суммами
            selected_tb: Выбранный ТБ
            selected_sum: Сумма выбранного варианта
        """
        # Нормализуем табельный номер для поиска в трекере (используем ту же логику, что и _normalize_tab_number)
        tab_number_str = str(tab_number).strip()
        if not tab_number_str or tab_number_str.lower() == 'nan':
            return
        # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
        tab_number_clean = tab_number_str.lstrip('0') if tab_number_str.lstrip('0') else '0'
        tab_number_normalized = tab_number_clean.zfill(8)
        
        if tab_number_normalized not in self.tab_data:
            # Пробуем найти без нормализации (оригинальный формат)
            if tab_number_str not in self.tab_data:
                return
            tab_number_normalized = tab_number_str
        
        self.tab_data[tab_number_normalized]["source_files"][file_name] = {
            "group": group,
            "month": month,
            "clients": clients_data,
            "tb_variants": tb_variants,
            "selected_tb": selected_tb,
            "selected_sum": selected_sum
        }
    
    def add_raw_data(self, tab_number: str, raw_data: Dict[str, Dict[str, Any]]) -> None:
        """
        Добавляет данные после схлопывания (RAW) для табельного номера.
        
        Args:
            tab_number: Табельный номер
            raw_data: Данные по ИНН после схлопывания
        """
        # Нормализуем табельный номер для поиска в трекере (используем ту же логику, что и _normalize_tab_number)
        tab_number_str = str(tab_number).strip()
        if not tab_number_str or tab_number_str.lower() == 'nan':
            return
        # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
        tab_number_clean = tab_number_str.lstrip('0') if tab_number_str.lstrip('0') else '0'
        tab_number_normalized = tab_number_clean.zfill(8)
        
        if tab_number_normalized not in self.tab_data:
            # Пробуем найти без нормализации (оригинальный формат)
            if tab_number_str not in self.tab_data:
                return
            tab_number_normalized = tab_number_str
        
        self.tab_data[tab_number_normalized]["raw_data"] = raw_data
    
    def add_calculations(self, tab_number: str, calculations: Dict[str, Dict[str, float]]) -> None:
        """
        Добавляет результаты расчетов для табельного номера.
        
        Args:
            tab_number: Табельный номер
            calculations: Результаты расчетов по месяцам
        """
        # Нормализуем табельный номер для поиска в трекере (используем ту же логику, что и _normalize_tab_number)
        tab_number_str = str(tab_number).strip()
        if not tab_number_str or tab_number_str.lower() == 'nan':
            return
        # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
        tab_number_clean = tab_number_str.lstrip('0') if tab_number_str.lstrip('0') else '0'
        tab_number_normalized = tab_number_clean.zfill(8)
        
        if tab_number_normalized not in self.tab_data:
            # Пробуем найти без нормализации (оригинальный формат)
            if tab_number_str not in self.tab_data:
                return
            tab_number_normalized = tab_number_str
        
        self.tab_data[tab_number_normalized]["calculations"] = calculations
    
    def add_normalization(self, tab_number: str, normalization: Dict[str, Dict[str, float]]) -> None:
        """
        Добавляет нормализованные значения для табельного номера.
        
        Args:
            tab_number: Табельный номер
            normalization: Нормализованные значения по месяцам
        """
        # Нормализуем табельный номер для поиска в трекере (используем ту же логику, что и _normalize_tab_number)
        tab_number_str = str(tab_number).strip()
        if not tab_number_str or tab_number_str.lower() == 'nan':
            return
        # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
        tab_number_clean = tab_number_str.lstrip('0') if tab_number_str.lstrip('0') else '0'
        tab_number_normalized = tab_number_clean.zfill(8)
        
        if tab_number_normalized not in self.tab_data:
            # Пробуем найти без нормализации (оригинальный формат)
            if tab_number_str not in self.tab_data:
                return
            tab_number_normalized = tab_number_str
        
        self.tab_data[tab_number_normalized]["normalization"] = normalization
    
    def add_scores(self, tab_number: str, scores: Dict[str, float], best_month: str) -> None:
        """
        Добавляет Score и лучший месяц для табельного номера.
        
        Args:
            tab_number: Табельный номер
            scores: Score по месяцам
            best_month: Лучший месяц
        """
        # Нормализуем табельный номер для поиска в трекере (используем ту же логику, что и _normalize_tab_number)
        tab_number_str = str(tab_number).strip()
        if not tab_number_str or tab_number_str.lower() == 'nan':
            return
        # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
        tab_number_clean = tab_number_str.lstrip('0') if tab_number_str.lstrip('0') else '0'
        tab_number_normalized = tab_number_clean.zfill(8)
        
        if tab_number_normalized not in self.tab_data:
            # Пробуем найти без нормализации (оригинальный формат)
            if tab_number_str not in self.tab_data:
                return
            tab_number_normalized = tab_number_str
        
        self.tab_data[tab_number_normalized]["scores"] = scores
        self.tab_data[tab_number_normalized]["best_month"] = best_month
    
    def set_unique_inn_count(self, tab_number: str, count: int) -> None:
        """
        Устанавливает количество уникальных ИНН для табельного номера.
        
        Args:
            tab_number: Табельный номер
            count: Количество уникальных ИНН
        """
        # Нормализуем табельный номер для поиска в трекере (используем ту же логику, что и _normalize_tab_number)
        tab_number_str = str(tab_number).strip()
        if not tab_number_str or tab_number_str.lower() == 'nan':
            return
        # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
        tab_number_clean = tab_number_str.lstrip('0') if tab_number_str.lstrip('0') else '0'
        tab_number_normalized = tab_number_clean.zfill(8)
        
        if tab_number_normalized not in self.tab_data:
            # Пробуем найти без нормализации (оригинальный формат)
            if tab_number_str not in self.tab_data:
                return
            tab_number_normalized = tab_number_str
        
        self.tab_data[tab_number_normalized]["unique_inn_count"] = count
    
    def get_tab_data(self, tab_number: str) -> Optional[Dict[str, Any]]:
        """
        Получает собранные данные для табельного номера.
        
        Args:
            tab_number: Табельный номер
            
        Returns:
            Dict с данными или None, если табельный номер не отслеживается
        """
        # Нормализуем табельный номер для поиска в трекере (используем ту же логику, что и _normalize_tab_number)
        tab_number_str = str(tab_number).strip()
        if not tab_number_str or tab_number_str.lower() == 'nan':
            return None
        # Удаляем лидирующие нули для корректной нормализации (как в _normalize_tab_number)
        tab_number_clean = tab_number_str.lstrip('0') if tab_number_str.lstrip('0') else '0'
        tab_number_normalized = tab_number_clean.zfill(8)
        
        if tab_number_normalized in self.tab_data:
            return self.tab_data[tab_number_normalized]
        # Пробуем найти без нормализации (оригинальный формат)
        if tab_number_str in self.tab_data:
            return self.tab_data[tab_number_str]
        return None
    
    def get_all_tab_numbers(self) -> List[str]:
        """Возвращает список всех отслеживаемых табельных номеров."""
        return list(self.tab_data.keys())


class FileProcessor:
    """Класс для обработки Excel файлов."""
    
    def _create_debug_tab_mask(self, df: pd.DataFrame, tab_column: str) -> pd.Series:
        """
        Создает маску для детального логирования табельных номеров из списка DEBUG_TAB_NUMBER.
        
        Args:
            df: DataFrame для создания маски
            tab_column: Название колонки с табельными номерами
            
        Returns:
            pd.Series: Булева маска (True для строк с табельными номерами из DEBUG_TAB_NUMBER)
        """
        if DEBUG_TAB_NUMBER is None or len(DEBUG_TAB_NUMBER) == 0 or tab_column not in df.columns:
            return pd.Series([False] * len(df), index=df.index)
        
        debug_mask = pd.Series([False] * len(df), index=df.index)
        for debug_tab in DEBUG_TAB_NUMBER:
            if debug_tab is None:
                continue
            debug_tab_str = str(debug_tab).strip().lstrip('0')
            tab_mask = df[tab_column].astype(str).str.strip().str.lstrip('0') == debug_tab_str
            debug_mask = debug_mask | tab_mask
        
        return debug_mask
    
    def __init__(self, input_dir: str = INPUT_DIR, logger_instance: Optional[Logger] = None):
        """
        Инициализация процессора файлов.
        
        Args:
            input_dir: Директория с входными файлами
            logger_instance: Экземпляр логгера
        
        Args:
            input_dir: Путь к каталогу с входными данными
            logger_instance: Экземпляр логгера
        """
        self.input_dir = Path(input_dir)
        self.groups = ["OD", "RA", "PS"]
        self.processed_files: Dict[str, Dict[str, pd.DataFrame]] = {}
        self.unique_tab_numbers: Dict[str, Dict[str, Any]] = {}
        self.logger = logger_instance
        
        # Инициализируем трекер для детальной статистики по табельным номерам
        self.debug_tracker = DebugTabNumberTracker(logger_instance=logger_instance)
        
        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Информируем о включенном режиме детального логирования
        if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0:
            tab_numbers_str = ", ".join(DEBUG_TAB_NUMBER)
            self.logger.info(
                f"Включено детальное логирование для табельных номеров: {tab_numbers_str}. "
                f"Все операции с этими табельными номерами будут подробно логироваться.",
                "FileProcessor",
                "__init__"
            )
        
        # Статистика обработки (собирается только если ENABLE_STATISTICS = True)
        self.statistics = {
            "files": {},  # Статистика по файлам: {group: {file_name: {initial_rows, dropped_by_rule, kept_by_rule, final_rows}}}
            "tab_selection": {},  # Статистика выбора табельных: {group: {file_name: {total_variants, selected_count}}}
            "summary": {}  # Итоговая статистика: {total_km, total_clients, by_tb: {tb: count}}
        }
    
    def load_all_files(self) -> None:
        """
        Загружает все файлы из подкаталогов OD, RA, PS.
        
        Файлы загружаются с учетом конфигурации для каждой группы.
        Используются только файлы из списка expected_files.
        
        ОПТИМИЗАЦИЯ: Все группы (OD, RA, PS) загружаются параллельно.
        """
        self.logger.info("Начало загрузки файлов", "FileProcessor", "load_all_files")
        
        # ОПТИМИЗАЦИЯ: Параллельная загрузка всех групп
        # Загружаем все группы параллельно: OD, RA, PS одновременно
        self.logger.debug(f"Параллельная загрузка всех групп: {', '.join(self.groups)} (max_workers={MAX_WORKERS})", "FileProcessor", "load_all_files")
        
        # Для сводной статистики
        total_rows = 0
        all_client_ids = set()
        all_tab_numbers = set()
        
        # Инициализируем словарь для обработанных файлов
        self.processed_files = {}
        
        # Загружаем все группы параллельно
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            # Создаем задачи для загрузки всех групп
            future_to_group = {
                executor.submit(self._load_group_files, group): group
                for group in self.groups
            }
            
            # Обрабатываем результаты по мере завершения
            for future in as_completed(future_to_group):
                group = future_to_group[future]
                try:
                    result = future.result()
                    group_name = result['group']
                    group_files = result['files']
                    group_stats = result['stats']
                    
                    # Сохраняем загруженные файлы
                    self.processed_files[group_name] = group_files
                    
                    # Собираем статистику
                    total_rows += group_stats['rows']
                    all_client_ids.update(group_stats['clients'])
                    all_tab_numbers.update(group_stats['tabs'])
                    
                except Exception as e:
                    self.logger.error(f"Ошибка при загрузке группы {group}: {str(e)}", "FileProcessor", "load_all_files")
        
        # Сводная статистика (INFO)
        stats_parts = [f"{total_rows} строк"]
        if len(all_client_ids) > 0:
            stats_parts.append(f"{len(all_client_ids)} уникальных клиентов (ИНН)")
        if len(all_tab_numbers) > 0:
            stats_parts.append(f"{len(all_tab_numbers)} уникальных табельных номеров")
        
        # Сохраняем статистику по клиентам
        if ENABLE_STATISTICS:
            self.statistics["summary"]["total_clients"] = len(all_client_ids)
        
        self.logger.info(f"Загрузка завершена. Обработано групп: {len(self.processed_files)}. Итого: {', '.join(stats_parts)}", "FileProcessor", "load_all_files")
    
    def _load_group_files(self, group: str) -> Dict[str, Any]:
        """
        Загружает все файлы одной группы.
        
        Args:
            group: Название группы (OD, RA, PS)
        
        Returns:
            Словарь с результатами загрузки: {
                'group': group,
                'files': {file_name: df},
                'stats': {'rows': int, 'clients': set, 'tabs': set}
            }
        """
        group_path = self.input_dir / group
        if not group_path.exists():
            self.logger.warning(f"Каталог {group_path} не найден, пропускаем", "FileProcessor", "_load_group_files")
            return {'group': group, 'files': {}, 'stats': {'rows': 0, 'clients': set(), 'tabs': set()}}
        
        self.logger.info(f"Обработка группы {group}", "FileProcessor", "_load_group_files")
        group_files = {}
        
        # Получаем конфигурацию группы
        group_config = config_manager.get_group_config(group)
        items = group_config.items
        defaults = group_config.defaults
        
        if not items:
            self.logger.warning(f"Список файлов (items) пуст для группы {group}", "FileProcessor", "_load_group_files")
            return {'group': group, 'files': {}, 'stats': {'rows': 0, 'clients': set(), 'tabs': set()}}
        
        self.logger.debug(f"Ожидается {len(items)} файлов в группе {group}", "FileProcessor", "_load_group_files")
        
        # ОПТИМИЗАЦИЯ: Параллельная загрузка файлов
        # Подготавливаем список файлов для загрузки
        files_to_load = []
        for item in items:
            if not item.file_name or item.file_name.strip() == "":
                continue
            file_path = group_path / item.file_name
            if file_path.exists():
                files_to_load.append((file_path, item, group, defaults))
        
        if not files_to_load:
            return {'group': group, 'files': {}, 'stats': {'rows': 0, 'clients': set(), 'tabs': set()}}
        
        # Статистика по группе
        total_rows = 0
        all_client_ids = set()
        all_tab_numbers = set()
        
        # Выбираем метод загрузки: параллельный или последовательный
        if ENABLE_PARALLEL_LOADING and len(files_to_load) > 1:
            self.logger.debug(f"Параллельная загрузка {len(files_to_load)} файлов группы {group} (max_workers={MAX_WORKERS})", "FileProcessor", "_load_group_files")
            
            # Загружаем файлы параллельно
            # ВАЖНО: Используем ThreadPoolExecutor для I/O операций (чтение Excel файлов)
            # pandas.read_excel может блокироваться на уровне GIL, но ThreadPoolExecutor должен справиться
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                # Создаем задачи для загрузки - все файлы отправляются в очередь одновременно
                future_to_file = {
                    executor.submit(self._load_file, file_path, group): (file_path, item, defaults)
                    for file_path, item, group, defaults in files_to_load
                }
                
                self.logger.debug(f"Отправлено {len(future_to_file)} задач на параллельную загрузку файлов группы {group}", "FileProcessor", "_load_group_files")
                
                # Обрабатываем результаты по мере завершения (as_completed гарантирует обработку по готовности)
                completed_count = 0
                for future in as_completed(future_to_file):
                    file_path, item, defaults = future_to_file[future]
                    completed_count += 1
                    self.logger.debug(f"Завершена загрузка файла {file_path.name} ({completed_count}/{len(future_to_file)})", "FileProcessor", "_load_group_files")
                    try:
                        df = future.result()
                        if df is not None and not df.empty:
                            # ВАЖНО: Запись в словарь происходит последовательно, но это быстро
                            group_files[file_path.name] = df
                            
                            # Статистика по файлу
                            rows_count = len(df)
                            total_rows += rows_count
                            
                            tab_number_col = defaults.tab_number_column
                            client_id_col = "client_id"
                            
                            unique_clients = 0
                            unique_tabs = 0
                            
                            if client_id_col in df.columns:
                                unique_clients = df[client_id_col].nunique()
                                if len(all_client_ids) < 10000:
                                    valid_client_ids = df[client_id_col].dropna().astype(str).str.strip()
                                    valid_client_ids = valid_client_ids[(valid_client_ids != 'nan') & (valid_client_ids != '')]
                                    all_client_ids.update(valid_client_ids.unique())
                            
                            if tab_number_col in df.columns:
                                unique_tabs = df[tab_number_col].nunique()
                                if len(all_tab_numbers) < 10000:
                                    valid_tabs = df[tab_number_col].dropna().astype(str).str.strip()
                                    valid_tabs = valid_tabs[(valid_tabs != 'nan') & (valid_tabs != '')]
                                    all_tab_numbers.update(valid_tabs.unique())
                            
                            # Логируем статистику по файлу (INFO)
                            stats_parts = [f"{rows_count} строк"]
                            if unique_clients > 0:
                                stats_parts.append(f"{unique_clients} уникальных клиентов (ИНН)")
                            if unique_tabs > 0:
                                stats_parts.append(f"{unique_tabs} уникальных табельных номеров")
                            
                            stats_message = f"Загружен файл {file_path.name} ({item.label}): {', '.join(stats_parts)}"
                            self.logger.info(stats_message, "FileProcessor", "_load_group_files")
                        else:
                            self.logger.warning(f"Файл {file_path.name} ({item.label}) загружен, но пуст", "FileProcessor", "_load_group_files")
                    except Exception as e:
                        self.logger.error(f"Ошибка при загрузке файла {file_path.name}: {str(e)}", "FileProcessor", "_load_group_files")
        else:
            # Последовательная загрузка (старый метод)
            for file_path, item, group, defaults in files_to_load:
                try:
                    df = self._load_file(file_path, group)
                    if df is not None and not df.empty:
                        group_files[file_path.name] = df
                        
                        # Статистика по файлу
                        rows_count = len(df)
                        total_rows += rows_count
                        
                        tab_number_col = defaults.tab_number_column
                        client_id_col = "client_id"
                        
                        unique_clients = 0
                        unique_tabs = 0
                        
                        if client_id_col in df.columns:
                            unique_clients = df[client_id_col].nunique()
                            if len(all_client_ids) < 10000:
                                valid_client_ids = df[client_id_col].dropna().astype(str).str.strip()
                                valid_client_ids = valid_client_ids[(valid_client_ids != 'nan') & (valid_client_ids != '')]
                                all_client_ids.update(valid_client_ids.unique())
                        
                        if tab_number_col in df.columns:
                            unique_tabs = df[tab_number_col].nunique()
                            if len(all_tab_numbers) < 10000:
                                valid_tabs = df[tab_number_col].dropna().astype(str).str.strip()
                                valid_tabs = valid_tabs[(valid_tabs != 'nan') & (valid_tabs != '')]
                                all_tab_numbers.update(valid_tabs.unique())
                        
                        # Логируем статистику по файлу (INFO)
                        stats_parts = [f"{rows_count} строк"]
                        if unique_clients > 0:
                            stats_parts.append(f"{unique_clients} уникальных клиентов (ИНН)")
                        if unique_tabs > 0:
                            stats_parts.append(f"{unique_tabs} уникальных табельных номеров")
                        
                        stats_message = f"Загружен файл {file_path.name} ({item.label}): {', '.join(stats_parts)}"
                        self.logger.info(stats_message, "FileProcessor", "_load_group_files")
                    else:
                        self.logger.warning(f"Файл {file_path.name} ({item.label}) загружен, но пуст", "FileProcessor", "_load_group_files")
                except Exception as e:
                    self.logger.error(f"Ошибка при загрузке файла {file_path.name} ({item.label}): {str(e)}", "FileProcessor", "_load_group_files")
        
        return {
            'group': group,
            'files': group_files,
            'stats': {'rows': total_rows, 'clients': all_client_ids, 'tabs': all_tab_numbers}
        }
    
    def _normalize_tab_number(self, value: Any, length: int, fill_char: str) -> str:
        """
        Нормализует табельный номер: преобразует в строку заданной длины с лидирующими нулями.
        
        Args:
            value: Значение табельного номера
            length: Длина строки
            fill_char: Символ для заполнения
            
        Returns:
            str: Нормализованный табельный номер
        """
        if pd.isna(value):
            return ""
        value_str = str(value).strip()
        if not value_str or value_str.lower() == 'nan':
            return ""
        # Удаляем лидирующие нули для корректной нормализации
        value_clean = value_str.lstrip('0') if value_str.lstrip('0') else '0'
        return value_clean.zfill(length)
    
    def _normalize_inn(self, value: Any, length: int, fill_char: str) -> str:
        """
        Нормализует ИНН: преобразует в строку заданной длины с лидирующими нулями.
        
        Args:
            value: Значение ИНН
            length: Длина строки
            fill_char: Символ для заполнения
            
        Returns:
            str: Нормализованный ИНН
        """
        if pd.isna(value):
            return ""
        value_str = str(value).strip()
        if not value_str or value_str.lower() == 'nan':
            return ""
        # Удаляем лидирующие нули для корректной нормализации
        value_clean = value_str.lstrip('0') if value_str.lstrip('0') else '0'
        return value_clean.zfill(length)
    
    def _load_file(self, file_path: Path, group_name: str) -> Optional[pd.DataFrame]:
        """
        Загружает один файл с применением конфигурации.
        
        Args:
            file_path: Путь к файлу
            group_name: Название группы
            
        Returns:
            Optional[pd.DataFrame]: DataFrame с данными или None при ошибке
        """
        try:
            # Получаем конфигурацию для файла
            config = config_manager.get_config_for_file(group_name, file_path.name)
            
            self.logger.debug(f"Загрузка файла {file_path.name} с конфигурацией: {config}", "FileProcessor", "_load_file")
            
            # Подготавливаем параметры для чтения Excel
            read_params = {}
            
            # Определяем engine
            if OPENPYXL_AVAILABLE:
                read_params['engine'] = 'openpyxl'
            
            # Параметры листа
            # Используем sheet_name из конфигурации (может быть из item.sheet или default_sheet)
            if config["sheet_name"]:
                read_params['sheet_name'] = config["sheet_name"]
            elif config["sheet_index"] is not None:
                read_params['sheet_name'] = config["sheet_index"]
            
            # Параметры пропуска строк
            if config["skip_rows"] > 0:
                read_params['skiprows'] = config["skip_rows"]
            
            if config["skip_footer"] > 0:
                read_params['skipfooter'] = config["skip_footer"]
            
            # Параметр заголовка
            if config["header_row"] is not None:
                read_params['header'] = config["header_row"]
            
            # ОПТИМИЗАЦИЯ: Определяем usecols для ускорения загрузки (если известны колонки)
            # Это позволяет загружать только нужные колонки, что значительно ускоряет загрузку больших файлов
            if config["columns"]:
                source_columns = [col["source"] for col in config["columns"]]
                read_params['usecols'] = source_columns
            
            # ОПТИМИЗАЦИЯ: Chunking для больших файлов
            # ВАЖНО: Chunking через openpyxl очень медленный, поэтому отключен по умолчанию
            # Используется только для очень больших файлов (>200 МБ)
            df = None
            if ENABLE_CHUNKING and OPENPYXL_AVAILABLE:
                # Проверяем размер файла (приблизительно)
                file_size_mb = file_path.stat().st_size / (1024 * 1024)
                # Если файл больше порога, используем chunking
                if file_size_mb > CHUNKING_THRESHOLD_MB:
                    self.logger.debug(f"Файл {file_path.name} очень большой ({file_size_mb:.1f} МБ), используем chunking", "FileProcessor", "_load_file")
                    try:
                        df = self._load_file_with_chunking(file_path, config, read_params)
                    except Exception as chunk_error:
                        self.logger.warning(f"Ошибка при chunking файла {file_path.name}, используем обычную загрузку: {str(chunk_error)}", "FileProcessor", "_load_file")
                        df = None  # Продолжим с обычной загрузкой
            
            # Обычная загрузка (если chunking не использовался или не сработал)
            if df is None:
                # ДИАГНОСТИКА: Логируем начало загрузки файла для проверки параллельности
                self.logger.debug(f"Начало загрузки файла {file_path.name} (группа {group_name})", "FileProcessor", "_load_file")
                try:
                    df = pd.read_excel(file_path, **read_params)
                    self.logger.debug(f"Завершена загрузка файла {file_path.name} (группа {group_name}): {len(df)} строк", "FileProcessor", "_load_file")
                except Exception as e:
                    # Если не удалось загрузить с параметрами, пробуем без usecols
                    self.logger.warning(f"Ошибка при загрузке с параметрами, пробуем без usecols: {str(e)}", "FileProcessor", "_load_file")
                    try:
                        read_params_fallback = {k: v for k, v in read_params.items() if k != 'usecols'}
                        df = pd.read_excel(file_path, **read_params_fallback)
                        
                        # Фильтруем колонки после загрузки
                        if config["columns"]:
                            source_columns = [col["source"] for col in config["columns"]]
                            available_columns = [col for col in source_columns if col in df.columns]
                            if available_columns:
                                df = df[available_columns]
                    except Exception as e2:
                        # Если все еще не получилось, пробуем без всех параметров
                        self.logger.warning(f"Ошибка при загрузке, пробуем без параметров: {str(e2)}", "FileProcessor", "_load_file")
                        try:
                            df = pd.read_excel(file_path)
                            # Фильтруем колонки после загрузки
                            if config["columns"]:
                                source_columns = [col["source"] for col in config["columns"]]
                                available_columns = [col for col in source_columns if col in df.columns]
                                if available_columns:
                                    df = df[available_columns]
                        except Exception as e3:
                            self.logger.error(f"Не удалось загрузить файл {file_path.name}: {str(e3)}", "FileProcessor", "_load_file")
                            return None
            
            # Собираем статистику: исходное количество строк
            if ENABLE_STATISTICS:
                initial_rows = len(df)
                if group_name not in self.statistics["files"]:
                    self.statistics["files"][group_name] = {}
                if file_path.name not in self.statistics["files"][group_name]:
                    self.statistics["files"][group_name][file_path.name] = {
                        "initial_rows": initial_rows,
                        "dropped_by_rule": {},
                        "kept_by_rule": {},
                        "final_rows": initial_rows
                    }
                else:
                    self.statistics["files"][group_name][file_path.name]["initial_rows"] = initial_rows
                    self.statistics["files"][group_name][file_path.name]["final_rows"] = initial_rows
            
            # Нормализуем названия колонок (убираем пробелы)
            df.columns = df.columns.str.strip()
            
            # Применяем маппинг колонок (source -> alias)
            if config["columns"]:
                # Формируем словарь маппинга: source -> alias
                column_maps = {col["source"]: col["alias"] for col in config["columns"]}
                
                # Проверяем наличие всех source колонок
                missing_columns = [col["source"] for col in config["columns"] if col["source"] not in df.columns]
                if missing_columns:
                    self.logger.warning(f"Отсутствующие колонки в файле {file_path.name}: {missing_columns}", "FileProcessor", "_load_file")
                
                # Переименовываем колонки из source в alias
                available_maps = {k: v for k, v in column_maps.items() if k in df.columns}
                df = df.rename(columns=available_maps)
                
                # Оставляем только нужные колонки (по alias)
                required_columns = [col["alias"] for col in config["columns"]]
                available_columns = [col for col in required_columns if col in df.columns]
                df = df[available_columns]
            
            # Применяем правила удаления строк (drop_rules)
            if config["drop_rules"]:
                df = self._apply_drop_rules(df, config["drop_rules"], file_path.name, group_name)
            
            # Применяем правила включения строк (in_rules)
            if config["in_rules"]:
                df = self._apply_in_rules(df, config["in_rules"], file_path.name, group_name)
            
            # Обновляем статистику: финальное количество строк
            if ENABLE_STATISTICS:
                if group_name in self.statistics["files"] and file_path.name in self.statistics["files"][group_name]:
                    self.statistics["files"][group_name][file_path.name]["final_rows"] = len(df)
            
            # Нормализуем табельные номера и ИНН
            group_config = config_manager.get_group_config(group_name)
            defaults = group_config.defaults
            
            # ОПТИМИЗАЦИЯ: Векторизованная нормализация табельных номеров и ИНН
            # Используем векторизованные операции вместо apply() для ускорения
            tab_number_col = defaults.tab_number_column
            if tab_number_col in df.columns:
                # Преобразуем в строку и очищаем
                df[tab_number_col] = df[tab_number_col].astype(str).str.strip()
                # Заменяем NaN и пустые значения
                mask_nan = (df[tab_number_col] == 'nan') | (df[tab_number_col] == 'None') | (df[tab_number_col] == '')
                df.loc[mask_nan, tab_number_col] = ''
                # Нормализуем: удаляем лидирующие нули и заполняем до нужной длины
                df[tab_number_col] = df[tab_number_col].apply(
                    lambda x: x.lstrip('0').zfill(defaults.tab_number_length) if x and x.lstrip('0') else ('0' * defaults.tab_number_length)
                )
            
            # Нормализация ИНН
            client_id_col = "client_id"
            if client_id_col in df.columns:
                # Преобразуем в строку и очищаем
                df[client_id_col] = df[client_id_col].astype(str).str.strip()
                # Заменяем NaN и пустые значения
                mask_nan = (df[client_id_col] == 'nan') | (df[client_id_col] == 'None') | (df[client_id_col] == '')
                df.loc[mask_nan, client_id_col] = ''
                # Нормализуем: удаляем лидирующие нули и заполняем до нужной длины
                df[client_id_col] = df[client_id_col].apply(
                    lambda x: x.lstrip('0').zfill(defaults.inn_length) if x and x.lstrip('0') else ('0' * defaults.inn_length)
                )
            
            # Нормализация ТБ (территориального банка)
            tb_col = defaults.tb_column
            if tb_col in df.columns:
                # ОПТИМИЗАЦИЯ: Векторизованная нормализация ТБ через функцию normalize_tb_value
                # Применяем нормализацию к каждому значению ТБ
                df[tb_col] = df[tb_col].apply(normalize_tb_value)
                # Заменяем None на пустую строку для единообразия
                df[tb_col] = df[tb_col].fillna('')
            
            # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Если указан DEBUG_TAB_NUMBER, логируем данные по этим табельным
            if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0 and tab_number_col in df.columns:
                debug_mask = self._create_debug_tab_mask(df, tab_number_col)
                debug_rows = df[debug_mask]
                if len(debug_rows) > 0:
                    indicator_col = defaults.indicator_column
                    client_id_col = "client_id"
                    
                    for idx, row in debug_rows.iterrows():
                        client_id = row.get(client_id_col, '')
                        tb_value = row.get(tb_col, '')
                        gosb_value = row.get(defaults.gosb_column, '')
                        fio_value = row.get(defaults.fio_column, '')
                        indicator_value = row.get(indicator_col, 0)
                        
                        # ФИО и табельный номер уже будут замаскированы в _mask_sensitive_data при логировании
                        tab_number_value = str(row.get(tab_number_col, ''))
                        self.logger.debug_tab(
                            f"Загрузка файла {file_path.name} (группа {group_name}): найдена строка для ТН. "
                            f"Табельный: {tab_number_value}, Клиент: {client_id}, ТБ: {tb_value}, ФИО: {fio_value}, "
                            f"Показатель ({indicator_col}): {indicator_value}",
                            tab_number=row.get(tab_number_col),
                            class_name="FileProcessor",
                            func_name="_load_file"
                        )
            
            # Добавляем метаданные о файле
            df.attrs['file_name'] = file_path.name
            df.attrs['group_name'] = group_name
            df.attrs['file_path'] = str(file_path)
            
            return df
            
        except Exception as e:
            self.logger.error(f"Ошибка при обработке файла {file_path}: {str(e)}", "FileProcessor", "_load_file")
            return None
    
    def _load_file_with_chunking(self, file_path: Path, config: Dict[str, Any], read_params: Dict[str, Any]) -> pd.DataFrame:
        """
        Загружает большой Excel файл по частям (chunking) для оптимизации памяти и производительности.
        
        Args:
            file_path: Путь к файлу
            config: Конфигурация для файла
            read_params: Параметры для чтения Excel
            
        Returns:
            pd.DataFrame: DataFrame с данными
        """
        try:
            if not OPENPYXL_AVAILABLE:
                # Если openpyxl недоступен, используем обычную загрузку
                return pd.read_excel(file_path, **read_params)
            
            from openpyxl import load_workbook
            
            # Загружаем рабочую книгу
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # Определяем лист для чтения
            sheet_name = read_params.get('sheet_name', wb.sheetnames[0] if wb.sheetnames else None)
            if sheet_name is None:
                sheet_name = wb.sheetnames[0] if wb.sheetnames else None
            
            if sheet_name is None:
                wb.close()
                return pd.DataFrame()
            
            ws = wb[sheet_name]
            
            # Определяем заголовки
            header_row = read_params.get('header', 0)
            if isinstance(header_row, int):
                headers = []
                for cell in ws[header_row + 1]:
                    headers.append(cell.value if cell.value else f"Column_{len(headers)}")
            else:
                headers = None
            
            # Определяем usecols
            usecols = read_params.get('usecols', None)
            if usecols and headers:
                # Фильтруем заголовки по usecols
                header_indices = []
                header_names = []
                for idx, header in enumerate(headers):
                    if header in usecols:
                        header_indices.append(idx)
                        header_names.append(header)
                headers = header_names
            elif headers:
                header_indices = list(range(len(headers)))
            else:
                header_indices = None
            
            # Читаем данные по частям
            chunks = []
            start_row = header_row + 1 + read_params.get('skiprows', 0)
            end_row = ws.max_row - read_params.get('skipfooter', 0)
            
            self.logger.debug(f"Чтение файла {file_path.name} по частям: строки {start_row}-{end_row}, размер chunk={CHUNK_SIZE}", "FileProcessor", "_load_file_with_chunking")
            
            for chunk_start in range(start_row, end_row + 1, CHUNK_SIZE):
                chunk_end = min(chunk_start + CHUNK_SIZE, end_row + 1)
                
                # Читаем chunk
                chunk_data = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=chunk_start, max_row=chunk_end, values_only=True), start=chunk_start):
                    if header_indices:
                        # Фильтруем колонки по usecols
                        row_data = [row[i] if i < len(row) else None for i in header_indices]
                    else:
                        row_data = list(row)
                    chunk_data.append(row_data)
                
                if chunk_data:
                    chunk_df = pd.DataFrame(chunk_data, columns=headers if headers else None)
                    chunks.append(chunk_df)
                
                # Логируем прогресс каждые 5 chunks
                if (chunk_start - start_row) // CHUNK_SIZE % 5 == 0:
                    self.logger.debug(f"Загружено {chunk_end - start_row} из {end_row - start_row + 1} строк файла {file_path.name}", "FileProcessor", "_load_file_with_chunking")
            
            wb.close()
            
            # Объединяем chunks
            if chunks:
                df = pd.concat(chunks, ignore_index=True)
                self.logger.debug(f"Файл {file_path.name} загружен по частям: {len(df)} строк", "FileProcessor", "_load_file_with_chunking")
                return df
            else:
                return pd.DataFrame()
                
        except Exception as e:
            self.logger.warning(f"Ошибка при chunking файла {file_path.name}, используем обычную загрузку: {str(e)}", "FileProcessor", "_load_file_with_chunking")
            # Fallback на обычную загрузку
            return pd.read_excel(file_path, **read_params)
    
    def _apply_drop_rules(self, df: pd.DataFrame, drop_rules: List[DropRule], file_name: str, group_name: str = "") -> pd.DataFrame:
        """
        Применяет правила удаления строк (drop_rules) с оптимизацией.
        
        ОПТИМИЗАЦИЯ: Объединяет несколько правил для одной колонки в одну операцию.
        
        Args:
            df: DataFrame для обработки
            drop_rules: Список правил удаления
            file_name: Имя файла для логирования
            group_name: Название группы для статистики
            
        Returns:
            DataFrame после применения правил
        """
        if not drop_rules:
            return df
        
        cleaned = df.copy()
        
        # ОПТИМИЗАЦИЯ: Группируем правила по колонкам для объединения операций
        rules_by_column: Dict[str, List[DropRule]] = {}
        for rule in drop_rules:
            if rule.alias not in cleaned.columns:
                # Колонка может отсутствовать в некоторых файлах - это нормальная ситуация
                self.logger.debug(f"Колонка {rule.alias} отсутствует в файле {file_name}, пропускаем правило", "FileProcessor", "_apply_drop_rules")
                continue
            
            if not rule.remove_unconditionally:
                self.logger.debug(f"Колонка {rule.alias}: remove_unconditionally=False, строки не удаляются", "FileProcessor", "_apply_drop_rules")
                continue
            
            if rule.alias not in rules_by_column:
                rules_by_column[rule.alias] = []
            rules_by_column[rule.alias].append(rule)
        
        # Применяем правила по колонкам (объединенные)
        for column, column_rules in rules_by_column.items():
            # ОПТИМИЗАЦИЯ: Объединяем все запрещенные значения для этой колонки в одно множество
            all_forbidden = set()
            for rule in column_rules:
                all_forbidden.update({str(v).strip().lower() for v in rule.values})
            
            # ОПТИМИЗАЦИЯ: Векторизация вместо apply() для ускорения в 10-50 раз
            # Преобразуем в строки и нормализуем один раз для всех правил колонки
            col_str = cleaned[column].astype(str).str.strip().str.lower()
            
            # Исключаем строки "nan" (которые были NaN) из проверки
            mask_not_nan = col_str != 'nan'
            
            # Проверяем принадлежность к запрещенным значениям (векторизованная операция)
            mask_forbidden = col_str.isin(all_forbidden)
            
            # Исключаем NaN из результата (NaN не считаются запрещенными)
            mask_forbidden = mask_forbidden & mask_not_nan
            
            if not mask_forbidden.any():
                # Нет запрещенных значений для этой колонки
                continue
            
            # ОПТИМИЗАЦИЯ: Проверяем условия для всех правил колонки одновременно
            # Если хотя бы одно правило имеет check_by_inn или check_by_tn, применяем условную логику
            has_conditional_rules = any(rule.check_by_inn or rule.check_by_tn for rule in column_rules)
            
            if not has_conditional_rules:
                # Простое удаление без условий (для всех правил колонки сразу)
                before = len(cleaned)
                cleaned = cleaned[~mask_forbidden]
                dropped_count = before - len(cleaned)
                
                if dropped_count > 0:
                    self.logger.debug(f"Колонка {column}: удалено {dropped_count} строк (безусловно, объединено {len(column_rules)} правил)", "FileProcessor", "_apply_drop_rules")
                    
                    # Собираем статистику для всех правил
                    if ENABLE_STATISTICS and group_name and file_name:
                        if group_name not in self.statistics["files"]:
                            self.statistics["files"][group_name] = {}
                        if file_name not in self.statistics["files"][group_name]:
                            self.statistics["files"][group_name][file_name] = {"dropped_by_rule": {}, "kept_by_rule": {}}
                        if "dropped_by_rule" not in self.statistics["files"][group_name][file_name]:
                            self.statistics["files"][group_name][file_name]["dropped_by_rule"] = {}
                        
                        # Записываем статистику для каждого правила отдельно
                        for rule in column_rules:
                            rule_key = f"{rule.alias}: {', '.join(map(str, rule.values))}"
                            # Приблизительное распределение удаленных строк между правилами
                            rule_dropped = dropped_count // len(column_rules) if len(column_rules) > 1 else dropped_count
                            if rule_key not in self.statistics["files"][group_name][file_name]["dropped_by_rule"]:
                                self.statistics["files"][group_name][file_name]["dropped_by_rule"][rule_key] = 0
                            self.statistics["files"][group_name][file_name]["dropped_by_rule"][rule_key] += rule_dropped
            else:
                # Условное удаление - обрабатываем каждое правило отдельно (сложная логика)
                for rule in column_rules:
                    rule_forbidden = {str(v).strip().lower() for v in rule.values}
                    rule_mask = col_str.isin(rule_forbidden) & mask_not_nan
                    
                    if not rule_mask.any():
                        continue
                    
                    rows_to_remove = rule_mask.copy()
                    
                    # ОПТИМИЗАЦИЯ: Векторизация проверки по ИНН
                    if rule.check_by_inn and "client_id" in cleaned.columns:
                        grouped_by_inn = cleaned.groupby("client_id")[column].apply(
                            lambda x: (~x.astype(str).str.strip().str.lower().isin(rule_forbidden) & (x.astype(str).str.strip().str.lower() != 'nan')).any()
                        )
                        keep_by_inn = cleaned["client_id"].map(grouped_by_inn).fillna(False)
                        rows_to_remove = rows_to_remove & ~keep_by_inn
                    
                    # ОПТИМИЗАЦИЯ: Векторизация проверки по ТН
                    if rule.check_by_tn:
                        tab_col = None
                        if "tab_number" in cleaned.columns:
                            tab_col = "tab_number"
                        elif "manager_id" in cleaned.columns:
                            tab_col = "manager_id"
                        
                        if tab_col:
                            grouped_by_tn = cleaned.groupby(tab_col)[column].apply(
                                lambda x: (~x.astype(str).str.strip().str.lower().isin(rule_forbidden) & (x.astype(str).str.strip().str.lower() != 'nan')).any()
                            )
                            keep_by_tn = cleaned[tab_col].map(grouped_by_tn).fillna(False)
                            rows_to_remove = rows_to_remove & ~keep_by_tn
                    
                    before = len(cleaned)
                    cleaned = cleaned[~rows_to_remove]
                    dropped_count = before - len(cleaned)
                    
                    if dropped_count > 0:
                        self.logger.debug(
                            f"Колонка {column}: удалено {dropped_count} строк "
                            f"(условно: check_by_inn={rule.check_by_inn}, check_by_tn={rule.check_by_tn})",
                            "FileProcessor", "_apply_drop_rules"
                        )
                        
                        # Собираем статистику
                        if ENABLE_STATISTICS and group_name and file_name:
                            rule_key = f"{rule.alias}: {', '.join(map(str, rule.values))} [условно: check_by_inn={rule.check_by_inn}, check_by_tn={rule.check_by_tn}]"
                            if group_name not in self.statistics["files"]:
                                self.statistics["files"][group_name] = {}
                            if file_name not in self.statistics["files"][group_name]:
                                self.statistics["files"][group_name][file_name] = {"dropped_by_rule": {}, "kept_by_rule": {}}
                            if "dropped_by_rule" not in self.statistics["files"][group_name][file_name]:
                                self.statistics["files"][group_name][file_name]["dropped_by_rule"] = {}
                            if rule_key not in self.statistics["files"][group_name][file_name]["dropped_by_rule"]:
                                self.statistics["files"][group_name][file_name]["dropped_by_rule"][rule_key] = 0
                            self.statistics["files"][group_name][file_name]["dropped_by_rule"][rule_key] += dropped_count
        
        return cleaned
    
    def _apply_in_rules(self, df: pd.DataFrame, in_rules: List[IncludeRule], file_name: str, group_name: str = "") -> pd.DataFrame:
        """
        Применяет правила включения строк (in_rules).
        
        Строка попадает в расчет только если она проходит ВСЕ условия из in_rules (И).
        
        Args:
            df: DataFrame для обработки
            in_rules: Список правил включения
            file_name: Имя файла для логирования
            
        Returns:
            DataFrame после применения правил
        """
        if not in_rules:
            return df
        
        # Начинаем с маски True для всех строк
        final_mask = pd.Series(True, index=df.index)
        
        for rule in in_rules:
            if rule.alias not in df.columns:
                # Колонка может отсутствовать в некоторых файлах - это нормальная ситуация
                self.logger.debug(f"Колонка {rule.alias} отсутствует в файле {file_name}, пропускаем правило", "FileProcessor", "_apply_in_rules")
                continue
            
            # Формируем множество разрешенных значений
            allowed = {str(v).strip().lower() for v in rule.values}
            
            def check_value(value: Any) -> bool:
                """Проверяет значение по условию."""
                if pd.isna(value):
                    return False
                value_str = str(value).strip().lower()
                if rule.condition == "in":
                    return value_str in allowed
                elif rule.condition == "not_in":
                    return value_str not in allowed
                return False
            
            # Применяем условие (И - все условия должны выполняться)
            rule_mask = df[rule.alias].apply(check_value)
            final_mask = final_mask & rule_mask
        
        before = len(df)
        result = df[final_mask]
        kept_count = len(result)
        dropped_count = before - kept_count
        self.logger.debug(f"После применения in_rules: оставлено {kept_count} строк из {before}", "FileProcessor", "_apply_in_rules")
        
        # Собираем статистику
        if ENABLE_STATISTICS and group_name and file_name and in_rules:
            if group_name not in self.statistics["files"]:
                self.statistics["files"][group_name] = {}
            if file_name not in self.statistics["files"][group_name]:
                self.statistics["files"][group_name][file_name] = {"dropped_by_rule": {}, "kept_by_rule": {}}
            if "kept_by_rule" not in self.statistics["files"][group_name][file_name]:
                self.statistics["files"][group_name][file_name]["kept_by_rule"] = {}
            
            for rule in in_rules:
                rule_key = f"{rule.alias}: {rule.condition} {', '.join(map(str, rule.values))}"
                if rule_key not in self.statistics["files"][group_name][file_name]["kept_by_rule"]:
                    self.statistics["files"][group_name][file_name]["kept_by_rule"][rule_key] = 0
                # Приблизительная оценка: считаем, что все правила вносят равный вклад
                self.statistics["files"][group_name][file_name]["kept_by_rule"][rule_key] = kept_count
        
        return result
    
    def collect_unique_tab_numbers(self) -> None:
        """
        Собирает уникальные табельные номера из всех файлов.
        
        Алгоритм:
        1. В каждом файле табельные номера должны быть уникальны (если есть дубликаты, берется первая строка)
        2. Поиск выполняется в порядке приоритета:
           - Группы: OD -> RA -> PS
           - Месяцы: декабрь (M-12) -> ноябрь (M-11) -> ... -> январь (M-1)
        3. Для каждого табельного номера берется ПЕРВЫЙ найденный ТБ
        4. Если табельный номер уже найден в файле с более высоким приоритетом,
           он НЕ обновляется - остается ранее найденный ТБ
        
        Результат: каждый табельный номер встречается в итоговом списке только один раз.
        """
        self.logger.info("Начало сбора уникальных табельных номеров", "FileProcessor", "collect_unique_tab_numbers")
        
        # Порядок приоритета групп
        group_priority = {"OD": 1, "RA": 2, "PS": 3}

        # ОПТИМИЗАЦИЯ: Кэш для номеров месяцев
        month_cache = {}
        
        # Извлекаем номер месяца из имени файла
        def extract_month_number(file_name: str) -> int:
            """
            Извлекает номер месяца из имени файла.

            Поддерживает форматы:
            - M-{номер}_{группа}.xlsx (например, M-1_RA.xlsx, M-12_OD.xlsx)
            - {группа}_{номер}.xlsx (например, RA_01.xlsx, OD_12.xlsx)
            - T-{номер} (например, T-11, T-0) - где T-11 = январь, T-0 = декабрь

            Args:
                file_name: Имя файла
                
            Returns:
                int: Номер месяца (1-12) или 0, если не удалось определить
            """
            # ОПТИМИЗАЦИЯ: Проверяем кэш
            if file_name in month_cache:
                return month_cache[file_name]
            
            # Паттерн для формата M-{номер}_{группа}.xlsx
            match = re.search(r'M-(\d{1,2})_', file_name)
            if match:
                month = int(match.group(1))
                if 1 <= month <= 12:
                    month_cache[file_name] = month
                    return month
            
            # Паттерн для формата {группа}_{номер}.xlsx (например, RA_01.xlsx)
            match = re.search(r'_(\d{2})\.', file_name)
            if match:
                month = int(match.group(1))
                if 1 <= month <= 12:
                    month_cache[file_name] = month
                    return month
            
            # Паттерн для формата T-{номер} (T-11 = январь, T-0 = декабрь)
            match = re.search(r'T-(\d{1,2})', file_name)
            if match:
                t_value = int(match.group(1))
                # Преобразуем T-11 -> 1 (январь), T-0 -> 12 (декабрь)
                if 0 <= t_value <= 11:
                    month = 12 - t_value
                    if 1 <= month <= 12:
                        month_cache[file_name] = month
                        return month
            
            # Если не нашли, возвращаем 0 (низкий приоритет)
            result = 0
            month_cache[file_name] = result
            return result
        
        # Собираем все табельные номера с информацией о файлах
        all_tab_data: Dict[str, Dict[str, Any]] = {}
        
        # Проходим по группам в порядке приоритета
        for group in sorted(self.groups, key=lambda x: group_priority.get(x, 999)):
            if group not in self.processed_files:
                continue
            
            group_config = config_manager.get_group_config(group)
            defaults = group_config.defaults
            tab_col = defaults.tab_number_column
            tb_col = defaults.tb_column
            gosb_col = defaults.gosb_column
            fio_col = defaults.fio_column
            
            # Сортируем файлы по номеру месяца (от большего к меньшему)
            files_sorted = sorted(
                self.processed_files[group].items(),
                key=lambda x: extract_month_number(x[0]),
                reverse=True
            )
            
            for file_name, df in files_sorted:
                month = extract_month_number(file_name)
                self.logger.debug(f"Обработка файла {file_name} группы {group}, месяц {month}", "FileProcessor", "collect_unique_tab_numbers")
                
                if tab_col not in df.columns:
                    self.logger.warning(f"Колонка '{tab_col}' не найдена в файле {file_name}", "FileProcessor", "collect_unique_tab_numbers")
                    continue
                
                # Табельные номера уже нормализованы при загрузке файла
                df_normalized = df.copy()
                # Фильтруем пустые и некорректные значения
                df_normalized = df_normalized[df_normalized[tab_col].notna()]
                df_normalized = df_normalized[df_normalized[tab_col] != '']
                
                if len(df_normalized) == 0:
                    continue
                
                # ВАЖНО: В каждом файле табельные номера должны быть уникальны
                # Если у табельного номера несколько разных ТБ, выбираем тот, у которого сумма показателя больше
                # Это делается только если табельный номер еще не встречался ранее
                current_priority = group_priority[group] * 100 + month
                indicator_col = defaults.indicator_column
                
                # ОПТИМИЗАЦИЯ: Выбираем уникальные строки для каждого табельного номера
                # Сначала суммируем показатели по комбинациям ТН+ТБ+ФИО, затем выбираем максимум
                if indicator_col in df_normalized.columns:
                    # Шаг 1: Группируем по ТН+ТБ+ФИО и суммируем показатели (быстро, векторизовано)
                    # ВАЖНО: Включаем fio_col в группировку, чтобы он был доступен после merge
                    # ГОСБ не используется для группировки, но остается в параметрах для обратной совместимости
                    group_cols = [tab_col]
                    if tb_col in df_normalized.columns:
                        group_cols.append(tb_col)
                    if fio_col in df_normalized.columns:
                        group_cols.append(fio_col)
                    
                    grouped = df_normalized.groupby(group_cols, as_index=False)[indicator_col].sum()
                    
                    # Шаг 2: Для каждого ТН находим строку с максимальной суммой (векторизовано)
                    # Используем groupby().idxmax() - это векторизованная операция, заменяет цикл
                    max_indices = grouped.groupby(tab_col)[indicator_col].idxmax()
                    max_rows = grouped.loc[max_indices]
                    
                    # Собираем статистику по выбору табельных номеров
                    if ENABLE_STATISTICS:
                        if group not in self.statistics["tab_selection"]:
                            self.statistics["tab_selection"][group] = {}
                        if file_name not in self.statistics["tab_selection"][group]:
                            self.statistics["tab_selection"][group][file_name] = {
                                "total_variants": 0,
                                "selected_count": 0,
                                "variants_with_multiple": 0
                            }
                        
                        # Подсчитываем количество вариантов для каждого табельного
                        for tab_num in grouped[tab_col].unique():
                            tab_data = grouped[grouped[tab_col] == tab_num]
                            variant_count = len(tab_data)
                            self.statistics["tab_selection"][group][file_name]["total_variants"] += variant_count
                            if variant_count > 1:
                                self.statistics["tab_selection"][group][file_name]["variants_with_multiple"] += 1
                        
                        self.statistics["tab_selection"][group][file_name]["selected_count"] = len(max_rows)
                    
                    # Собираем данные для трекера детальной статистики
                    for _, max_row in max_rows.iterrows():
                        tab_num = max_row[tab_col]
                        tab_data = grouped[grouped[tab_col] == tab_num]
                        
                        # Собираем данные о клиентах (ИНН) для этого табельного номера из исходного DataFrame
                        clients_data = []
                        if "client_id" in df_normalized.columns:
                            tab_rows = df_normalized[df_normalized[tab_col] == tab_num]
                            for _, row in tab_rows.iterrows():
                                client_inn = str(row.get("client_id", ""))
                                client_tb = str(row.get(tb_col, ""))
                                client_fio = str(row.get(fio_col, ""))
                                client_indicator = float(row.get(indicator_col, 0))
                                selected_tb_value = max_row.get(tb_col, '')
                                is_selected = (client_tb == selected_tb_value)
                                
                                clients_data.append({
                                    "ИНН": client_inn,
                                    "ТБ": client_tb,
                                    "ФИО": client_fio,
                                    "Показатель": client_indicator,
                                    "Выбран": is_selected
                                })
                        
                        # Собираем варианты ТБ с суммами
                        tb_variants = {}
                        for _, variant_row in tab_data.iterrows():
                            tb_value = variant_row.get(tb_col, '')
                            sum_value = float(variant_row.get(indicator_col, 0))
                            tb_variants[tb_value] = sum_value
                        
                        selected_tb = max_row.get(tb_col, '')
                        selected_sum = float(max_row.get(indicator_col, 0))
                        
                        # Добавляем данные в трекер
                        # ВАЖНО: tab_num уже нормализован через _normalize_tab_number при загрузке файла
                        # Используем его как есть, но также пробуем нормализовать для совместимости
                        tab_num_str = str(tab_num).strip()
                        self.debug_tracker.add_source_file_data(
                            tab_number=tab_num_str,
                            file_name=file_name,
                            group=group,
                            month=month,
                            clients_data=clients_data,
                            tb_variants=tb_variants,
                            selected_tb=selected_tb,
                            selected_sum=selected_sum
                        )
                        
                        # Логируем для диагностики
                        if self.logger._is_debug_tab_number(tab_num_str):
                            self.logger.debug(
                                f"Добавлены данные в трекер для табельного {tab_num_str} из файла {file_name}: "
                                f"клиентов={len(clients_data)}, вариантов ТБ={len(tb_variants)}, выбран ТБ={selected_tb}",
                                "FileProcessor",
                                "collect_unique_tab_numbers"
                            )
                        
                        if len(tab_data) > 1:
                            # Формируем детальную информацию о вариантах
                            # ВАЖНО: Сумма не маскируется и форматируется с разделителем разрядов
                            variants_list = []
                            for _, variant_row in tab_data.iterrows():
                                sum_value = variant_row.get(indicator_col, 0)
                                # Форматируем сумму с разделителем разрядов и двумя знаками после запятой
                                sum_formatted = f"{sum_value:,.2f}".replace(",", " ").replace(".", ",")
                                variants_list.append(f"ТБ='{variant_row.get(tb_col, '')}' (сумма={sum_formatted})")
                            
                            # Форматируем выбранную сумму с разделителем разрядов
                            selected_sum_formatted = f"{selected_sum:,.2f}".replace(",", " ").replace(".", ",")
                            
                            self.logger.debug(
                                f"В файле {file_name} для табельного {tab_num} найдено {len(tab_data)} вариантов ТБ: "
                                f"{', '.join(variants_list)}. "
                                f"Выбран вариант: ТБ='{selected_tb}' с максимальной суммой показателя: {selected_sum_formatted}",
                                "FileProcessor", "collect_unique_tab_numbers"
                            )
                            
                        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем выбор варианта для указанного табельного
                        if self.logger._is_debug_tab_number(tab_num):
                            variants_info = []
                            for _, variant_row in tab_data.iterrows():
                                sum_value = variant_row.get(indicator_col, 0)
                                # Форматируем сумму с разделителем разрядов (не маскируется)
                                sum_formatted = f"{sum_value:,.2f}".replace(",", " ").replace(".", ",")
                                variants_info.append({
                                    "ТБ": variant_row.get(tb_col, ''),
                                    "ФИО": variant_row.get(fio_col, ''),  # ФИО будет замаскировано в _mask_sensitive_data
                                    "Показатель": sum_formatted  # Форматированная сумма с разделителем
                                })
                            
                            selected_sum_formatted = f"{selected_sum:,.2f}".replace(",", " ").replace(".", ",")
                            
                            self.logger.debug_tab(
                                f"Выбор варианта ТБ для табельного в файле {file_name} (группа {group}): "
                                f"найдено {len(tab_data)} вариантов. Все варианты: {variants_info}. "
                                f"Выбран вариант с максимальной суммой показателя: ТБ='{max_row.get(tb_col, '')}', "
                                f"ФИО='{max_row.get(fio_col, '')}', Показатель={selected_sum_formatted}",
                                tab_number=tab_num,
                                class_name="FileProcessor",
                                func_name="collect_unique_tab_numbers"
                            )
                    
                    # Шаг 3: Находим соответствующие строки в исходном DataFrame через merge (быстро)
                    # Используем merge вместо циклов с mask - это векторизованная операция
                    # ВАЖНО: Включаем все нужные колонки в merge, чтобы они были доступны в df_unique
                    # ГОСБ не используется для merge, но остается в параметрах для обратной совместимости
                    merge_cols = [tab_col]
                    if tb_col in max_rows.columns:
                        merge_cols.append(tb_col)
                    if fio_col in max_rows.columns:
                        merge_cols.append(fio_col)
                    
                    # ВАЖНО: merge сохраняет все колонки из df_normalized, включая те, что не в merge_cols
                    df_unique = df_normalized.merge(
                        max_rows[merge_cols],
                        on=merge_cols,
                        how='inner'
                    ).drop_duplicates(subset=[tab_col], keep='first')
                    
                    # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем, что нужные колонки есть в df_unique
                    missing_cols = []
                    if tb_col not in df_unique.columns:
                        missing_cols.append(tb_col)
                    if fio_col not in df_unique.columns:
                        missing_cols.append(fio_col)
                    
                    if missing_cols:
                        self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: Колонки {missing_cols} не найдены в df_unique после merge для файла {file_name}. Доступные колонки: {list(df_unique.columns)}. merge_cols={merge_cols}, max_rows.columns={list(max_rows.columns)}", "FileProcessor", "collect_unique_tab_numbers")
                    else:
                        # Проверяем, что данные не пустые
                        if len(df_unique) > 0:
                            sample_tb = df_unique[tb_col].iloc[0] if tb_col in df_unique.columns else None
                            sample_fio = df_unique[fio_col].iloc[0] if fio_col in df_unique.columns else None
                            self.logger.debug(f"df_unique после merge для файла {file_name}: {len(df_unique)} строк. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "collect_unique_tab_numbers")
                else:
                    # Если нет колонки с показателем, используем старую логику
                    df_unique = df_normalized.drop_duplicates(subset=[tab_col], keep='first')
                    
                    # Собираем статистику для случая без показателя
                    if ENABLE_STATISTICS:
                        if group not in self.statistics["tab_selection"]:
                            self.statistics["tab_selection"][group] = {}
                        if file_name not in self.statistics["tab_selection"][group]:
                            self.statistics["tab_selection"][group][file_name] = {
                                "total_variants": len(df_normalized),
                                "selected_count": len(df_unique),
                                "variants_with_multiple": 0
                            }
                        else:
                            self.statistics["tab_selection"][group][file_name]["total_variants"] = len(df_normalized)
                            self.statistics["tab_selection"][group][file_name]["selected_count"] = len(df_unique)
                
                if len(df_unique) < len(df_normalized):
                    duplicates_count = len(df_normalized) - len(df_unique)
                    self.logger.debug(f"В файле {file_name} найдено {duplicates_count} дубликатов табельных номеров, оставлено уникальных: {len(df_unique)}", "FileProcessor", "collect_unique_tab_numbers")
                
                # ВАЖНО: Используем нормализованные значения из df_unique напрямую
                # ОПТИМИЗАЦИЯ: Используем itertuples() вместо iterrows() для ускорения (12x быстрее)
                # Получаем индексы колонок один раз для быстрого доступа
                # ГОСБ не используется для обработки, но остается в параметрах для обратной совместимости
                tab_col_idx = df_unique.columns.get_loc(tab_col) if tab_col in df_unique.columns else -1
                tb_col_idx = df_unique.columns.get_loc(tb_col) if tb_col in df_unique.columns else -1
                fio_col_idx = df_unique.columns.get_loc(fio_col) if fio_col in df_unique.columns else -1
                
                # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем наличие колонок
                if tab_col_idx < 0:
                    self.logger.warning(f"Колонка '{tab_col}' не найдена в df_unique для файла {file_name}. Доступные колонки: {list(df_unique.columns)}", "FileProcessor", "collect_unique_tab_numbers")
                if tb_col_idx < 0:
                    self.logger.warning(f"Колонка '{tb_col}' не найдена в df_unique для файла {file_name}. Доступные колонки: {list(df_unique.columns)}", "FileProcessor", "collect_unique_tab_numbers")
                if fio_col_idx < 0:
                    self.logger.warning(f"Колонка '{fio_col}' не найдена в df_unique для файла {file_name}. Доступные колонки: {list(df_unique.columns)}", "FileProcessor", "collect_unique_tab_numbers")
                
                for row_tuple in df_unique.itertuples(index=False):
                    # Получаем нормализованные значения из df_unique (уже нормализованы при загрузке)
                    # itertuples() возвращает tuple, доступ к колонкам по индексу
                    if tab_col_idx >= 0 and tab_col_idx < len(row_tuple):
                        tab_number = str(row_tuple[tab_col_idx])
                    else:
                        tab_number = ""
                    
                    if not tab_number or tab_number == '' or tab_number.lower() == 'nan':
                        continue
                    
                    # ВАЖНО: Если табельный номер уже найден ранее (в файле с более высоким приоритетом),
                    # НЕ обновляем его - оставляем ранее найденный ТБ
                    # Алгоритм: ищем от OD к PS, от декабря к январю, берем ПЕРВЫЙ найденный
                    if tab_number not in all_tab_data:
                        # Табельный номер еще не встречался - добавляем его
                        # ВАЖНО: Извлекаем значения с проверкой на NaN и пустые строки
                        # ГОСБ не используется для обработки, но остается в словаре для обратной совместимости
                        tb_val = row_tuple[tb_col_idx] if tb_col_idx >= 0 and tb_col_idx < len(row_tuple) else None
                        fio_val = row_tuple[fio_col_idx] if fio_col_idx >= 0 and fio_col_idx < len(row_tuple) else None
                        
                        # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Логируем первые несколько записей с детальной информацией
                        if len(all_tab_data) < 5:
                            self.logger.debug(f"Извлечение данных для табельного {tab_number}: tb_col_idx={tb_col_idx}, fio_col_idx={fio_col_idx}, len(row_tuple)={len(row_tuple)}, tb_val={tb_val}, fio_val={fio_val}", "FileProcessor", "collect_unique_tab_numbers")
                        
                        # Преобразуем в строку с обработкой NaN и пустых значений
                        if tb_val is not None and pd.notna(tb_val):
                            tb_str = str(tb_val).strip()
                            if tb_str.lower() in ['nan', 'none', '']:
                                tb_str = ""
                        else:
                            tb_str = ""
                        
                        if fio_val is not None and pd.notna(fio_val):
                            fio_str = str(fio_val).strip()
                            if fio_str.lower() in ['nan', 'none', '']:
                                fio_str = ""
                        else:
                            fio_str = ""
                        
                        # Логируем первые несколько записей для отладки
                        if len(all_tab_data) < 5:
                            self.logger.debug(f"Добавлен табельный {tab_number}: ТБ='{tb_str}', ФИО='{fio_str}' (из файла {file_name})", "FileProcessor", "collect_unique_tab_numbers")
                        
                        all_tab_data[tab_number] = {
                            "tab_number": tab_number,
                            "tb": tb_str,
                            "gosb": "",  # Оставляем для обратной совместимости, но не используем
                            "fio": fio_str,
                            "group": group,
                            "month": month,
                            "priority": current_priority
                        }
                        
                        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем сбор уникального табельного номера
                        if self.logger._is_debug_tab_number(tab_number):
                            self.logger.debug_tab(
                                f"Собран уникальный табельный номер из файла {file_name} (группа {group}, месяц M-{month}): "
                                f"ТБ='{tb_str}', ФИО='{fio_str}', приоритет={current_priority}",
                                tab_number=tab_number,
                                class_name="FileProcessor",
                                func_name="collect_unique_tab_numbers"
                            )
                    # Если табельный номер уже найден, НЕ обновляем - оставляем ранее найденный
                    elif self.logger._is_debug_tab_number(tab_number):
                        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем, что табельный номер уже был найден ранее
                        existing_data = all_tab_data[tab_number]
                        self.logger.debug_tab(
                            f"Табельный номер уже найден ранее в файле {existing_data.get('group', '?')} M-{existing_data.get('month', '?')} "
                            f"(приоритет {existing_data.get('priority', '?')}). "
                            f"Текущий файл {file_name} (группа {group}, месяц M-{month}, приоритет {current_priority}) пропущен - "
                            f"используются ранее найденные значения: ТБ='{existing_data.get('tb', '')}', ФИО='{existing_data.get('fio', '')}'",
                            tab_number=tab_number,
                            class_name="FileProcessor",
                            func_name="collect_unique_tab_numbers"
                        )
        
        self.unique_tab_numbers = all_tab_data
        
        # ВАЖНО: Проверяем на дубликаты (должно быть уникально)
        if len(all_tab_data) != len(set(all_tab_data.keys())):
            self.logger.warning(f"Обнаружены дубликаты табельных номеров в all_tab_data! Всего ключей: {len(all_tab_data)}, уникальных: {len(set(all_tab_data.keys()))}", "FileProcessor", "collect_unique_tab_numbers")
        
        # Логируем статистику по группам и месяцам
        group_stats = {}
        for tab_number, data in all_tab_data.items():
            group = data["group"]
            month = data["month"]
            key = f"{group}_M-{month}"
            if key not in group_stats:
                group_stats[key] = 0
            group_stats[key] += 1
        
        self.logger.debug(f"Распределение табельных номеров по группам и месяцам: {group_stats}", "FileProcessor", "collect_unique_tab_numbers")
        self.logger.info(f"Собрано {len(self.unique_tab_numbers)} уникальных табельных номеров", "FileProcessor", "collect_unique_tab_numbers")
    
    def _process_file_for_raw(self, group: str, file_name: str, df: pd.DataFrame, defaults, month: int) -> Optional[pd.DataFrame]:
        """
        Обрабатывает один файл для листа RAW.
        
        ВАЖНО: Работает с уже загруженными данными из self.processed_files, которые уже содержат
        правильные колонки в зависимости от DATA_MODE (TEST/PROM). Алиасы колонок (tab_number_column,
        indicator_column, tb_column, gosb_column, fio_column) одинаковые для обоих режимов.
        
        Args:
            group: Название группы (OD, RA, PS)
            file_name: Имя файла
            df: DataFrame с данными файла (уже загружен с правильными колонками)
            defaults: Конфигурация по умолчанию для группы (содержит алиасы колонок)
            month: Номер месяца
        
        Returns:
            DataFrame с обработанными данными или None в случае ошибки
        """
        tab_col = defaults.tab_number_column
        tb_col = defaults.tb_column
        gosb_col = defaults.gosb_column
        fio_col = defaults.fio_column
        indicator_col = defaults.indicator_column
        
        # Проверяем наличие необходимых колонок (ГОСБ не требуется для группировки, но оставляем в параметрах)
        required_cols = [tab_col, tb_col, fio_col, indicator_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            self.logger.warning(f"В файле {file_name} отсутствуют колонки: {missing_cols}", "FileProcessor", "_process_file_for_raw")
            return None
        
        # Группируем по уникальным комбинациям ТН+ФИО+ТБ+ИНН (без ГОСБ) и суммируем показатель
        grouped = df.groupby([tab_col, fio_col, tb_col, "client_id"], as_index=False)[indicator_col].sum()
        
        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем данные для указанных табельных в RAW
        # И собираем данные для трекера
        if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0 and tab_col in grouped.columns:
            debug_mask = self._create_debug_tab_mask(grouped, tab_col)
            debug_rows = grouped[debug_mask]
            if len(debug_rows) > 0:
                for _, row in debug_rows.iterrows():
                    tab_num = str(row.get(tab_col, '')).strip()
                    # Собираем данные для трекера
                    # Пробуем найти с нормализацией и без
                    tab_num_normalized = tab_num.zfill(8)
                    tab_data_key = None
                    if tab_num_normalized in self.debug_tracker.tab_data:
                        tab_data_key = tab_num_normalized
                    elif tab_num in self.debug_tracker.tab_data:
                        tab_data_key = tab_num
                    
                    if tab_data_key:
                        if "raw_data" not in self.debug_tracker.tab_data[tab_data_key]:
                            self.debug_tracker.tab_data[tab_data_key]["raw_data"] = {}
                        
                        inn = str(row.get('client_id', ''))
                        if inn not in self.debug_tracker.tab_data[tab_data_key]["raw_data"]:
                            self.debug_tracker.tab_data[tab_data_key]["raw_data"][inn] = {
                                "ТБ": str(row.get(tb_col, '')),
                                "ФИО": str(row.get(fio_col, '')),
                                "sums_by_file": {}
                            }
                        
                        # Добавляем сумму для этого файла
                        file_key = f"{group} (M-{month})"
                        self.debug_tracker.tab_data[tab_data_key]["raw_data"][inn]["sums_by_file"][file_key] = float(row.get(indicator_col, 0))
                    
                    self.logger.debug_tab(
                        f"Подготовка RAW данных для файла {file_name} (группа {group}, месяц M-{month}): "
                        f"ТБ='{row.get(tb_col, '')}', ФИО='{row.get(fio_col, '')}', "
                        f"ИНН={row.get('client_id', '')}, Показатель={row.get(indicator_col, 0):.2f}",
                        tab_number=tab_num,
                        class_name="FileProcessor",
                        func_name="_process_file_for_raw"
                    )
        
        # ВАЖНО: Нормализуем табельные номера перед переименованием (для корректного сравнения с final_df)
        # Применяем нормализацию табельных номеров (8 знаков с лидирующими нулями)
        if tab_col in grouped.columns:
            grouped[tab_col] = grouped[tab_col].apply(
                lambda x: self._normalize_tab_number(x, defaults.tab_number_length, defaults.tab_number_fill_char)
            )
        
        # Переименовываем колонки для единообразия (без ГОСБ)
        grouped = grouped.rename(columns={
            tab_col: "Табельный",
            fio_col: "ФИО",
            tb_col: "ТБ",
            "client_id": "ИНН",
            indicator_col: "Показатель"
        })
        
        # Добавляем информацию о группе и месяце для создания колонок
        grouped["Группа"] = group
        grouped["Месяц"] = month
        grouped["Файл"] = file_name
        grouped["Файл_колонка"] = f"{group} (M-{month})"
        
        return grouped
    
    def prepare_raw_data(self) -> pd.DataFrame:
        """
        Подготавливает сырые данные для листа RAW.
        
        Для каждого файла создает уникальные комбинации ТН+ФИО+ТБ+ГОСБ+ИНН с суммой показателя.
        
        Returns:
            pd.DataFrame: DataFrame с сырыми данными
        """
        self.logger.info("=== Начало подготовки сырых данных для листа 'RAW' ===", "FileProcessor", "prepare_raw_data")
        
        raw_data_list = []
        
        # ОПТИМИЗАЦИЯ: Кэш для номеров месяцев
        month_cache = {}
        
        def extract_month_number(file_name: str) -> int:
            """Извлекает номер месяца из имени файла."""
            if file_name in month_cache:
                return month_cache[file_name]
            match = re.search(r'M-(\d{1,2})_', file_name)
            if match:
                month = int(match.group(1))
                if 1 <= month <= 12:
                    month_cache[file_name] = month
                    return month
            month_cache[file_name] = 0
            return 0
        
        # ОПТИМИЗАЦИЯ: Параллельная обработка всех файлов (независимо от группы)
        # Подготавливаем список всех файлов для обработки
        files_to_process = []
        for group in self.groups:
            if group not in self.processed_files:
                continue
            
            group_config = config_manager.get_group_config(group)
            defaults = group_config.defaults
            
            # Сортируем файлы по номеру месяца
            files_sorted = sorted(
                self.processed_files[group].items(),
                key=lambda x: extract_month_number(x[0])
            )
            
            for file_name, df in files_sorted:
                month = extract_month_number(file_name)
                files_to_process.append((group, file_name, df, defaults, month))
        
        # ОПТИМИЗАЦИЯ: Обрабатываем все файлы параллельно
        self.logger.debug(f"Параллельная обработка {len(files_to_process)} файлов для листа RAW (max_workers={MAX_WORKERS})", "FileProcessor", "prepare_raw_data")
        
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = []
            for group, file_name, df, defaults, month in files_to_process:
                future = executor.submit(self._process_file_for_raw, group, file_name, df, defaults, month)
                futures.append(future)
            
            # Собираем результаты по мере завершения
            for future in as_completed(futures):
                try:
                    grouped = future.result()
                    if grouped is not None:
                        raw_data_list.append(grouped)
                except Exception as e:
                    self.logger.error(f"Ошибка при обработке файла для RAW: {str(e)}", "FileProcessor", "prepare_raw_data")
        
        if not raw_data_list:
            self.logger.warning("Нет данных для листа RAW", "FileProcessor", "prepare_raw_data")
            return pd.DataFrame()
        
        # Объединяем все данные
        raw_df = pd.concat(raw_data_list, ignore_index=True)
        
        # ОПТИМИЗАЦИЯ: Используем pivot_table для создания сводной таблицы (быстрее чем циклы)
        base_cols = ["Табельный", "ФИО", "ТБ", "ИНН"]
        
        # Функция для сортировки колонок: сначала по группе (OD, RA, PS), затем по номеру месяца
        def sort_column_key(col_name: str) -> tuple:
            """
            Функция для сортировки колонок: сначала по группе (OD, RA, PS), затем по номеру месяца.
            
            Returns:
                tuple: (приоритет_группы, номер_месяца) для сортировки
            """
            # Базовые колонки идут первыми
            if col_name in base_cols:
                return (-1, 0)
            
            # Парсим название колонки: "OD (M-1)", "RA (M-12)" и т.д.
            match = re.search(r'^([A-Z]+)\s+\(M-(\d{1,2})\)', col_name)
            if match:
                group = match.group(1)
                month = int(match.group(2))
                
                # Приоритет групп: OD=1, RA=2, PS=3
                group_priority = {"OD": 1, "RA": 2, "PS": 3}.get(group, 999)
                
                return (group_priority, month)
            
            # Если не удалось распарсить, идем в конец
            return (999, 999)
        
        # Используем pivot_table для создания сводной таблицы
        # Индекс - базовые колонки, колонки - файлы, значения - показатели
        try:
            pivot_df = raw_df.pivot_table(
                index=base_cols,
                columns="Файл_колонка",
                values="Показатель",
                aggfunc='sum',
                fill_value=0
            )
            
            # Сбрасываем индекс для получения плоской таблицы
            raw_pivot_df = pivot_df.reset_index()
            
            # Переименовываем колонки (убираем иерархию если есть)
            if isinstance(raw_pivot_df.columns, pd.MultiIndex):
                raw_pivot_df.columns = [col[1] if col[1] else col[0] for col in raw_pivot_df.columns.values]
        except Exception as e:
            # Если pivot_table не сработал, используем альтернативный метод
            self.logger.warning(f"Ошибка при создании pivot_table, используем альтернативный метод: {str(e)}", "FileProcessor", "prepare_raw_data")
            
            # Альтернативный метод: группируем и создаем колонки вручную
            unique_combinations = raw_df[base_cols].drop_duplicates()
            file_data_dict = {}
            for _, row in raw_df.iterrows():
                key = tuple(row[col] for col in base_cols)
                file_col = row["Файл_колонка"]
                if key not in file_data_dict:
                    file_data_dict[key] = {}
                file_data_dict[key][file_col] = row["Показатель"]
            
            result_data = []
            for _, combo_row in unique_combinations.iterrows():
                key = tuple(combo_row[col] for col in base_cols)
                row = {col: combo_row[col] for col in base_cols}
                if key in file_data_dict:
                    row.update(file_data_dict[key])
                result_data.append(row)
            
            raw_pivot_df = pd.DataFrame(result_data)
            
            # Применяем правильную сортировку колонок и для альтернативного метода
            indicator_cols_alt = [col for col in raw_pivot_df.columns if col not in base_cols]
            indicator_cols_sorted_alt = sorted(indicator_cols_alt, key=sort_column_key)
            all_cols_alt = base_cols + indicator_cols_sorted_alt
            raw_pivot_df = raw_pivot_df[all_cols_alt]
        
        # Заполняем NaN нулями
        indicator_cols = [col for col in raw_pivot_df.columns if col not in base_cols]
        if indicator_cols:
            raw_pivot_df[indicator_cols] = raw_pivot_df[indicator_cols].fillna(0)
        
        # Сортируем колонки по приоритету группы и номеру месяца
        indicator_cols_sorted = sorted(indicator_cols, key=sort_column_key)
        all_cols = base_cols + indicator_cols_sorted
        raw_pivot_df = raw_pivot_df[all_cols]
        
        self.logger.info(f"Лист 'RAW': Подготовлено {len(raw_pivot_df)} уникальных комбинаций", "FileProcessor", "prepare_raw_data")
        self.logger.info("=== Завершена подготовка сырых данных для листа 'RAW' ===", "FileProcessor", "prepare_raw_data")
        
        return raw_pivot_df
    
    def _create_file_index(self, group: str, file_name: str, full_name: str, df: pd.DataFrame, defaults) -> Dict[str, float]:
        """
        Создает индекс (словарь) для одного файла: {tab_number: sum}.
        
        ВАЖНО: Работает с уже загруженными данными из self.processed_files, которые уже содержат
        правильные колонки в зависимости от DATA_MODE (TEST/PROM). Алиасы колонок (tab_number_column,
        indicator_column) одинаковые для обоих режимов.
        
        Args:
            group: Название группы (OD, RA, PS)
            file_name: Имя файла
            full_name: Полное имя файла (group_file_name)
            df: DataFrame с данными файла (уже загружен с правильными колонками)
            defaults: Конфигурация по умолчанию для группы (содержит алиасы колонок)
        
        Returns:
            Словарь {tab_number: sum} с суммами показателей по табельным номерам
        """
        tab_col = defaults.tab_number_column
        indicator_col = defaults.indicator_column
        
        if tab_col not in df.columns or indicator_col not in df.columns:
            return {}
        
        # ОПТИМИЗАЦИЯ: Нормализуем табельные номера один раз
        df_normalized = df.copy()
        df_normalized[tab_col] = df_normalized[tab_col].astype(str).str.strip()
        df_normalized = df_normalized[df_normalized[tab_col] != 'nan']
        df_normalized = df_normalized[df_normalized[tab_col] != '']
        
        # ОПТИМИЗАЦИЯ: Группируем по табельным номерам и суммируем показатели один раз для всего файла
        grouped = df_normalized.groupby(tab_col)[indicator_col].sum()
        return grouped.to_dict()
    
    def prepare_summary_data(self) -> pd.DataFrame:
        """
        Подготавливает сводные данные для итогового файла.
        
        Для каждого табельного номера собирает суммы показателей из каждого файла.
        
        Returns:
            pd.DataFrame: DataFrame со сводными данными
        """
        self.logger.info("=== Начало подготовки сводных данных для листа 'Данные' ===", "FileProcessor", "prepare_summary_data")
        
        if not self.unique_tab_numbers:
            self.logger.warning("Уникальные табельные номера не собраны", "FileProcessor", "prepare_summary_data")
            self.collect_unique_tab_numbers()
        
        # ОПТИМИЗАЦИЯ: Кэш для номеров месяцев
        month_cache = {}
        
        # Извлекаем номер месяца из имени файла для сортировки
        def extract_month_number(file_name: str) -> int:
            """Извлекает номер месяца из имени файла."""
            if file_name in month_cache:
                return month_cache[file_name]
            match = re.search(r'M-(\d{1,2})_', file_name)
            if match:
                month = int(match.group(1))
                if 1 <= month <= 12:
                    month_cache[file_name] = month
                    return month
            month_cache[file_name] = 0
            return 0
        
        # Создаем список всех файлов в порядке обработки
        # Порядок: для каждой группы (OD, RA, PS) файлы сортируются по месяцам (M-1, M-2, ..., M-12)
        all_files: List[Tuple[str, str, str]] = []  # (group, file_name, full_name)
        
        # Логируем информацию о группах и месяцах (DEBUG - детальная информация)
        for group in self.groups:
            if group in self.processed_files:
                # Сортируем файлы по номеру месяца (1-12)
                files_sorted = sorted(
                    self.processed_files[group].keys(),
                    key=lambda x: extract_month_number(x)
                )
                months_list = [extract_month_number(fn) for fn in files_sorted]
                self.logger.debug(f"Лист 'Данные': Группа {group}, обрабатываем месяцы: {months_list} (M-{min(months_list)} ... M-{max(months_list)})", "FileProcessor", "prepare_summary_data")
                for file_name in files_sorted:
                    full_name = f"{group}_{file_name}"
                    all_files.append((group, file_name, full_name))
        
        self.logger.debug(f"Лист 'Данные': Всего колонок для обработки: {len(all_files)} (базовые: Табельный, ТБ, ФИО + данные по группам и месяцам)", "FileProcessor", "prepare_summary_data")
        
        # ОПТИМИЗАЦИЯ: Предварительно создаем индексы для всех файлов параллельно
        # Кэшируем конфигурации групп
        self.logger.debug("Лист 'Данные': Параллельное создание индексов по табельным номерам для всех файлов", "FileProcessor", "prepare_summary_data")
        file_indexes = {}  # {full_name: {tab_number: sum}}
        group_configs_cache = {}  # Кэш конфигураций
        
        # Подготавливаем список файлов для обработки
        files_to_index = []
        for group, file_name, full_name in all_files:
            if group in self.processed_files and file_name in self.processed_files[group]:
                df = self.processed_files[group][file_name]
                
                # Кэшируем конфигурацию группы
                if group not in group_configs_cache:
                    group_configs_cache[group] = config_manager.get_group_config(group)
                
                defaults = group_configs_cache[group].defaults
                files_to_index.append((group, file_name, full_name, df, defaults))
        
        # ОПТИМИЗАЦИЯ: Создаем индексы параллельно для всех файлов
        if files_to_index:
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                future_to_file = {
                    executor.submit(self._create_file_index, group, file_name, full_name, df, defaults): (group, file_name, full_name)
                    for group, file_name, full_name, df, defaults in files_to_index
                }
                
                # Обрабатываем результаты по мере завершения
                for future in as_completed(future_to_file):
                    group, file_name, full_name = future_to_file[future]
                    try:
                        file_index = future.result()
                        file_indexes[full_name] = file_index
                    except Exception as e:
                        self.logger.error(f"Ошибка при создании индекса для файла {full_name}: {str(e)}", "FileProcessor", "prepare_summary_data")
                        file_indexes[full_name] = {}
        
        self.logger.debug(f"Лист 'Данные': Индексы созданы для {len(file_indexes)} файлов", "FileProcessor", "prepare_summary_data")
        
        # Создаем структуру данных
        result_data = []
        total_tab_numbers = len(self.unique_tab_numbers)
        self.logger.info(f"Лист 'Данные': Обработка {total_tab_numbers} уникальных табельных номеров", "FileProcessor", "prepare_summary_data")
        
        processed_count = 0
        for tab_number, tab_info in self.unique_tab_numbers.items():
            processed_count += 1
            # Логируем прогресс каждые 100 записей или в начале/конце (DEBUG - детальная информация)
            if processed_count == 1 or processed_count % 100 == 0 or processed_count == total_tab_numbers:
                self.logger.debug(f"Лист 'Данные': Обработано {processed_count} из {total_tab_numbers} табельных номеров ({processed_count * 100 // total_tab_numbers if total_tab_numbers > 0 else 0}%)", "FileProcessor", "prepare_summary_data")
            # Форматируем табельный номер: 8 знаков с лидирующими нулями
            tab_number_formatted = str(tab_number).zfill(8) if tab_number else "00000000"
            
            # ВАЖНО: Извлекаем значения напрямую из словаря (не через get с проверкой)
            # ГОСБ не используется для вывода, но остается в tab_info для обратной совместимости
            tb_value = tab_info.get("tb", "") or ""
            fio_value = tab_info.get("fio", "") or ""
            
            # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Логируем первые несколько записей и каждую 100-ю для отладки
            if processed_count <= 5 or processed_count % 100 == 0:
                self.logger.debug(f"Подготовка строки для табельного {tab_number_formatted}: ТБ='{tb_value}', ФИО='{fio_value}' (из tab_info: {list(tab_info.keys())}, значения: {tab_info})", "FileProcessor", "prepare_summary_data")
                
                # Проверяем, что значения не пустые
                if not tb_value and not fio_value:
                    self.logger.warning(f"ВНИМАНИЕ: Для табельного {tab_number_formatted} все значения (ТБ, ФИО) пустые! tab_info={tab_info}", "FileProcessor", "prepare_summary_data")
            
            row = {
                "Табельный": tab_number_formatted,
                "ТБ": str(tb_value) if tb_value else "",
                "ФИО": str(fio_value) if fio_value else ""
            }
            
            # ОПТИМИЗАЦИЯ: Используем предварительно созданные индексы вместо фильтрации
            for group, file_name, full_name in all_files:
                if full_name in file_indexes:
                    row[full_name] = file_indexes[full_name].get(tab_number, 0)
                else:
                    row[full_name] = 0
            
            # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Если это табельный номер для детального логирования
            if self.logger._is_debug_tab_number(tab_number):
                # Собираем информацию о всех значениях по месяцам
                month_values = {}
                for group, file_name, full_name in all_files:
                    value = row.get(full_name, 0)
                    if value != 0:
                        month_values[full_name] = value
                
                self.logger.debug_tab(
                    f"Подготовка сводных данных для ТН: ТБ='{tb_value}', ФИО='{fio_value}'. "
                    f"Найдено значений по месяцам: {len(month_values)}. "
                    f"Детали: {dict(list(month_values.items())[:10])}",
                    tab_number=tab_number,
                    class_name="FileProcessor",
                    func_name="prepare_summary_data"
                )
            
            result_data.append(row)
        
        self.logger.debug(f"Лист 'Данные': Завершена обработка всех табельных номеров, формирование DataFrame из {len(result_data)} строк", "FileProcessor", "prepare_summary_data")
        result_df = pd.DataFrame(result_data)
        self.logger.debug(f"Лист 'Данные': DataFrame создан, размер: {len(result_df)} строк x {len(result_df.columns)} колонок", "FileProcessor", "prepare_summary_data")
        
        # ВАЖНО: Проверяем, что базовые колонки заполнены данными
        if len(result_df) > 0:
            sample_tb = result_df["ТБ"].iloc[0] if "ТБ" in result_df.columns else None
            sample_fio = result_df["ФИО"].iloc[0] if "ФИО" in result_df.columns else None
            self.logger.debug(f"summary_df (result_df) создан: {len(result_df)} строк. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "prepare_summary_data")
            
            # Проверяем, что не все значения пустые
            if "ТБ" in result_df.columns:
                non_empty_tb = result_df["ТБ"].notna() & (result_df["ТБ"] != "")
                if non_empty_tb.sum() == 0:
                    self.logger.warning(f"В summary_df все значения ТБ пустые!", "FileProcessor", "prepare_summary_data")
                else:
                    self.logger.debug(f"В summary_df заполнено ТБ: {non_empty_tb.sum()}/{len(result_df)} строк", "FileProcessor", "prepare_summary_data")
        
        # Собираем итоговую статистику
        if ENABLE_STATISTICS:
            # Количество КМ (табельных номеров)
            self.statistics["summary"]["total_km"] = len(result_df)
            
            # Количество уникальных клиентов (уже сохранено при загрузке файлов)
            # Если не было сохранено, пытаемся получить из result_df
            if "total_clients" not in self.statistics["summary"] or self.statistics["summary"]["total_clients"] == 0:
                if "ИНН" in result_df.columns:
                    unique_clients = result_df["ИНН"].nunique()
                    self.statistics["summary"]["total_clients"] = unique_clients
                elif "ID_Clients" in result_df.columns:
                    unique_clients = result_df["ID_Clients"].nunique()
                    self.statistics["summary"]["total_clients"] = unique_clients
            
            # Количество КМ по ТБ
            if "ТБ" in result_df.columns:
                by_tb = result_df["ТБ"].value_counts().to_dict()
                self.statistics["summary"]["by_tb"] = by_tb
            
            # Количество КМ по ГОСБ - убрано по требованию (считаем только по ТБ)
        
        # Нормализуем табельные номера и ИНН в выходных данных
        # Получаем параметры нормализации из первой группы (все группы должны иметь одинаковые параметры)
        first_group = list(config_manager.groups.keys())[0] if config_manager.groups else None
        if first_group:
            defaults = config_manager.get_group_config(first_group).defaults
            if "Табельный" in result_df.columns:
                result_df["Табельный"] = result_df["Табельный"].apply(
                    lambda x: self._normalize_tab_number(x, defaults.tab_number_length, defaults.tab_number_fill_char)
                )
            if "ID_Clients" in result_df.columns:
                result_df["ID_Clients"] = result_df["ID_Clients"].apply(
                    lambda x: self._normalize_inn(x, defaults.inn_length, defaults.inn_fill_char)
                )

        # ВАЖНО: Проверяем на дубликаты табельных номеров в итоговом результате
        if "Табельный" in result_df.columns:
            duplicates = result_df[result_df.duplicated(subset=["Табельный"], keep=False)]
            if len(duplicates) > 0:
                duplicate_tabs = duplicates["Табельный"].unique()
                self.logger.warning(f"Лист 'Данные': Обнаружено {len(duplicate_tabs)} дубликатов табельных номеров в итоговом результате! Примеры: {list(duplicate_tabs[:5])}", "FileProcessor", "prepare_summary_data")
                # Удаляем дубликаты, оставляя первую запись
                # ВАЖНО: Сохраняем базовые колонки при удалении дубликатов
                result_df = result_df.drop_duplicates(subset=["Табельный"], keep='first')
                self.logger.warning(f"Лист 'Данные': Дубликаты удалены, осталось {len(result_df)} уникальных табельных номеров", "FileProcessor", "prepare_summary_data")
                
                # Проверяем, что базовые колонки не потерялись после drop_duplicates
                if len(result_df) > 0:
                    sample_tb = result_df["ТБ"].iloc[0] if "ТБ" in result_df.columns else None
                    sample_fio = result_df["ФИО"].iloc[0] if "ФИО" in result_df.columns else None
                    self.logger.debug(f"После drop_duplicates: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "prepare_summary_data")
        
        # Упорядочиваем колонки: сначала базовые, потом по группам и месяцам
        self.logger.debug("Лист 'Данные': Упорядочивание колонок", "FileProcessor", "prepare_summary_data")
        base_columns = ["Табельный", "ТБ", "ФИО"]
        data_columns = [full_name for _, _, full_name in all_files]
        ordered_columns = base_columns + data_columns
        
        # Оставляем только существующие колонки
        existing_columns = [col for col in ordered_columns if col in result_df.columns]
        # Добавляем колонки, которых нет в списке (на случай если что-то пропущено)
        other_columns = [col for col in result_df.columns if col not in existing_columns]
        final_columns = existing_columns + other_columns
        
        result_df = result_df[final_columns]
        self.logger.debug(f"Лист 'Данные': Колонки упорядочены, итоговое количество: {len(result_df.columns)}", "FileProcessor", "prepare_summary_data")
        
        # ВАЖНО: Финальная проверка перед возвратом
        if len(result_df) > 0:
            # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем наличие базовых колонок
            base_columns_check = ["Табельный", "ТБ", "ФИО"]
            missing_base = [col for col in base_columns_check if col not in result_df.columns]
            if missing_base:
                self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: В summary_df отсутствуют базовые колонки: {missing_base}. Доступные колонки: {list(result_df.columns)}", "FileProcessor", "prepare_summary_data")
            else:
                self.logger.debug(f"Проверка базовых колонок: все базовые колонки присутствуют в summary_df", "FileProcessor", "prepare_summary_data")
            
            sample_tb = result_df["ТБ"].iloc[0] if "ТБ" in result_df.columns else None
            sample_fio = result_df["ФИО"].iloc[0] if "ФИО" in result_df.columns else None
            self.logger.debug(f"Финальный summary_df: {len(result_df)} строк x {len(result_df.columns)} колонок. Пример первой строки: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "prepare_summary_data")
            
            # Проверяем, что не все значения пустые
            if "ТБ" in result_df.columns:
                non_empty_tb = result_df["ТБ"].notna() & (result_df["ТБ"] != "")
                non_empty_fio = result_df["ФИО"].notna() & (result_df["ФИО"] != "") if "ФИО" in result_df.columns else pd.Series([False] * len(result_df))
                self.logger.debug(f"Финальная проверка заполненности: ТБ={non_empty_tb.sum()}/{len(result_df)}, ФИО={non_empty_fio.sum()}/{len(result_df)}", "FileProcessor", "prepare_summary_data")
                
                if non_empty_tb.sum() == 0:
                    self.logger.warning(f"ВНИМАНИЕ: В summary_df все значения ТБ пустые!", "FileProcessor", "prepare_summary_data")
                if non_empty_fio.sum() == 0:
                    self.logger.warning(f"ВНИМАНИЕ: В summary_df все значения ФИО пустые!", "FileProcessor", "prepare_summary_data")
        
        self.logger.info(f"Лист 'Данные': Подготовлено {len(result_df)} строк сводных данных, колонок: {len(result_df.columns)}", "FileProcessor", "prepare_summary_data")
        self.logger.info("=== Завершена подготовка сводных данных для листа 'Данные' ===", "FileProcessor", "prepare_summary_data")
        
        return result_df
    
    def prepare_calculated_data(self, summary_df: pd.DataFrame) -> pd.DataFrame:
        """
        Подготавливает данные с расчетами для второго листа.
        
        Варианты расчета:
        1: Как есть - просто сумма
        2: Прирост по 2 месяцам (текущий - предыдущий)
        3: Прирост по трем периодам (М-3 - 2*М-2 + М-1)
        
        Args:
            summary_df: DataFrame с исходными данными из prepare_summary_data
            
        Returns:
            pd.DataFrame: DataFrame с расчетными данными
        """
        self.logger.info("=== Начало подготовки расчетных данных для листа 'Расчеты' ===", "FileProcessor", "prepare_calculated_data")

        # ВАЖНО: Базовые текстовые колонки, которые НЕ должны конвертироваться в числа
        base_text_columns = ['Табельный', 'ТБ', 'ФИО', 'ИНН']

        # ОПТИМИЗАЦИЯ: Кэш для номеров месяцев
        month_cache = {}
        
        # Извлекаем номер месяца из имени файла
        def extract_month_number(file_name: str) -> int:
            """Извлекает номер месяца из имени файла."""
            if file_name in month_cache:
                return month_cache[file_name]
            match = re.search(r'M-(\d{1,2})_', file_name)
            if match:
                month = int(match.group(1))
                if 1 <= month <= 12:
                    month_cache[file_name] = month
                    return month
            month_cache[file_name] = 0
            return 0
        
        # Функция для генерации понятного имени колонки на основе типа расчета
        def generate_column_name(group: str, month: int, calc_type: int, 
                                 prev_month: Optional[int] = None, 
                                 prev2_month: Optional[int] = None) -> str:
            """
            Генерирует понятное имя колонки в формате: OD (M-1) [как считалось]
            
            Args:
                group: Название группы (OD, RA, PS)
                month: Номер текущего месяца
                calc_type: Тип расчета (1, 2, 3)
                prev_month: Номер предыдущего месяца (для типа 2 и 3)
                prev2_month: Номер пред-предыдущего месяца (для типа 3)
                
            Returns:
                str: Понятное имя колонки в формате "OD (M-1) [описание расчета]"
            """
            month_str = f"M-{month}"
            period_part = f"{group} ({month_str})"
            
            if calc_type == 1:
                # Тип 1: Как есть - просто факт
                return f"{period_part} [факт]"
            elif calc_type == 2:
                # Тип 2: Прирост по 2 месяцам
                if prev_month is not None:
                    prev_month_str = f"M-{prev_month}"
                    return f"{period_part} [{month_str}→{prev_month_str}]"
                else:
                    # Первый месяц
                    return f"{period_part} [факт]"
            elif calc_type == 3:
                # Тип 3: Прирост по трем периодам
                if prev_month is not None and prev2_month is not None:
                    prev_month_str = f"M-{prev_month}"
                    prev2_month_str = f"M-{prev2_month}"
                    return f"{period_part} [{month_str}-2*{prev_month_str}+{prev2_month_str}]"
                elif prev_month is not None:
                    # Второй месяц
                    prev_month_str = f"M-{prev_month}"
                    return f"{period_part} [{month_str}→{prev_month_str}]"
                else:
                    # Первый месяц
                    return f"{period_part} [факт]"
            else:
                return f"{period_part} [факт]"
        
        # Базовые колонки
        base_columns = ["Табельный", "ТБ", "ФИО"]
        
        # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем, что базовые колонки есть в summary_df перед копированием
        if not all(col in summary_df.columns for col in base_columns):
            missing_cols = [col for col in base_columns if col not in summary_df.columns]
            self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: В summary_df отсутствуют базовые колонки: {missing_cols}. Доступные колонки: {list(summary_df.columns)}", "FileProcessor", "prepare_calculated_data")
            raise ValueError(f"Отсутствуют базовые колонки в summary_df: {missing_cols}")
        else:
            self.logger.debug(f"Проверка summary_df: все базовые колонки присутствуют перед копированием", "FileProcessor", "prepare_calculated_data")
        
        # ВАЖНО: Создаем копию ПЕРЕД конвертацией, чтобы не испортить исходные данные
        # Сбрасываем индекс, чтобы гарантировать совпадение строк
        calculated_df = summary_df.copy().reset_index(drop=True)
        
        # ВАЖНО: Конвертируем числовые колонки ТОЛЬКО в calculated_df, а не в summary_df
        # Это нужно делать ПОСЛЕ копирования, чтобы не испортить исходные данные
        
        # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем данные сразу после копирования, ДО конвертации
        if len(calculated_df) > 0:
            sample_tb_before = calculated_df["ТБ"].iloc[0] if "ТБ" in calculated_df.columns else None
            sample_fio_before = calculated_df["ФИО"].iloc[0] if "ФИО" in calculated_df.columns else None
            self.logger.debug(f"calculated_df сразу после копирования (ДО конвертации): ТБ='{sample_tb_before}', ФИО='{sample_fio_before}'", "FileProcessor", "prepare_calculated_data")
        
        # ВАЖНО: Проверяем, что базовые колонки есть и не пустые в calculated_df
        if not all(col in calculated_df.columns for col in base_columns):
            missing_cols = [col for col in base_columns if col not in calculated_df.columns]
            self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: После копирования в calculated_df отсутствуют базовые колонки: {missing_cols}. Доступные колонки: {list(calculated_df.columns)[:10]}", "FileProcessor", "prepare_calculated_data")
            raise ValueError(f"Отсутствуют базовые колонки после копирования: {missing_cols}")
        
        # ОПТИМИЗАЦИЯ: Конвертируем все числовые колонки в числовой тип перед вычислениями
        # Это исправляет ошибку "unsupported operand type(s) for -: 'str' and 'float'"
        # ВАЖНО: Исключаем базовые текстовые колонки из конвертации!
        for col in calculated_df.columns:
            if col not in base_text_columns:  # Пропускаем текстовые колонки
                try:
                    # Пробуем конвертировать в числовой тип только если колонка не текстовая
                    calculated_df[col] = pd.to_numeric(calculated_df[col], errors='coerce')
                except Exception:
                    pass  # Если не получилось, оставляем как есть
        
        # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем, что данные не пустые
        if len(calculated_df) > 0:
            sample_tb = calculated_df["ТБ"].iloc[0] if "ТБ" in calculated_df.columns else None
            sample_fio = calculated_df["ФИО"].iloc[0] if "ФИО" in calculated_df.columns else None
            self.logger.debug(f"calculated_df создан из summary_df: {len(calculated_df)} строк x {len(calculated_df.columns)} колонок. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "prepare_calculated_data")
            
            # Проверяем заполненность базовых колонок
            non_empty_tb = calculated_df["ТБ"].notna() & (calculated_df["ТБ"] != "") if "ТБ" in calculated_df.columns else pd.Series([False] * len(calculated_df))
            non_empty_fio = calculated_df["ФИО"].notna() & (calculated_df["ФИО"] != "") if "ФИО" in calculated_df.columns else pd.Series([False] * len(calculated_df))
            self.logger.debug(f"Заполненность базовых колонок в calculated_df: ТБ={non_empty_tb.sum()}/{len(calculated_df)}, ФИО={non_empty_fio.sum()}/{len(calculated_df)}", "FileProcessor", "prepare_calculated_data")
        
        # Словарь для переименования колонок
        rename_dict = {}
        
        # Получаем список всех файлов в порядке обработки
        all_files: List[Tuple[str, str, str, int]] = []  # (group, file_name, full_name, month)
        
        for group in self.groups:
            if group in self.processed_files:
                files_sorted = sorted(
                    self.processed_files[group].keys(),
                    key=lambda x: extract_month_number(x)
                )
                for file_name in files_sorted:
                    month = extract_month_number(file_name)
                    full_name = f"{group}_{file_name}"
                    all_files.append((group, file_name, full_name, month))
        
        # Сортируем по группе и месяцу
        all_files_sorted = sorted(all_files, key=lambda x: (x[0], x[3]))
        
        # Для каждой группы обрабатываем файлы по порядку
        for group in self.groups:
            group_files = [(g, fn, fname, m) for g, fn, fname, m in all_files_sorted if g == group]
            if not group_files:
                continue
            
            self.logger.debug(f"Лист 'Расчеты': Обработка группы {group}, файлов: {len(group_files)}", "FileProcessor", "prepare_calculated_data")
            group_config = config_manager.get_group_config(group)
            
            for idx, (g, file_name, full_name, month) in enumerate(group_files):
                if full_name not in calculated_df.columns:
                    continue
                
                # Получаем конфигурацию для файла
                file_config = config_manager.get_config_for_file(group, file_name)
                calc_type = file_config.get("calculation_type", 1)
                first_month_val = file_config.get("first_month_value", "self")
                three_periods_mode = file_config.get("three_periods_first_months", "zero_both")
                
                # Определяем предыдущие месяцы для генерации имени
                prev_month = None
                prev2_month = None
                
                if calc_type == 2 and idx > 0:
                    prev_month = group_files[idx - 1][3]
                elif calc_type == 3:
                    if idx > 0:
                        prev_month = group_files[idx - 1][3]
                    if idx > 1:
                        prev2_month = group_files[idx - 2][3]
                
                # Генерируем понятное имя колонки
                new_name = generate_column_name(group, month, calc_type, prev_month, prev2_month)
                rename_dict[full_name] = new_name
                
                # Логируем информацию о типе расчета
                calc_type_names = {1: "Как есть (факт)", 2: "Прирост по 2 месяцам", 3: "Прирост по трем периодам"}
                calc_desc = calc_type_names.get(calc_type, f"Тип {calc_type}")
                if calc_type == 2:
                    if idx == 0:
                        calc_desc += f", первый месяц: {first_month_val}"
                    else:
                        calc_desc += f", M-{month} - M-{prev_month}"
                elif calc_type == 3:
                    if idx == 0:
                        calc_desc += f", режим: {three_periods_mode}"
                    elif idx == 1:
                        calc_desc += f", режим: {three_periods_mode}, M-{month} - M-{prev_month}"
                    else:
                        calc_desc += f", M-{month} - 2*M-{prev_month} + M-{prev2_month}"
                
                self.logger.debug(f"Лист 'Расчеты': Группа {group}, месяц M-{month}, тип расчета: {calc_desc}, колонка: {new_name}", "FileProcessor", "prepare_calculated_data")
                
                if calc_type == 1:
                    # Вариант 1: Как есть - просто копируем значение
                    # ОПТИМИЗАЦИЯ: Конвертируем в числовой тип
                    calculated_df[full_name] = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                
                elif calc_type == 2:
                    # Вариант 2: Прирост по 2 месяцам
                    if idx == 0:
                        # Первый месяц
                        if first_month_val == "self":
                            # ОПТИМИЗАЦИЯ: Конвертируем в числовой тип
                            calculated_df[full_name] = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                        else:  # "zero"
                            calculated_df[full_name] = 0
                    else:
                        # Текущий месяц минус предыдущий
                        prev_file_name = group_files[idx - 1][2]
                        if prev_file_name in summary_df.columns:
                            # ОПТИМИЗАЦИЯ: Конвертируем в числовой тип перед вычислением
                            curr_val = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                            prev_val = pd.to_numeric(summary_df[prev_file_name], errors='coerce').fillna(0)
                            calculated_df[full_name] = curr_val - prev_val
                            
                            # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем расчет для указанного табельного
                            if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0 and "Табельный" in summary_df.columns:
                                debug_mask = self._create_debug_tab_mask(summary_df, "Табельный")
                                if debug_mask.any():
                                    debug_idx = summary_df[debug_mask].index[0]
                                    curr_val_debug = curr_val.loc[debug_idx] if debug_idx in curr_val.index else 0
                                    prev_val_debug = prev_val.loc[debug_idx] if debug_idx in prev_val.index else 0
                                    result_debug = calculated_df.loc[debug_idx, full_name] if debug_idx in calculated_df.index else 0
                                    
                                    self.logger.debug_tab(
                                        f"Расчет типа 2 для группы {group}, месяц M-{month}: "
                                        f"текущее значение (M-{month})={curr_val_debug}, "
                                        f"предыдущее значение (M-{prev_month})={prev_val_debug}, "
                                        f"результат (прирост)={result_debug}",
                                        tab_number=None,  # Проверка уже сделана через debug_mask
                                        class_name="FileProcessor",
                                        func_name="prepare_calculated_data"
                                    )
                        else:
                            calculated_df[full_name] = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                
                elif calc_type == 3:
                    # Вариант 3: Прирост по трем периодам (М-3 - 2*М-2 + М-1)
                    if idx == 0:
                        # Первый месяц
                        if three_periods_mode == "self_first_diff_second":
                            # ОПТИМИЗАЦИЯ: Конвертируем в числовой тип
                            calculated_df[full_name] = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                        else:  # "zero_both" или "zero_first_diff_second"
                            calculated_df[full_name] = 0
                    elif idx == 1:
                        # Второй месяц
                        if three_periods_mode == "zero_both":
                            calculated_df[full_name] = 0
                        else:  # "zero_first_diff_second" или "self_first_diff_second"
                            prev_file_name = group_files[0][2]
                            if prev_file_name in summary_df.columns:
                                # ОПТИМИЗАЦИЯ: Конвертируем в числовой тип перед вычислением
                                curr_val = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                                prev_val = pd.to_numeric(summary_df[prev_file_name], errors='coerce').fillna(0)
                                calculated_df[full_name] = curr_val - prev_val
                            else:
                                calculated_df[full_name] = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                    else:
                        # М-3 - 2*М-2 + М-1
                        # ОПТИМИЗАЦИЯ: Конвертируем в числовой тип перед вычислением
                        curr_val = pd.to_numeric(summary_df[full_name], errors='coerce').fillna(0)
                        prev1_file_name = group_files[idx - 1][2]
                        prev2_file_name = group_files[idx - 2][2]
                        
                        if prev1_file_name in summary_df.columns:
                            prev1_val = pd.to_numeric(summary_df[prev1_file_name], errors='coerce').fillna(0)
                        else:
                            prev1_val = 0
                        
                        if prev2_file_name in summary_df.columns:
                            prev2_val = pd.to_numeric(summary_df[prev2_file_name], errors='coerce').fillna(0)
                        else:
                            prev2_val = 0
                        
                        calculated_df[full_name] = curr_val - 2 * prev1_val + prev2_val
                        
                        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем расчет для указанного табельного
                        if DEBUG_TAB_NUMBER and "Табельный" in summary_df.columns:
                            debug_mask = summary_df["Табельный"].astype(str).str.strip().str.lstrip('0') == str(DEBUG_TAB_NUMBER).strip().lstrip('0')
                            if debug_mask.any():
                                debug_idx = summary_df[debug_mask].index[0]
                                curr_val_debug = curr_val.loc[debug_idx] if debug_idx in curr_val.index else 0
                                prev1_val_debug = prev1_val.loc[debug_idx] if debug_idx in prev1_val.index else 0
                                prev2_val_debug = prev2_val.loc[debug_idx] if debug_idx in prev2_val.index else 0
                                result_debug = calculated_df.loc[debug_idx, full_name] if debug_idx in calculated_df.index else 0
                                
                                self.logger.debug_tab(
                                    f"Расчет типа 3 для группы {group}, месяц M-{month}: "
                                    f"текущее значение (M-{month})={curr_val_debug}, "
                                    f"предыдущее значение (M-{prev_month})={prev1_val_debug}, "
                                    f"пред-предыдущее значение (M-{prev2_month})={prev2_val_debug}, "
                                    f"результат (M-{month} - 2*M-{prev_month} + M-{prev2_month})={result_debug}",
                                    tab_number=DEBUG_TAB_NUMBER,
                                    class_name="FileProcessor",
                                    func_name="prepare_calculated_data"
                                )
        
        # Переименовываем колонки на понятные имена (только те, которые существуют в DataFrame)
        # ВАЖНО: Исключаем базовые колонки из переименования
        existing_rename_dict = {k: v for k, v in rename_dict.items() if k in calculated_df.columns and k not in base_columns}
        calculated_df = calculated_df.rename(columns=existing_rename_dict)
        self.logger.debug(f"Лист 'Расчеты': Переименовано колонок: {len(existing_rename_dict)}", "FileProcessor", "prepare_calculated_data")
        
        # ВАЖНО: Проверяем, что базовые колонки не потерялись после переименования
        if not all(col in calculated_df.columns for col in base_columns):
            missing_cols = [col for col in base_columns if col not in calculated_df.columns]
            self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: После переименования в calculated_df отсутствуют базовые колонки: {missing_cols}. Доступные колонки: {list(calculated_df.columns)}", "FileProcessor", "prepare_calculated_data")
            raise ValueError(f"Потеряны базовые колонки после переименования: {missing_cols}")
        else:
            self.logger.debug(f"Проверка после переименования: все базовые колонки присутствуют в calculated_df", "FileProcessor", "prepare_calculated_data")
        
        # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем, что данные не пустые
        if len(calculated_df) > 0:
            sample_tb = calculated_df["ТБ"].iloc[0] if "ТБ" in calculated_df.columns else None
            sample_fio = calculated_df["ФИО"].iloc[0] if "ФИО" in calculated_df.columns else None
            self.logger.debug(f"calculated_df после переименования: {len(calculated_df)} строк x {len(calculated_df.columns)} колонок. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "prepare_calculated_data")
            
            # Проверяем заполненность базовых колонок после переименования
            non_empty_tb = calculated_df["ТБ"].notna() & (calculated_df["ТБ"] != "") if "ТБ" in calculated_df.columns else pd.Series([False] * len(calculated_df))
            non_empty_fio = calculated_df["ФИО"].notna() & (calculated_df["ФИО"] != "") if "ФИО" in calculated_df.columns else pd.Series([False] * len(calculated_df))
            self.logger.debug(f"Заполненность базовых колонок после переименования: ТБ={non_empty_tb.sum()}/{len(calculated_df)}, ФИО={non_empty_fio.sum()}/{len(calculated_df)}", "FileProcessor", "prepare_calculated_data")
        
        # НЕ рассчитываем вертикальные ранги (убрано для варианта 3)
        # calculated_df = self._calculate_ranks(calculated_df, all_files_sorted, config_manager)

        self.logger.info(f"Лист 'Расчет': Подготовлено {len(calculated_df)} строк расчетных данных, колонок: {len(calculated_df.columns)}", "FileProcessor", "prepare_calculated_data")
        self.logger.info("=== Завершена подготовка расчетных данных для листа 'Расчет' ===", "FileProcessor", "prepare_calculated_data")

        return calculated_df
    
    def _normalize_group(self, group_name: str, direction: str, month_data: Dict[int, Dict[str, str]], calculated_df: pd.DataFrame) -> Dict[str, pd.Series]:
        """
        Нормализует показатели для одной группы (OD, RA или PS).
        
        Args:
            group_name: Название группы (OD, RA, PS)
            direction: Направление нормализации ("MAX" или "MIN")
            month_data: Словарь {month: {group: col_name}} с колонками по месяцам и группам
            calculated_df: DataFrame с расчетными данными
        
        Returns:
            Словарь {norm_col_name: normalized_series} с нормализованными значениями
        """
        # Собираем все колонки для данного показателя
        group_cols = {}
        for month in sorted(month_data.keys()):
            col = month_data[month].get(group_name)
            if col and col in calculated_df.columns:
                group_cols[month] = col
        
        if not group_cols:
            return {}
        
        normalized_cols = {}
        
        # ОПТИМИЗАЦИЯ: Создаем временный DataFrame с данными показателя
        group_data = pd.DataFrame(index=calculated_df.index)
        for month, col in group_cols.items():
            # Используем fillna(0) только для расчета, но сохраняем NaN для проверки
            group_data[f"M-{month}"] = calculated_df[col]
        
        # Нормализуем для каждого КМ (горизонтально по месяцам)
        # Для каждого КМ находим min и max по месяцам (игнорируя NaN)
        group_min = group_data.min(axis=1, skipna=True)
        group_max = group_data.max(axis=1, skipna=True)
        group_range = group_max - group_min
        
        # ОПТИМИЗАЦИЯ: Обрабатываем деление на ноль и одинаковые значения
        # Проверяем количество месяцев с данными (не NaN и не 0) для каждого КМ
        non_zero_count = (group_data.notna() & (group_data != 0)).sum(axis=1)
        mask_zero_range = (group_range < 1e-10) | group_range.isna()  # Все значения одинаковы или разница очень мала или все NaN
        mask_single_month = non_zero_count <= 1  # Только один месяц с данными или все нули/NaN
        
        # Защита от деления на ноль: заменяем нули в group_range на 1
        group_range_safe = group_range.where(~mask_zero_range, 1.0)
        
        # Нормализуем
        for month in sorted(group_cols.keys()):
            norm_col_name = f"{group_name}_norm (M-{month})"
            col_data = group_data[f"M-{month}"]
            
            # ОПТИМИЗАЦИЯ: Векторизованная нормализация с обработкой edge cases
            if direction == "MAX":
                # Больше = лучше: нормализуем к [0, 1]
                normalized = (col_data - group_min) / group_range_safe
            else:  # direction == "MIN"
                # Меньше = лучше: инвертируем нормализацию
                normalized = (group_max - col_data) / group_range_safe
            
            # Обрабатываем edge cases (векторизованно)
            # ВАЖНО: Сначала обрабатываем случай "только один месяц с данными", 
            # затем случай "все значения одинаковы"
            
            # Случай 1: Только один месяц с данными (не нулями)
            # Месяц с данными получает 1.0, остальные (нули) получают 0.0
            if direction == "MAX":
                # Для MAX: месяц с максимальным значением (ненулевым) = 1.0, остальные = 0.0
                # Проверяем, является ли текущий месяц максимальным и ненулевым
                is_max_and_nonzero = (col_data == group_max) & (col_data != 0) & (non_zero_count == 1)
                # Для КМ с одним месяцем данных: сначала всем 0, затем максимуму 1.0
                normalized = normalized.where(~mask_single_month, 0.0)  # Сначала всем 0
                normalized = normalized.where(~is_max_and_nonzero, 1.0)  # Затем максимуму 1.0
            else:  # direction == "MIN"
                # Для MIN: месяц с минимальным значением (ненулевым) = 1.0, остальные = 0.0
                # Проверяем, является ли текущий месяц минимальным ненулевым
                is_min_and_nonzero = (col_data == group_min) & (col_data != 0) & (non_zero_count == 1)
                # Для КМ с одним месяцем данных: сначала всем 0, затем минимуму 1.0
                normalized = normalized.where(~mask_single_month, 0.0)  # Сначала всем 0
                normalized = normalized.where(~is_min_and_nonzero, 1.0)  # Затем минимуму 1.0
            
            # Случай 2: Все значения одинаковы (включая все нули) - всем 0.5
            # Это применяется только если НЕ случай "один месяц с данными"
            # (mask_zero_range может быть True и для случая "один месяц", поэтому проверяем ~mask_single_month)
            mask_all_same_not_single = mask_zero_range & ~mask_single_month
            normalized = normalized.where(~mask_all_same_not_single, 0.5)
            
            # Защита от выхода за границы [0, 1] (из-за погрешности вычислений)
            normalized = normalized.clip(0.0, 1.0)
            
            normalized_cols[norm_col_name] = normalized
            
            # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем нормализацию для указанного табельного
            if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0 and "Табельный" in calculated_df.columns:
                debug_mask = self._create_debug_tab_mask(calculated_df, "Табельный")
                if debug_mask.any():
                    debug_idx = calculated_df[debug_mask].index[0]
                    col = group_cols[month]
                    original_value = calculated_df.loc[debug_idx, col] if col in calculated_df.columns else None
                    normalized_value = normalized.loc[debug_idx] if debug_idx in normalized.index else None
                    min_val = group_min.loc[debug_idx] if debug_idx in group_min.index else None
                    max_val = group_max.loc[debug_idx] if debug_idx in group_max.index else None
                    
                    self.logger.debug_tab(
                        f"Нормализация показателя {group_name} для месяца M-{month}: "
                        f"исходное значение={original_value}, нормализованное={normalized_value}, "
                        f"min={min_val}, max={max_val}, направление={direction}",
                        tab_number=DEBUG_TAB_NUMBER,
                        class_name="FileProcessor",
                        func_name="_normalize_group"
                    )
        
        return normalized_cols
    
    def _normalize_indicators(self, calculated_df: pd.DataFrame, config_manager) -> pd.DataFrame:
        """
        Нормализует показатели для каждого КМ по месяцам с учетом направления (вариант 3).
        
        Создает лист "Нормализация" с нормализованными значениями показателей.
        
        Args:
            calculated_df: DataFrame с расчетными данными
            config_manager: Менеджер конфигурации
        
        Returns:
            DataFrame с нормализованными значениями показателей
        """
        self.logger.info("=== Начало нормализации показателей (вариант 3) ===", "FileProcessor", "_normalize_indicators")
        
        # Базовые колонки
        base_columns = ["Табельный", "ТБ", "ФИО"]
        
        # Группируем колонки по месяцам и группам
        month_data = {}  # {month: {"OD": col_name, "RA": col_name, "PS": col_name}}
        
        for col in calculated_df.columns:
            if col in base_columns:
                continue
            
            match = re.search(r'^([A-Z]+)\s+\(M-(\d{1,2})\)', col)
            if match:
                group = match.group(1)
                month = int(match.group(2))
                
                if month not in month_data:
                    month_data[month] = {}
                month_data[month][group] = col
        
        # Создаем DataFrame для нормализованных данных
        # ВАЖНО: Убеждаемся, что базовые колонки существуют и не пустые
        if not all(col in calculated_df.columns for col in base_columns):
            missing_cols = [col for col in base_columns if col not in calculated_df.columns]
            self.logger.error(f"В calculated_df отсутствуют колонки: {missing_cols}. Доступные колонки: {list(calculated_df.columns)[:10]}", "FileProcessor", "_normalize_indicators")
            raise ValueError(f"Отсутствуют базовые колонки в calculated_df: {missing_cols}")
        
        # ВАЖНО: НЕ сбрасываем индекс, чтобы индексы совпадали с calculated_df при присваивании
        # Это критично для правильного присваивания нормализованных значений
        normalized_df = calculated_df[base_columns].copy()
        
        # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем, что базовые колонки скопированы правильно
        if not all(col in normalized_df.columns for col in base_columns):
            missing_cols = [col for col in base_columns if col not in normalized_df.columns]
            self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: В normalized_df отсутствуют базовые колонки после копирования: {missing_cols}. Доступные колонки: {list(normalized_df.columns)}", "FileProcessor", "_normalize_indicators")
            raise ValueError(f"Отсутствуют базовые колонки в normalized_df: {missing_cols}")
        else:
            self.logger.debug(f"Проверка normalized_df: все базовые колонки присутствуют после копирования", "FileProcessor", "_normalize_indicators")
        
        # ВАЖНО: Проверяем, что данные не пустые
        if len(normalized_df) > 0:
            sample_tb = normalized_df["ТБ"].iloc[0] if "ТБ" in normalized_df.columns else None
            sample_fio = normalized_df["ФИО"].iloc[0] if "ФИО" in normalized_df.columns else None
            self.logger.debug(f"normalized_df создан: {len(normalized_df)} строк x {len(normalized_df.columns)} колонок. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "_normalize_indicators")
            
            # Проверяем заполненность базовых колонок
            non_empty_tb = normalized_df["ТБ"].notna() & (normalized_df["ТБ"] != "") if "ТБ" in normalized_df.columns else pd.Series([False] * len(normalized_df))
            non_empty_fio = normalized_df["ФИО"].notna() & (normalized_df["ФИО"] != "") if "ФИО" in normalized_df.columns else pd.Series([False] * len(normalized_df))
            self.logger.debug(f"Заполненность базовых колонок в normalized_df: ТБ={non_empty_tb.sum()}/{len(normalized_df)}, ФИО={non_empty_fio.sum()}/{len(normalized_df)}", "FileProcessor", "_normalize_indicators")
        
        # Получаем направления для каждого показателя
        od_config = config_manager.get_group_config("OD").defaults if "OD" in config_manager.groups else None
        ra_config = config_manager.get_group_config("RA").defaults if "RA" in config_manager.groups else None
        ps_config = config_manager.get_group_config("PS").defaults if "PS" in config_manager.groups else None
        
        od_direction = od_config.indicator_direction if od_config else "MAX"
        ra_direction = ra_config.indicator_direction if ra_config else "MAX"
        ps_direction = ps_config.indicator_direction if ps_config else "MAX"
        
        # ОПТИМИЗАЦИЯ: Параллельная нормализация для всех групп (OD, RA, PS)
        # Нормализуем все группы параллельно
        self.logger.debug(f"Параллельная нормализация всех групп: OD, RA, PS (max_workers=3)", "FileProcessor", "_normalize_indicators")
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {
                executor.submit(self._normalize_group, group_name, direction, month_data, calculated_df): group_name
                for group_name, direction in [("OD", od_direction), ("RA", ra_direction), ("PS", ps_direction)]
            }
            
            # Обрабатываем результаты по мере завершения
            for future in as_completed(futures):
                group_name = futures[future]
                try:
                    normalized_cols = future.result()
                    # Добавляем нормализованные колонки в normalized_df
                    for norm_col_name, normalized in normalized_cols.items():
                        # ВАЖНО: Убеждаемся, что индексы совпадают при присваивании
                        normalized_df.loc[normalized.index, norm_col_name] = normalized
                except Exception as e:
                    self.logger.error(f"Ошибка при нормализации группы {group_name}: {str(e)}", "FileProcessor", "_normalize_indicators")
        
        # ВАЖНО: Сбрасываем индекс только в конце, после всех присваиваний
        normalized_df = normalized_df.reset_index(drop=True)
        
        # ВАЖНО: Финальная проверка перед возвратом
        if len(normalized_df) > 0:
            sample_tb = normalized_df["ТБ"].iloc[0] if "ТБ" in normalized_df.columns else None
            sample_fio = normalized_df["ФИО"].iloc[0] if "ФИО" in normalized_df.columns else None
            self.logger.debug(f"normalized_df финальный: {len(normalized_df)} строк. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "_normalize_indicators")
        
        self.logger.info(f"Нормализация завершена: {len(normalized_df)} строк, {len(normalized_df.columns)} колонок", "FileProcessor", "_normalize_indicators")
        return normalized_df
    
    def _normalize_with_direction(self, values: Dict[int, float], direction: str) -> Dict[int, float]:
        """
        Нормализует значения с учетом направления.
        
        Args:
            values: Словарь {month: value}
            direction: "MAX" или "MIN"
        
        Returns:
            Словарь {month: normalized_value} в диапазоне [0, 1]
        """
        if len(values) == 0:
            return {}
        
        if len(values) == 1:
            # Только один месяц - возвращаем 0.5 (среднее значение)
            return {month: 0.5 for month in values.keys()}
        
        min_val = min(values.values())
        max_val = max(values.values())
        
        if abs(max_val - min_val) < 1e-10:
            # Все значения одинаковы
            return {month: 0.5 for month in values.keys()}
        
        normalized = {}
        for month, value in values.items():
            if direction == "MAX":
                # Больше = лучше: нормализуем к [0, 1]
                normalized[month] = (value - min_val) / (max_val - min_val)
            else:  # direction == "MIN"
                # Меньше = лучше: инвертируем нормализацию
                normalized[month] = (max_val - value) / (max_val - min_val)
        
        return normalized
    
    def _calculate_score_for_month(self, month: int, normalized_df: pd.DataFrame, weight_od: float, weight_ra: float, weight_ps: float) -> Tuple[str, pd.Series]:
        """
        Рассчитывает Score для одного месяца.
        
        Args:
            month: Номер месяца
            normalized_df: DataFrame с нормализованными данными
            weight_od: Вес для OD
            weight_ra: Вес для RA
            weight_ps: Вес для PS
        
        Returns:
            Tuple[score_col_name, score_series] с именем колонки и значениями Score
        """
        score = pd.Series(0.0, index=normalized_df.index)
        
        # ОПТИМИЗАЦИЯ: Векторизованный расчет Score
        od_norm_col = f"OD_norm (M-{month})"
        ra_norm_col = f"RA_norm (M-{month})"
        ps_norm_col = f"PS_norm (M-{month})"
        
        if od_norm_col in normalized_df.columns:
            score += normalized_df[od_norm_col].fillna(0) * weight_od
        
        if ra_norm_col in normalized_df.columns:
            score += normalized_df[ra_norm_col].fillna(0) * weight_ra
        
        if ps_norm_col in normalized_df.columns:
            score += normalized_df[ps_norm_col].fillna(0) * weight_ps
        
        score_col_name = f"Score (M-{month})"
        
        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем расчет Score для указанного табельного
        if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0 and "Табельный" in normalized_df.columns:
            debug_mask = self._create_debug_tab_mask(normalized_df, "Табельный")
            if debug_mask.any():
                debug_idx = normalized_df[debug_mask].index[0]
                od_val = normalized_df.loc[debug_idx, od_norm_col] if od_norm_col in normalized_df.columns and debug_idx in normalized_df.index else 0
                ra_val = normalized_df.loc[debug_idx, ra_norm_col] if ra_norm_col in normalized_df.columns and debug_idx in normalized_df.index else 0
                ps_val = normalized_df.loc[debug_idx, ps_norm_col] if ps_norm_col in normalized_df.columns and debug_idx in normalized_df.index else 0
                score_val = score.loc[debug_idx] if debug_idx in score.index else 0
                
                self.logger.debug_tab(
                    f"Расчет Score для месяца M-{month}: "
                    f"OD_norm={od_val:.4f} × {weight_od} = {od_val * weight_od:.4f}, "
                    f"RA_norm={ra_val:.4f} × {weight_ra} = {ra_val * weight_ra:.4f}, "
                    f"PS_norm={ps_val:.4f} × {weight_ps} = {ps_val * weight_ps:.4f}, "
                    f"Итого Score={score_val:.4f}",
                    tab_number=None,  # Проверка уже сделана через debug_mask
                    class_name="FileProcessor",
                    func_name="_calculate_score_for_month"
                )
        
        return (score_col_name, score)
    
    def _calculate_best_month_variant3(self, calculated_df: pd.DataFrame, normalized_df: pd.DataFrame, config_manager, raw_df: Optional[pd.DataFrame] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Рассчитывает лучший месяц для каждого КМ на основе нормализованных значений (вариант 3).
        
        Создает листы "Места и выбор" и "Итог".
        
        Args:
            calculated_df: DataFrame с расчетными данными
            normalized_df: DataFrame с нормализованными данными
            config_manager: Менеджер конфигурации
            raw_df: DataFrame с сырыми данными (лист RAW) для подсчета уникальных ИНН (опционально)
        
        Returns:
            Tuple[DataFrame для "Места и выбор", DataFrame для "Итог"]
        """
        self.logger.info("=== Начало расчета лучшего месяца (вариант 3) ===", "FileProcessor", "_calculate_best_month_variant3")
        
        # Получаем веса и направления
        od_config = config_manager.get_group_config("OD").defaults if "OD" in config_manager.groups else None
        ra_config = config_manager.get_group_config("RA").defaults if "RA" in config_manager.groups else None
        ps_config = config_manager.get_group_config("PS").defaults if "PS" in config_manager.groups else None
        
        weight_od = od_config.weight if od_config else 0.33
        weight_ra = ra_config.weight if ra_config else 0.33
        weight_ps = ps_config.weight if ps_config else 0.34
        
        # Базовые колонки
        base_columns = ["Табельный", "ТБ", "ФИО"]
        
        # Группируем колонки по месяцам
        month_data = {}  # {month: {"OD": col_name, "RA": col_name, "PS": col_name}}
        
        for col in calculated_df.columns:
            if col in base_columns:
                continue
            
            match = re.search(r'^([A-Z]+)\s+\(M-(\d{1,2})\)', col)
            if match:
                group = match.group(1)
                month = int(match.group(2))
                
                if month not in month_data:
                    month_data[month] = {}
                month_data[month][group] = col
        
        # Создаем DataFrame для "Места и выбор"
        # ВАЖНО: Убеждаемся, что базовые колонки существуют
        if not all(col in normalized_df.columns for col in base_columns):
            missing_cols = [col for col in base_columns if col not in normalized_df.columns]
            self.logger.error(f"В normalized_df отсутствуют колонки: {missing_cols}. Доступные колонки: {list(normalized_df.columns)[:10]}", "FileProcessor", "_calculate_best_month_variant3")
            raise ValueError(f"Отсутствуют базовые колонки в normalized_df: {missing_cols}")
        
        # ВАЖНО: Сбрасываем индекс перед копированием, чтобы гарантировать совпадение строк
        places_df = normalized_df[base_columns].copy().reset_index(drop=True)
        
        # Создаем DataFrame для "Итог"
        # ВАЖНО: Убеждаемся, что базовые колонки существуют в calculated_df
        if not all(col in calculated_df.columns for col in base_columns):
            missing_cols = [col for col in base_columns if col not in calculated_df.columns]
            self.logger.error(f"В calculated_df отсутствуют колонки: {missing_cols}. Доступные колонки: {list(calculated_df.columns)[:10]}", "FileProcessor", "_calculate_best_month_variant3")
            raise ValueError(f"Отсутствуют базовые колонки в calculated_df: {missing_cols}")
        
        # ВАЖНО: Сбрасываем индекс перед копированием, чтобы гарантировать совпадение строк
        final_df = calculated_df[base_columns].copy().reset_index(drop=True)
        
        # Добавляем колонку с числом уникальных ИНН для каждого табельного номера (из RAW)
        if raw_df is not None and "Табельный" in raw_df.columns and "ИНН" in raw_df.columns:
            # ВАЖНО: Табельные номера в raw_df уже нормализованы (8 знаков с лидирующими нулями) в _process_file_for_raw
            # Подсчитываем количество уникальных ИНН для каждого табельного номера
            unique_inn_count = raw_df.groupby("Табельный")["ИНН"].nunique().to_dict()
            
            # Добавляем колонку в final_df (табельные номера в final_df тоже нормализованы)
            final_df["Количество уникальных ИНН"] = final_df["Табельный"].apply(
                lambda x: unique_inn_count.get(str(x), 0)
            )
            
            # Собираем данные для трекера
            for tab_num in self.debug_tracker.get_all_tab_numbers():
                # Пробуем найти в unique_inn_count с нормализацией
                tab_num_str = str(tab_num).strip()
                # Пробуем найти как есть (уже нормализован)
                count = unique_inn_count.get(tab_num_str, 0)
                if count == 0:
                    # Пробуем найти через сравнение без лидирующих нулей
                    tab_num_clean = tab_num_str.lstrip('0') if tab_num_str.lstrip('0') else '0'
                    for key, value in unique_inn_count.items():
                        key_clean = str(key).strip().lstrip('0') if str(key).strip().lstrip('0') else '0'
                        if key_clean == tab_num_clean:
                            count = value
                            break
                self.debug_tracker.set_unique_inn_count(tab_num_str, count)
            
            # Логируем статистику для диагностики
            non_zero_count = (final_df["Количество уникальных ИНН"] > 0).sum()
            total_count = len(final_df)
            sample_values = final_df["Количество уникальных ИНН"].head(5).tolist()
            self.logger.debug(f"Добавлена колонка 'Количество уникальных ИНН' в final_df: {non_zero_count}/{total_count} строк с ненулевыми значениями. Примеры значений: {sample_values}", "FileProcessor", "_calculate_best_month_variant3")
        else:
            # Если raw_df не передан или нет нужных колонок, заполняем нулями
            final_df["Количество уникальных ИНН"] = 0
            if raw_df is None:
                self.logger.debug(f"raw_df не передан, колонка 'Количество уникальных ИНН' заполнена нулями", "FileProcessor", "_calculate_best_month_variant3")
            else:
                available_cols = list(raw_df.columns) if raw_df is not None else []
                self.logger.warning(f"В raw_df отсутствуют колонки 'Табельный' или 'ИНН'. Доступные колонки: {available_cols[:10]}. Колонка 'Количество уникальных ИНН' заполнена нулями", "FileProcessor", "_calculate_best_month_variant3")
        
        # ВАЖНО: Проверяем, что данные не пустые
        if len(places_df) > 0:
            sample_tb = places_df["ТБ"].iloc[0] if "ТБ" in places_df.columns else None
            sample_fio = places_df["ФИО"].iloc[0] if "ФИО" in places_df.columns else None
            self.logger.debug(f"places_df создан: {len(places_df)} строк. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "_calculate_best_month_variant3")
        
        if len(final_df) > 0:
            sample_tb = final_df["ТБ"].iloc[0] if "ТБ" in final_df.columns else None
            sample_fio = final_df["ФИО"].iloc[0] if "ФИО" in final_df.columns else None
            self.logger.debug(f"final_df создан: {len(final_df)} строк. Пример: ТБ='{sample_tb}', ФИО='{sample_fio}'", "FileProcessor", "_calculate_best_month_variant3")
        
        # ОПТИМИЗАЦИЯ: Параллельный расчет Score для всех месяцев
        self.logger.debug(f"Параллельный расчет Score для всех месяцев (max_workers={MAX_WORKERS})", "FileProcessor", "_calculate_best_month_variant3")
        
        score_cols = {}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(self._calculate_score_for_month, month, normalized_df, weight_od, weight_ra, weight_ps): month
                for month in sorted(month_data.keys())
            }
            
            # Обрабатываем результаты по мере завершения
            for future in as_completed(futures):
                month = futures[future]
                try:
                    score_col_name, score = future.result()
                    places_df[score_col_name] = score
                    score_cols[month] = score_col_name
                except Exception as e:
                    self.logger.error(f"Ошибка при расчете Score для месяца M-{month}: {str(e)}", "FileProcessor", "_calculate_best_month_variant3")
        
        # ОПТИМИЗАЦИЯ: Векторизованный расчет горизонтального ранга
        # Создаем DataFrame со всеми Score для удобства работы
        score_df = pd.DataFrame(index=calculated_df.index)
        for month, col in score_cols.items():
            score_df[f"M-{month}"] = places_df[col]
        
        # Для каждого КМ рассчитываем ранг (горизонтально)
        # Используем rank с method='min' и ascending=False (больше = лучше)
        # na_option='keep' - NaN остаются NaN
        rank_df = score_df.rank(axis=1, method='min', ascending=False, na_option='keep')
        
        # Добавляем ранги в places_df
        for month in sorted(month_data.keys()):
            rank_col_name = f"Место (M-{month})"
            places_df[rank_col_name] = rank_df[f"M-{month}"].fillna(0).astype(int)
        
        # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем расчет рангов для указанного табельного
        if DEBUG_TAB_NUMBER and "Табельный" in calculated_df.columns:
            debug_mask = calculated_df["Табельный"].astype(str).str.strip().str.lstrip('0') == str(DEBUG_TAB_NUMBER).strip().lstrip('0')
            if debug_mask.any():
                debug_idx = calculated_df[debug_mask].index[0]
                ranks_info = {}
                scores_info = {}
                for month in sorted(month_data.keys()):
                    rank_val = rank_df.loc[debug_idx, f"M-{month}"] if debug_idx in rank_df.index and f"M-{month}" in rank_df.columns else None
                    score_val = score_df.loc[debug_idx, f"M-{month}"] if debug_idx in score_df.index and f"M-{month}" in score_df.columns else None
                    if rank_val is not None:
                        ranks_info[f"M-{month}"] = int(rank_val)
                    if score_val is not None:
                        scores_info[f"M-{month}"] = float(score_val)
                
                self.logger.debug_tab(
                    f"Расчет рангов (мест): Score по месяцам: {scores_info}, Места по месяцам: {ranks_info}",
                    tab_number=DEBUG_TAB_NUMBER,
                    class_name="FileProcessor",
                    func_name="_calculate_best_month_variant3"
                )
        
        # ОПТИМИЗАЦИЯ: Векторизованный поиск лучшего месяца
        # Находим все месяцы с рангом 1 для каждого КМ
        best_month_series = pd.Series("", index=calculated_df.index, dtype=str)
        
        # Создаем маску для месяцев с рангом 1 (заполняем NaN как False)
        rank_1_mask = (rank_df == 1).fillna(False)
        
        def get_month_values(month: int, idx: int) -> tuple:
            """Получает значения OD, RA, PS для указанного месяца и индекса."""
            od_val = None
            ra_val = None
            ps_val = None
            
            od_col = month_data[month].get("OD")
            ra_col = month_data[month].get("RA")
            ps_col = month_data[month].get("PS")
            
            if od_col and od_col in calculated_df.columns:
                od_val = calculated_df.loc[idx, od_col]
            if ra_col and ra_col in calculated_df.columns:
                ra_val = calculated_df.loc[idx, ra_col]
            if ps_col and ps_col in calculated_df.columns:
                ps_val = calculated_df.loc[idx, ps_col]
            
            return (od_val, ra_val, ps_val)
        
        def find_consecutive_groups(months: List[int]) -> List[List[int]]:
            """Находит группы подряд идущих месяцев."""
            if not months:
                return []
            
            sorted_months = sorted(months)
            groups = []
            current_group = [sorted_months[0]]
            
            for i in range(1, len(sorted_months)):
                if sorted_months[i] == sorted_months[i-1] + 1:
                    # Подряд идущий месяц
                    current_group.append(sorted_months[i])
                else:
                    # Разрыв в последовательности
                    groups.append(current_group)
                    current_group = [sorted_months[i]]
            
            groups.append(current_group)
            return groups
        
        # Для каждого КМ собираем месяцы с рангом 1 и обрабатываем их
        for idx in calculated_df.index:
            best_months = []
            for month in sorted(month_data.keys()):
                col_name = f"M-{month}"
                if col_name in rank_1_mask.columns:
                    if rank_1_mask.loc[idx, col_name]:
                        best_months.append(month)
            
            if not best_months:
                continue
            
            # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем процесс выбора лучшего месяца
            tab_number = calculated_df.loc[idx, "Табельный"] if "Табельный" in calculated_df.columns else None
            is_debug_tab = self.logger._is_debug_tab_number(tab_number) if tab_number is not None else False
            
            if is_debug_tab:
                self.logger.debug_tab(
                    f"Найдены месяцы с рангом 1: {best_months}",
                    tab_number=tab_number,
                    class_name="FileProcessor",
                    func_name="_calculate_best_month_variant3"
                )
            
            # Если только один месяц - просто добавляем его
            if len(best_months) == 1:
                best_month_series.loc[idx] = str(best_months[0])
                if is_debug_tab:
                    self.logger.debug_tab(
                        f"Выбран единственный лучший месяц: {best_months[0]}",
                        tab_number=tab_number,
                        class_name="FileProcessor",
                        func_name="_calculate_best_month_variant3"
                    )
                continue
            
            # Если несколько месяцев - проверяем значения и группируем
            # Создаем словарь: месяц -> (OD, RA, PS)
            month_values = {}
            for month in best_months:
                month_values[month] = get_month_values(month, idx)
            
            # Находим группы подряд идущих месяцев
            consecutive_groups = find_consecutive_groups(best_months)
            
            # Для каждой группы проверяем, одинаковые ли значения
            selected_months = []
            
            for group in consecutive_groups:
                if len(group) == 1:
                    # Один месяц - добавляем его
                    selected_months.append(group[0])
                else:
                    # Несколько месяцев - проверяем, одинаковые ли значения
                    first_month_values = month_values[group[0]]
                    all_same = True
                    
                    for month in group[1:]:
                        current_values = month_values[month]
                        # Сравниваем значения с учетом NaN и float (численное сравнение)
                        try:
                            # Проверяем OD
                            od_eq = (pd.isna(first_month_values[0]) and pd.isna(current_values[0])) or \
                                    (not pd.isna(first_month_values[0]) and not pd.isna(current_values[0]) and 
                                     abs(float(first_month_values[0]) - float(current_values[0])) < 1e-10)
                            # Проверяем RA
                            ra_eq = (pd.isna(first_month_values[1]) and pd.isna(current_values[1])) or \
                                    (not pd.isna(first_month_values[1]) and not pd.isna(current_values[1]) and 
                                     abs(float(first_month_values[1]) - float(current_values[1])) < 1e-10)
                            # Проверяем PS
                            ps_eq = (pd.isna(first_month_values[2]) and pd.isna(current_values[2])) or \
                                    (not pd.isna(first_month_values[2]) and not pd.isna(current_values[2]) and 
                                     abs(float(first_month_values[2]) - float(current_values[2])) < 1e-10)
                            
                            if not (od_eq and ra_eq and ps_eq):
                                all_same = False
                                break
                        except (ValueError, TypeError):
                            # Если не удалось сравнить как числа - считаем разными
                            all_same = False
                            break
                    
                    if all_same:
                        # Все значения одинаковые - берем только первый месяц
                        selected_months.append(group[0])
                    else:
                        # Значения разные - добавляем все месяцы
                        selected_months.extend(group)
            
            # Формируем строку с выбранными месяцами
            best_month_series.loc[idx] = ", ".join([str(m) for m in sorted(selected_months)])
            
            # ДЕТАЛЬНОЕ ЛОГИРОВАНИЕ: Логируем финальный выбор лучшего месяца
            if is_debug_tab:
                self.logger.debug_tab(
                    f"Финальный выбор лучшего месяца: {best_month_series.loc[idx]}. "
                    f"Исходные месяцы с рангом 1: {best_months}, "
                    f"Группы подряд идущих: {consecutive_groups}, "
                    f"Выбранные месяцы: {selected_months}",
                    tab_number=tab_number,
                    class_name="FileProcessor",
                    func_name="_calculate_best_month_variant3"
                )
        
        # Добавляем колонку "Лучший месяц" в places_df и final_df
        places_df["Лучший месяц"] = best_month_series
        final_df["Лучший месяц"] = best_month_series
        
        # Собираем данные для трекера: Score и лучший месяц
        if "Табельный" in places_df.columns:
            for tab_num in self.debug_tracker.get_all_tab_numbers():
                # Нормализуем табельный номер для поиска (используем ту же логику, что и в DataFrame)
                # В DataFrame табельные номера нормализованы через _normalize_tab_number (8 знаков с лидирующими нулями)
                tab_num_normalized = str(tab_num).strip()
                # Если табельный номер уже в формате 8 знаков, используем его как есть
                # Иначе пробуем найти через сравнение без лидирующих нулей
                tab_mask = places_df["Табельный"].astype(str).str.strip() == tab_num_normalized
                if not tab_mask.any():
                    # Пробуем найти через сравнение без лидирующих нулей
                    tab_num_clean = tab_num_normalized.lstrip('0') if tab_num_normalized.lstrip('0') else '0'
                    places_clean = places_df["Табельный"].astype(str).str.strip().str.lstrip('0')
                    places_clean = places_clean.apply(lambda x: x if x else '0')
                    tab_mask = places_clean == tab_num_clean
                
                if tab_mask.any():
                    tab_idx = places_df[tab_mask].index[0]
                    
                    # Собираем Score по месяцам
                    scores_dict = {}
                    for month in sorted(month_data.keys()):
                        score_col = score_cols.get(month)
                        if score_col and score_col in places_df.columns:
                            score_val = places_df.loc[tab_idx, score_col] if tab_idx in places_df.index else 0
                            scores_dict[str(month)] = float(score_val) if pd.notna(score_val) else 0
                    
                    best_month_val = best_month_series.loc[tab_idx] if tab_idx in best_month_series.index else ""
                    # Используем нормализованный номер из DataFrame для добавления в трекер
                    tab_num_from_df = str(places_df.loc[tab_idx, "Табельный"]).strip()
                    self.debug_tracker.add_scores(tab_num_from_df, scores_dict, str(best_month_val))
                    
                    # Логируем для диагностики
                    self.logger.debug(
                        f"Добавлены данные в трекер для табельного {tab_num_from_df}: "
                        f"scores={len(scores_dict)}, best_month={best_month_val}",
                        "FileProcessor",
                        "_calculate_best_month_variant3"
                    )
                    
                    # Собираем данные расчетов
                    if tab_idx in calculated_df.index:
                        calc_dict = {}
                        for month in sorted(month_data.keys()):
                            od_col = month_data[month].get("OD")
                            ra_col = month_data[month].get("RA")
                            ps_col = month_data[month].get("PS")
                            
                            fact = 0
                            if od_col and od_col in calculated_df.columns:
                                fact += float(calculated_df.loc[tab_idx, od_col]) if pd.notna(calculated_df.loc[tab_idx, od_col]) else 0
                            if ra_col and ra_col in calculated_df.columns:
                                fact += float(calculated_df.loc[tab_idx, ra_col]) if pd.notna(calculated_df.loc[tab_idx, ra_col]) else 0
                            if ps_col and ps_col in calculated_df.columns:
                                fact += float(calculated_df.loc[tab_idx, ps_col]) if pd.notna(calculated_df.loc[tab_idx, ps_col]) else 0
                            
                            calc_dict[str(month)] = {
                                "fact": fact,
                                "growth_2m": 0,  # Упрощенно, можно расширить
                                "growth_3m": 0  # Упрощенно, можно расширить
                            }
                        # Используем нормализованный номер из DataFrame для добавления в трекер
                        tab_num_from_df = str(places_df.loc[tab_idx, "Табельный"]).strip()
                        self.debug_tracker.add_calculations(tab_num_from_df, calc_dict)
                    
                    # Собираем данные нормализации
                    if tab_idx in normalized_df.index:
                        norm_dict = {}
                        for month in sorted(month_data.keys()):
                            od_norm_col = f"OD (M-{month})_norm" if f"OD (M-{month})_norm" in normalized_df.columns else None
                            ra_norm_col = f"RA (M-{month})_norm" if f"RA (M-{month})_norm" in normalized_df.columns else None
                            ps_norm_col = f"PS (M-{month})_norm" if f"PS (M-{month})_norm" in normalized_df.columns else None
                            
                            norm_dict[str(month)] = {
                                "OD": float(normalized_df.loc[tab_idx, od_norm_col]) if od_norm_col and pd.notna(normalized_df.loc[tab_idx, od_norm_col]) else 0,
                                "RA": float(normalized_df.loc[tab_idx, ra_norm_col]) if ra_norm_col and pd.notna(normalized_df.loc[tab_idx, ra_norm_col]) else 0,
                                "PS": float(normalized_df.loc[tab_idx, ps_norm_col]) if ps_norm_col and pd.notna(normalized_df.loc[tab_idx, ps_norm_col]) else 0
                            }
                        # Используем нормализованный номер из DataFrame для добавления в трекер
                        tab_num_from_df = str(places_df.loc[tab_idx, "Табельный"]).strip()
                        self.debug_tracker.add_normalization(tab_num_from_df, norm_dict)
                        
                        # Логируем для диагностики
                        self.logger.debug(
                            f"Добавлены данные нормализации в трекер для табельного {tab_num_from_df}: "
                            f"месяцев={len(norm_dict)}",
                            "FileProcessor",
                            "_calculate_best_month_variant3"
                        )
        
        self.logger.info(f"Расчет лучшего месяца завершен: определен для {len(best_month_series[best_month_series != ''])} КМ", "FileProcessor", "_calculate_best_month_variant3")
        
        return places_df, final_df
    
    def prepare_statistics_sheet(self) -> Optional[pd.DataFrame]:
        """
        Формирует лист со статистикой обработки данных.
        
        Returns:
            Optional[pd.DataFrame]: DataFrame со статистикой или None, если статистика отключена
        """
        if not ENABLE_STATISTICS:
            return None
        
        self.logger.info("=== Начало формирования листа 'Статистика' ===", "FileProcessor", "prepare_statistics_sheet")
        
        # Создаем список всех таблиц статистики
        statistics_tables = []
        
        # Таблица 1: Общая статистика
        summary_data = []
        summary_data.append(["Параметр", "Значение"])
        summary_data.append(["Всего обработано КМ (табельных номеров)", self.statistics["summary"].get("total_km", 0)])
        summary_data.append(["Всего уникальных клиентов", self.statistics["summary"].get("total_clients", 0)])
        summary_data.append(["", ""])  # Пустая строка для разделения
        
        # Таблица 2: Количество КМ по ТБ
        if "by_tb" in self.statistics["summary"]:
            summary_data.append(["Количество КМ по ТБ", ""])
            summary_data.append(["ТБ", "Количество КМ"])
            for tb, count in sorted(self.statistics["summary"]["by_tb"].items(), key=lambda x: x[1], reverse=True):
                summary_data.append([tb, count])
            summary_data.append(["", ""])  # Пустая строка для разделения
        
        # Таблица 3: Статистика обработки файлов (разделена по группам OD, RA, PS)
        # Функция для извлечения номера месяца из имени файла
        def extract_month_number(file_name: str) -> int:
            """Извлекает номер месяца из имени файла (M-1, M-2, ..., M-12)."""
            match = re.search(r'M-(\d{1,2})_', file_name)
            if match:
                month = int(match.group(1))
                if 1 <= month <= 12:
                    return month
            return 0
        
        # Создаем развернутые таблицы для каждой группы
        for group in ["OD", "RA", "PS"]:
            if group not in self.statistics["files"]:
                continue
            
            summary_data.append([f"Статистика обработки файлов - {group}", ""])
            
            # Собираем данные по месяцам
            months = list(range(1, 13))  # M-1 до M-12
            month_files = {}  # {month: file_name}
            file_data = {}  # {file_name: {initial, dropped, kept, final, drop_rules: {}, in_rules: {}}}
            
            for file_name in sorted(self.statistics["files"][group].keys()):
                month = extract_month_number(file_name)
                if month > 0:
                    month_files[month] = file_name
                    file_stats = self.statistics["files"][group][file_name]
                    file_data[file_name] = {
                        "initial": file_stats.get("initial_rows", 0),
                        "final": file_stats.get("final_rows", 0),
                        "dropped": sum(file_stats.get("dropped_by_rule", {}).values()),
                        "kept": sum(file_stats.get("kept_by_rule", {}).values()),
                        "drop_rules": file_stats.get("dropped_by_rule", {}),
                        "in_rules": file_stats.get("kept_by_rule", {})
                    }
            
            # Создаем заголовки: строка данных, M-1, M-2, ..., M-12
            header = ["Параметр"] + [f"M-{m}" for m in months]
            summary_data.append(header)
            
            # Строка 1: Исходно строк
            row = ["Исходно строк"]
            for m in months:
                file_name = month_files.get(m, "")
                if file_name:
                    row.append(file_data[file_name]["initial"])
                else:
                    row.append("")
            summary_data.append(row)
            
            # Строка 2: Удалено по drop_rules (всего)
            row = ["Удалено по drop_rules (всего)"]
            for m in months:
                file_name = month_files.get(m, "")
                if file_name:
                    row.append(file_data[file_name]["dropped"])
                else:
                    row.append("")
            summary_data.append(row)
            
            # Строка 3: Оставлено по in_rules (всего)
            row = ["Оставлено по in_rules (всего)"]
            for m in months:
                file_name = month_files.get(m, "")
                if file_name:
                    row.append(file_data[file_name]["kept"])
                else:
                    row.append("")
            summary_data.append(row)
            
            # Строка 4: Итогово строк
            row = ["Итогово строк"]
            for m in months:
                file_name = month_files.get(m, "")
                if file_name:
                    row.append(file_data[file_name]["final"])
                else:
                    row.append("")
            summary_data.append(row)
            
            # Детальная статистика по drop_rules
            # Собираем все уникальные правила drop_rules
            all_drop_rules = set()
            for file_name in file_data.keys():
                all_drop_rules.update(file_data[file_name]["drop_rules"].keys())
            
            if all_drop_rules:
                summary_data.append(["", ""])  # Пустая строка
                summary_data.append(["Детальная статистика по drop_rules", ""])
                summary_data.append(header)
                
                for rule in sorted(all_drop_rules):
                    row = [f"Удалено: {rule}"]
                    for m in months:
                        file_name = month_files.get(m, "")
                        if file_name and rule in file_data[file_name]["drop_rules"]:
                            row.append(file_data[file_name]["drop_rules"][rule])
                        else:
                            row.append("")
                    summary_data.append(row)
            
            # Детальная статистика по in_rules
            # Собираем все уникальные правила in_rules
            all_in_rules = set()
            for file_name in file_data.keys():
                all_in_rules.update(file_data[file_name]["in_rules"].keys())
            
            if all_in_rules:
                summary_data.append(["", ""])  # Пустая строка
                summary_data.append(["Детальная статистика по in_rules", ""])
                summary_data.append(header)
                
                for rule in sorted(all_in_rules):
                    row = [f"Оставлено: {rule}"]
                    for m in months:
                        file_name = month_files.get(m, "")
                        if file_name and rule in file_data[file_name]["in_rules"]:
                            row.append(file_data[file_name]["in_rules"][rule])
                        else:
                            row.append("")
                    summary_data.append(row)
            
            summary_data.append(["", ""])  # Пустая строка для разделения
        
        # Таблица 4: Статистика выбора табельных номеров (разделена по группам)
        for group in ["OD", "RA", "PS"]:
            if group not in self.statistics["tab_selection"]:
                continue
            
            summary_data.append([f"Статистика выбора табельных номеров - {group}", ""])
            
            # Собираем данные по месяцам
            months = list(range(1, 13))  # M-1 до M-12
            month_files = {}  # {month: file_name}
            tab_data = {}  # {file_name: {total_variants, selected_count, variants_with_multiple}}
            
            for file_name in sorted(self.statistics["tab_selection"][group].keys()):
                month = extract_month_number(file_name)
                if month > 0:
                    month_files[month] = file_name
                    tab_stats = self.statistics["tab_selection"][group][file_name]
                    tab_data[file_name] = {
                        "total_variants": tab_stats.get("total_variants", 0),
                        "selected_count": tab_stats.get("selected_count", 0),
                        "variants_with_multiple": tab_stats.get("variants_with_multiple", 0)
                    }
            
            # Создаем заголовки: строка данных, M-1, M-2, ..., M-12
            header = ["Параметр"] + [f"M-{m}" for m in months]
            summary_data.append(header)
            
            # Строка 1: Всего вариантов ТБ
            row = ["Всего вариантов ТБ"]
            for m in months:
                file_name = month_files.get(m, "")
                if file_name:
                    row.append(tab_data[file_name]["total_variants"])
                else:
                    row.append("")
            summary_data.append(row)
            
            # Строка 2: Выбрано уникальных
            row = ["Выбрано уникальных"]
            for m in months:
                file_name = month_files.get(m, "")
                if file_name:
                    row.append(tab_data[file_name]["selected_count"])
                else:
                    row.append("")
            summary_data.append(row)
            
            # Строка 3: Табельных с несколькими вариантами
            row = ["Табельных с несколькими вариантами"]
            for m in months:
                file_name = month_files.get(m, "")
                if file_name:
                    row.append(tab_data[file_name]["variants_with_multiple"])
                else:
                    row.append("")
            summary_data.append(row)
            
            summary_data.append(["", ""])  # Пустая строка для разделения
        
        # Создаем DataFrame
        if len(summary_data) > 0:
            # Находим максимальную длину строки для определения количества колонок
            max_cols = max(len(row) for row in summary_data) if summary_data else 2
            
            # Дополняем все строки до максимальной длины
            summary_data_padded = [row + [""] * (max_cols - len(row)) for row in summary_data]
            
            statistics_df = pd.DataFrame(summary_data_padded)
        else:
            statistics_df = pd.DataFrame([["Статистика недоступна", ""]])
        
        self.logger.info(f"Лист 'Статистика': Подготовлено {len(statistics_df)} строк статистики", "FileProcessor", "prepare_statistics_sheet")
        self.logger.info("=== Завершена подготовка листа 'Статистика' ===", "FileProcessor", "prepare_statistics_sheet")
        
        # Выводим статистику в лог
        self._log_statistics()
        
        return statistics_df
    
    def _log_statistics(self) -> None:
        """Выводит статистику в лог."""
        if not ENABLE_STATISTICS:
            return
        
        self.logger.info("=" * 80, "FileProcessor", "_log_statistics")
        self.logger.info("СТАТИСТИКА ОБРАБОТКИ ДАННЫХ", "FileProcessor", "_log_statistics")
        self.logger.info("=" * 80, "FileProcessor", "_log_statistics")
        
        # Общая статистика
        self.logger.info(f"Всего обработано КМ (табельных номеров): {self.statistics['summary'].get('total_km', 0)}", "FileProcessor", "_log_statistics")
        self.logger.info(f"Всего уникальных клиентов: {self.statistics['summary'].get('total_clients', 0)}", "FileProcessor", "_log_statistics")
        
        # Статистика по ТБ
        if "by_tb" in self.statistics["summary"]:
            self.logger.info("Количество КМ по ТБ:", "FileProcessor", "_log_statistics")
            for tb, count in sorted(self.statistics["summary"]["by_tb"].items(), key=lambda x: x[1], reverse=True):
                self.logger.info(f"  {tb}: {count}", "FileProcessor", "_log_statistics")
        
        # Статистика по файлам
        total_initial = 0
        total_dropped = 0
        total_final = 0
        
        for group in sorted(self.statistics["files"].keys()):
            self.logger.info(f"Группа {group}:", "FileProcessor", "_log_statistics")
            for file_name in sorted(self.statistics["files"][group].keys()):
                file_stats = self.statistics["files"][group][file_name]
                initial = file_stats.get("initial_rows", 0)
                final = file_stats.get("final_rows", 0)
                dropped_count = sum(file_stats.get("dropped_by_rule", {}).values())
                
                self.logger.info(f"  {file_name}: исходно {initial}, удалено {dropped_count}, итого {final}", "FileProcessor", "_log_statistics")
                
                total_initial += initial
                total_dropped += dropped_count
                total_final += final
        
        self.logger.info(f"ИТОГО: исходно {total_initial}, удалено {total_dropped}, итого {total_final}", "FileProcessor", "_log_statistics")
        
        # Статистика выбора табельных
        for group in sorted(self.statistics["tab_selection"].keys()):
            self.logger.info(f"Выбор табельных номеров - группа {group}:", "FileProcessor", "_log_statistics")
            for file_name in sorted(self.statistics["tab_selection"][group].keys()):
                tab_stats = self.statistics["tab_selection"][group][file_name]
                self.logger.info(
                    f"  {file_name}: всего вариантов {tab_stats.get('total_variants', 0)}, "
                    f"выбрано {tab_stats.get('selected_count', 0)}, "
                    f"с несколькими вариантами {tab_stats.get('variants_with_multiple', 0)}",
                    "FileProcessor", "_log_statistics"
                )
        
        self.logger.info("=" * 80, "FileProcessor", "_log_statistics")


# ============================================================================
# МОДУЛЬ ФОРМАТИРОВАНИЯ EXCEL
# ============================================================================

class ExcelFormatter:
    """Класс для форматирования Excel файлов с использованием только базовых модулей Anaconda."""
    
    def __init__(self, logger_instance: Optional[Logger] = None):
        """
        Инициализация форматтера.
        
        Args:
            logger_instance: Экземпляр логгера
        """
        self.min_width = 15
        self.max_width = 150
        self.logger = logger_instance
    
    def _calculate_column_width(self, df: pd.DataFrame, col_name: str) -> float:
        """
        Вычисляет оптимальную ширину колонки на основе содержимого.
        
        Args:
            df: DataFrame с данными
            col_name: Название колонки
            
        Returns:
            float: Ширина колонки
        """
        if col_name not in df.columns:
            return self.min_width
        
        # Максимальная длина в заголовке
        max_length = len(str(col_name))
        
        # Максимальная длина в данных (первые 100 строк для производительности)
        sample_size = min(100, len(df))
        for idx in range(sample_size):
            value = df[col_name].iloc[idx]
            if pd.notna(value):
                max_length = max(max_length, len(str(value)))
        
        # Применяем ограничения
        width = max(self.min_width, min(max_length + 2, self.max_width))
        return width
    
    def create_formatted_excel(self, raw_df: pd.DataFrame, summary_df: pd.DataFrame, calculated_df: pd.DataFrame, 
                              normalized_df: pd.DataFrame, places_df: pd.DataFrame, final_df: pd.DataFrame,
                              output_path: str, statistics_df: Optional[pd.DataFrame] = None, 
                              debug_tracker: Optional[DebugTabNumberTracker] = None) -> None:
        """
        Создает новый Excel файл с форматированием используя только базовые модули Anaconda.
        Используется только openpyxl
        
        Создает 6 основных листов + лист "Статистика" (если включен):
        1. "RAW" - сырые данные после фильтрации (уникальные комбинации ТН+ФИО+ТБ+ИНН с суммами по файлам)
        2. "Исходник" - исходные отфильтрованные данные
        3. "Расчет" - расчетные данные (факт, прирост по 2м, прирост по 3м)
        4. "Нормализация" - нормализованные значения показателей
        5. "Места и выбор" - Score, ранги и лучший месяц
        6. "Итог" - итоговые данные с выбором месяца и значениями показателей
        7. "Статистика" - статистика обработки (если ENABLE_STATISTICS = True и statistics_df не None)
        
        Args:
            raw_df: DataFrame с сырыми данными (лист "RAW")
            summary_df: DataFrame с исходными данными (лист "Исходник")
            calculated_df: DataFrame с расчетными данными (лист "Расчет")
            normalized_df: DataFrame с нормализованными данными (лист "Нормализация")
            places_df: DataFrame с Score и рангами (лист "Места и выбор")
            final_df: DataFrame с итоговыми данными (лист "Итог")
            output_path: Путь для сохранения файла
            statistics_df: DataFrame со статистикой (лист "Статистика") или None, если статистика отключена
        """
        self.logger.info(f"Создание форматированного Excel файла {output_path}")
        
        # РАСШИРЕННОЕ ЛОГИРОВАНИЕ: Проверяем наличие базовых колонок во всех DataFrame перед сохранением
        base_columns = ["Табельный", "ТБ", "ФИО"]
        
        for df_name, df in [("summary_df (Исходник)", summary_df), 
                            ("calculated_df (Расчет)", calculated_df),
                            ("normalized_df (Нормализация)", normalized_df),
                            ("places_df (Места и выбор)", places_df),
                            ("final_df (Итог)", final_df)]:
            if df is not None and len(df) > 0:
                missing_cols = [col for col in base_columns if col not in df.columns]
                if missing_cols:
                    self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: В {df_name} отсутствуют базовые колонки: {missing_cols}. Доступные колонки: {list(df.columns)[:20]}", "ExcelFormatter", "create_formatted_excel")
                else:
                    self.logger.debug(f"Проверка {df_name}: все базовые колонки присутствуют. Размер: {len(df)} строк x {len(df.columns)} колонок", "ExcelFormatter", "create_formatted_excel")
                    # Проверяем заполненность
                    if "ТБ" in df.columns:
                        non_empty = df["ТБ"].notna() & (df["ТБ"] != "")
                        self.logger.debug(f"Заполненность ТБ в {df_name}: {non_empty.sum()}/{len(df)} строк", "ExcelFormatter", "create_formatted_excel")
        
        try:
            if OPENPYXL_AVAILABLE:
                # Используем openpyxl для форматирования
                self._create_with_openpyxl(raw_df, summary_df, calculated_df, normalized_df, places_df, final_df, output_path, statistics_df, debug_tracker)
            else:
                # Используем pandas ExcelWriter без форматирования
                self.logger.warning("openpyxl недоступен, создается файл без форматирования", "ExcelFormatter", "create_formatted_excel")
                # Разбиваем raw_df на чанки (если больше 900 000 строк)
                raw_chunks = self._split_raw_df(raw_df, chunk_size=900_000)
                # Пробуем использовать доступный engine
                try:
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        # Сохраняем все чанки RAW
                        for sheet_name, chunk_df in raw_chunks:
                            if len(chunk_df) > 0:
                                chunk_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        summary_df.to_excel(writer, sheet_name="Исходник", index=False)
                        calculated_df.to_excel(writer, sheet_name="Расчет", index=False)
                        normalized_df.to_excel(writer, sheet_name="Нормализация", index=False)
                        places_df.to_excel(writer, sheet_name="Места и выбор", index=False)
                        final_df.to_excel(writer, sheet_name="Итог", index=False)
                        if statistics_df is not None:
                            statistics_df.to_excel(writer, sheet_name="Статистика", index=False, header=False)
                except Exception as e:
                    try:
                        with pd.ExcelWriter(output_path) as writer:
                            # Сохраняем все чанки RAW
                            for sheet_name, chunk_df in raw_chunks:
                                if len(chunk_df) > 0:
                                    chunk_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            summary_df.to_excel(writer, sheet_name="Исходник", index=False)
                            calculated_df.to_excel(writer, sheet_name="Расчет", index=False)
                            normalized_df.to_excel(writer, sheet_name="Нормализация", index=False)
                            places_df.to_excel(writer, sheet_name="Места и выбор", index=False)
                            final_df.to_excel(writer, sheet_name="Итог", index=False)
                            if statistics_df is not None:
                                statistics_df.to_excel(writer, sheet_name="Статистика", index=False, header=False)
                    except Exception:
                        # Если не получилось, используем любой доступный engine
                        with pd.ExcelWriter(output_path) as writer:
                            # Сохраняем все чанки RAW
                            for sheet_name, chunk_df in raw_chunks:
                                if len(chunk_df) > 0:
                                    chunk_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            summary_df.to_excel(writer, sheet_name="Исходник", index=False)
                            calculated_df.to_excel(writer, sheet_name="Расчет", index=False)
                            normalized_df.to_excel(writer, sheet_name="Нормализация", index=False)
                            places_df.to_excel(writer, sheet_name="Места и выбор", index=False)
                            final_df.to_excel(writer, sheet_name="Итог", index=False)
                            if statistics_df is not None:
                                statistics_df.to_excel(writer, sheet_name="Статистика", index=False, header=False)
                self.logger.info(f"Файл {output_path} создан без форматирования", "ExcelFormatter", "create_formatted_excel")
            
        except Exception as e:
            self.logger.error(f"Ошибка при создании Excel файла {output_path}: {str(e)}", "ExcelFormatter", "create_formatted_excel")
            # Разбиваем raw_df на чанки (если больше 900 000 строк) для fallback режима
            raw_chunks = self._split_raw_df(raw_df, chunk_size=900_000)
            # Пробуем создать без форматирования
            try:
                with pd.ExcelWriter(output_path) as writer:
                    # Сохраняем все чанки RAW
                    for sheet_name, chunk_df in raw_chunks:
                        if len(chunk_df) > 0:
                            chunk_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    summary_df.to_excel(writer, sheet_name="Исходник", index=False)
                    calculated_df.to_excel(writer, sheet_name="Расчет", index=False)
                    normalized_df.to_excel(writer, sheet_name="Нормализация", index=False)
                    places_df.to_excel(writer, sheet_name="Места и выбор", index=False)
                    final_df.to_excel(writer, sheet_name="Итог", index=False)
                    if statistics_df is not None:
                        statistics_df.to_excel(writer, sheet_name="Статистика", index=False, header=False)
                self.logger.warning(f"Файл создан без форматирования из-за ошибки: {str(e)}", "ExcelFormatter", "create_formatted_excel")
            except Exception as e2:
                self.logger.error(f"Критическая ошибка при создании файла: {str(e2)}", "ExcelFormatter", "create_formatted_excel")
                raise
    
    def _split_raw_df(self, raw_df: pd.DataFrame, chunk_size: int = 900_000) -> list[tuple[str, pd.DataFrame]]:
        """
        Разбивает raw_df на несколько чанков для сохранения в отдельные листы Excel.
        
        Args:
            raw_df: DataFrame с сырыми данными
            chunk_size: Максимальный размер чанка (по умолчанию 900 000 строк)
        
        Returns:
            list[tuple[str, pd.DataFrame]]: Список кортежей (имя_листа, DataFrame_чанка)
        """
        if len(raw_df) == 0:
            return []
        
        chunks = []
        total_rows = len(raw_df)
        
        if total_rows <= chunk_size:
            # Если данных меньше или равно chunk_size, создаем один лист RAW
            chunks.append(("RAW", raw_df))
        else:
            # Разбиваем на несколько листов: RAW, RAW_2, RAW_3, ...
            num_chunks = (total_rows + chunk_size - 1) // chunk_size  # Округление вверх
            
            self.logger.info(
                f"Лист RAW слишком большой ({total_rows} строк), будет разбит на {num_chunks} листа(ов) "
                f"(по {chunk_size} строк в каждом)",
                "ExcelFormatter",
                "_split_raw_df"
            )
            
            for i in range(num_chunks):
                start_idx = i * chunk_size
                end_idx = min((i + 1) * chunk_size, total_rows)
                chunk_df = raw_df.iloc[start_idx:end_idx].copy()
                
                if i == 0:
                    sheet_name = "RAW"
                else:
                    sheet_name = f"RAW_{i + 1}"
                
                chunks.append((sheet_name, chunk_df))
                self.logger.debug(
                    f"Создан чанк {sheet_name}: строки {start_idx + 1}-{end_idx} (всего {len(chunk_df)} строк)",
                    "ExcelFormatter",
                    "_split_raw_df"
                )
        
        return chunks
    
    def _create_debug_tab_sheets(self, debug_tracker: DebugTabNumberTracker, writer: pd.ExcelWriter) -> None:
        """
        Создает детальные листы Excel для табельных номеров из DEBUG_TAB_NUMBER.
        
        Args:
            debug_tracker: Трекер с собранными данными
            writer: ExcelWriter для записи листов
        """
        self.logger.info("=== НАЧАЛО СОЗДАНИЯ ДЕТАЛЬНЫХ ЛИСТОВ ===", "ExcelFormatter", "_create_debug_tab_sheets")
        
        if not debug_tracker:
            self.logger.error("debug_tracker не передан в _create_debug_tab_sheets", "ExcelFormatter", "_create_debug_tab_sheets")
            return
        
        all_tab_numbers = debug_tracker.get_all_tab_numbers()
        self.logger.info(f"Получены табельные номера из трекера: {all_tab_numbers} (всего {len(all_tab_numbers)})", "ExcelFormatter", "_create_debug_tab_sheets")
        self.logger.info(f"Все ключи в tab_data: {list(debug_tracker.tab_data.keys())}", "ExcelFormatter", "_create_debug_tab_sheets")
        
        if len(all_tab_numbers) == 0:
            self.logger.error("В debug_tracker нет табельных номеров для создания детальных листов. Проверьте, что DEBUG_TAB_NUMBER указан и данные собираются в трекер.", "ExcelFormatter", "_create_debug_tab_sheets")
            return
        
        self.logger.info(f"Создание детальных листов для {len(all_tab_numbers)} табельных номеров: {all_tab_numbers}", "ExcelFormatter", "_create_debug_tab_sheets")
        
        for tab_number in all_tab_numbers:
            self.logger.info(f"Обработка табельного номера: {tab_number} (тип: {type(tab_number)})", "ExcelFormatter", "_create_debug_tab_sheets")
            
            # Пробуем получить данные разными способами
            tab_data = debug_tracker.get_tab_data(tab_number)
            if not tab_data:
                # Пробуем найти через прямой доступ к tab_data
                self.logger.warning(f"get_tab_data вернул None для {tab_number}. Пробуем прямой доступ...", "ExcelFormatter", "_create_debug_tab_sheets")
                self.logger.info(f"Доступные ключи в tab_data: {list(debug_tracker.tab_data.keys())}", "ExcelFormatter", "_create_debug_tab_sheets")
                
                # Пробуем найти через нормализацию
                tab_num_str = str(tab_number).strip()
                tab_num_clean = tab_num_str.lstrip('0') if tab_num_str.lstrip('0') else '0'
                tab_num_normalized = tab_num_clean.zfill(8)
                
                if tab_num_normalized in debug_tracker.tab_data:
                    tab_data = debug_tracker.tab_data[tab_num_normalized]
                    self.logger.info(f"Найдены данные по нормализованному ключу: {tab_num_normalized}", "ExcelFormatter", "_create_debug_tab_sheets")
                elif tab_num_str in debug_tracker.tab_data:
                    tab_data = debug_tracker.tab_data[tab_num_str]
                    self.logger.info(f"Найдены данные по оригинальному ключу: {tab_num_str}", "ExcelFormatter", "_create_debug_tab_sheets")
                else:
                    self.logger.error(
                        f"Нет данных для табельного номера {tab_number} в debug_tracker. "
                        f"Доступные ключи: {list(debug_tracker.tab_data.keys())}. "
                        f"Пробовали: нормализованный={tab_num_normalized}, оригинальный={tab_num_str}",
                        "ExcelFormatter",
                        "_create_debug_tab_sheets"
                    )
                    continue
            
            source_files_count = len(tab_data.get('source_files', {}))
            raw_data_count = len(tab_data.get('raw_data', {}))
            calculations_count = len(tab_data.get('calculations', {}))
            normalization_count = len(tab_data.get('normalization', {}))
            scores_count = len(tab_data.get('scores', {}))
            
            self.logger.info(
                f"Обработка табельного номера {tab_number}: "
                f"source_files={source_files_count}, raw_data={raw_data_count}, "
                f"calculations={calculations_count}, normalization={normalization_count}, scores={scores_count}",
                "ExcelFormatter",
                "_create_debug_tab_sheets"
            )
            
            # Создаем лист для каждого табельного номера
            sheet_name = f"Детально_{tab_number}"
            # Ограничиваем длину имени листа (Excel ограничение - 31 символ)
            if len(sheet_name) > 31:
                sheet_name = f"Дет_{tab_number[-8:]}"
            
            # Создаем список таблиц для листа
            tables_data = []
            
            # Таблица 1: Данные из исходных файлов
            if "source_files" in tab_data and tab_data["source_files"]:
                source_rows = []
                source_rows.append(["Файл", "Группа", "Месяц", "ТБ", "ФИО", "ИНН", "Показатель", "Выбран"])
                
                for file_name, file_data in sorted(tab_data["source_files"].items()):
                    group = file_data.get("group", "")
                    month = file_data.get("month", 0)
                    selected_tb = file_data.get("selected_tb", "")
                    selected_sum = file_data.get("selected_sum", 0)
                    
                    # Добавляем строку с выбранным вариантом
                    source_rows.append([
                        file_name,
                        group,
                        f"M-{month}",
                        selected_tb,
                        "",  # ФИО будет в отдельных строках
                        "",  # ИНН будет в отдельных строках
                        selected_sum,
                        "Да"
                    ])
                    
                    # Добавляем строки с клиентами
                    clients = file_data.get("clients", [])
                    for client in clients:
                        source_rows.append([
                            "",  # Файл уже указан выше
                            "",  # Группа уже указана выше
                            "",  # Месяц уже указан выше
                            client.get("ТБ", ""),
                            client.get("ФИО", ""),
                            client.get("ИНН", ""),
                            client.get("Показатель", 0),
                            "Да" if client.get("Выбран", False) else "Нет"
                        ])
                    
                    # Добавляем строку с вариантами ТБ
                    tb_variants = file_data.get("tb_variants", {})
                    if len(tb_variants) > 1:
                        source_rows.append(["", "", "", "Варианты ТБ:", "", "", "", ""])
                        for tb, sum_val in sorted(tb_variants.items(), key=lambda x: x[1], reverse=True):
                            source_rows.append([
                                "", "", "",
                                tb,
                                "", "", "",
                                f"{sum_val:,.2f}".replace(",", " ").replace(".", ",")
                            ])
                    
                    # Пустая строка между файлами
                    source_rows.append([""] * 8)
                
                if len(source_rows) > 1:  # Если есть данные кроме заголовка
                    source_df = pd.DataFrame(source_rows[1:], columns=source_rows[0])
                    tables_data.append(("Исходные файлы", source_df))
            
            # Таблица 2: Данные после схлопывания (RAW)
            if "raw_data" in tab_data and tab_data["raw_data"]:
                raw_rows = []
                raw_rows.append(["ИНН", "ТБ", "ФИО"] + sorted(set(
                    file_key
                    for inn_data in tab_data["raw_data"].values()
                    for file_key in inn_data.get("sums_by_file", {}).keys()
                )))
                
                for inn, inn_data in sorted(tab_data["raw_data"].items()):
                    row = [
                        inn,
                        inn_data.get("ТБ", ""),
                        inn_data.get("ФИО", "")
                    ]
                    sums_by_file = inn_data.get("sums_by_file", {})
                    for file_key in raw_rows[0][3:]:  # Все колонки кроме первых трех
                        row.append(sums_by_file.get(file_key, 0))
                    raw_rows.append(row)
                
                if len(raw_rows) > 1:
                    raw_df_sheet = pd.DataFrame(raw_rows[1:], columns=raw_rows[0])
                    tables_data.append(("Данные после схлопывания (RAW)", raw_df_sheet))
            
            # Таблица 3: Результаты расчетов
            if "calculations" in tab_data and tab_data["calculations"]:
                calc_rows = []
                calc_rows.append(["Месяц", "Факт", "Прирост 2м", "Прирост 3м"])
                
                for month, calc_data in sorted(tab_data["calculations"].items()):
                    calc_rows.append([
                        f"M-{month}",
                        calc_data.get("fact", 0),
                        calc_data.get("growth_2m", 0),
                        calc_data.get("growth_3m", 0)
                    ])
                
                if len(calc_rows) > 1:
                    calc_df = pd.DataFrame(calc_rows[1:], columns=calc_rows[0])
                    tables_data.append(("Результаты расчетов", calc_df))
            
            # Таблица 4: Нормализация
            if "normalization" in tab_data and tab_data["normalization"]:
                norm_rows = []
                norm_rows.append(["Месяц", "OD_norm", "RA_norm", "PS_norm"])
                
                for month, norm_data in sorted(tab_data["normalization"].items()):
                    norm_rows.append([
                        f"M-{month}",
                        norm_data.get("OD", 0),
                        norm_data.get("RA", 0),
                        norm_data.get("PS", 0)
                    ])
                
                if len(norm_rows) > 1:
                    norm_df = pd.DataFrame(norm_rows[1:], columns=norm_rows[0])
                    tables_data.append(("Нормализация", norm_df))
            
            # Таблица 5: Score и лучший месяц
            if "scores" in tab_data and tab_data["scores"]:
                score_rows = []
                score_rows.append(["Месяц", "Score", "Лучший месяц"])
                
                best_month = tab_data.get("best_month", "")
                for month, score_val in sorted(tab_data["scores"].items()):
                    score_rows.append([
                        f"M-{month}",
                        score_val,
                        "Да" if str(month) in best_month.split(", ") else "Нет"
                    ])
                
                if len(score_rows) > 1:
                    score_df = pd.DataFrame(score_rows[1:], columns=score_rows[0])
                    tables_data.append(("Score и выбор месяца", score_df))
            
            # Таблица 6: Итоговая статистика
            summary_rows = []
            summary_rows.append(["Параметр", "Значение"])
            summary_rows.append(["Табельный номер", tab_number])
            summary_rows.append(["Количество уникальных ИНН", tab_data.get("unique_inn_count", 0)])
            summary_rows.append(["Лучший месяц", tab_data.get("best_month", "")])
            summary_rows.append(["Количество исходных файлов", len(tab_data.get("source_files", {}))])
            summary_rows.append(["Количество клиентов в RAW", len(tab_data.get("raw_data", {}))])
            
            summary_df = pd.DataFrame(summary_rows[1:], columns=summary_rows[0])
            tables_data.append(("Итоговая статистика", summary_df))
            
            self.logger.info(f"Создано {len(tables_data)} таблиц для листа {sheet_name}", "ExcelFormatter", "_create_debug_tab_sheets")
            
            if len(tables_data) == 0:
                self.logger.warning(f"Нет данных для создания листа {sheet_name}, пропускаем", "ExcelFormatter", "_create_debug_tab_sheets")
                continue
            
            # Объединяем все таблицы в один DataFrame для листа
            # Создаем вертикальное объединение таблиц с заголовками
            all_rows = []
            for table_name, table_df in tables_data:
                # Добавляем заголовок таблицы
                all_rows.append([table_name] + [""] * (len(table_df.columns) - 1))
                # Добавляем заголовки колонок
                all_rows.append(list(table_df.columns))
                # Добавляем данные
                for _, row in table_df.iterrows():
                    all_rows.append(list(row))
                # Пустая строка между таблицами
                all_rows.append([""] * len(table_df.columns))
            
            if all_rows:
                # Создаем DataFrame с максимальным количеством колонок
                max_cols = max(len(row) for row in all_rows) if all_rows else 1
                for row in all_rows:
                    while len(row) < max_cols:
                        row.append("")
                
                debug_df = pd.DataFrame(all_rows)
                self.logger.info(f"Сохранение листа {sheet_name} с {len(debug_df)} строками и {len(debug_df.columns)} колонками", "ExcelFormatter", "_create_debug_tab_sheets")
                try:
                    debug_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    self.logger.info(f"Лист {sheet_name} успешно создан для табельного номера {tab_number}", "ExcelFormatter", "_create_debug_tab_sheets")
                except Exception as e:
                    self.logger.error(f"Ошибка при сохранении листа {sheet_name}: {str(e)}", "ExcelFormatter", "_create_debug_tab_sheets", exc_info=True)
            else:
                self.logger.warning(f"Нет строк для создания листа {sheet_name} для табельного номера {tab_number}", "ExcelFormatter", "_create_debug_tab_sheets")
        
        self.logger.info("=== ЗАВЕРШЕНО СОЗДАНИЕ ДЕТАЛЬНЫХ ЛИСТОВ ===", "ExcelFormatter", "_create_debug_tab_sheets")
    
    def _create_with_openpyxl(self, raw_df: pd.DataFrame, summary_df: pd.DataFrame, calculated_df: pd.DataFrame,
                             normalized_df: pd.DataFrame, places_df: pd.DataFrame, final_df: pd.DataFrame,
                             output_path: str, statistics_df: Optional[pd.DataFrame] = None,
                             debug_tracker: Optional[DebugTabNumberTracker] = None) -> None:
        """
        Создает Excel файл с форматированием используя openpyxl.
        
        Args:
            raw_df: DataFrame с сырыми данными (может быть разбит на несколько листов)
            summary_df: DataFrame с исходными данными
            calculated_df: DataFrame с расчетными данными
            normalized_df: DataFrame с нормализованными данными
            places_df: DataFrame с Score и рангами
            final_df: DataFrame с итоговыми данными
            output_path: Путь для сохранения файла
            statistics_df: DataFrame со статистикой (опционально)
        """
        self.logger.info("Использование openpyxl для форматирования")
        self.logger.info(f"Режим форматирования: {FORMATTING_MODE} (full=полное, off=выключено, simple=упрощенное)", "ExcelFormatter", "_create_with_openpyxl")
        
        # Разбиваем raw_df на чанки (если больше 900 000 строк) только если RAW листы включены
        if ENABLE_RAW_SHEETS:
            raw_chunks = self._split_raw_df(raw_df, chunk_size=900_000)
            self.logger.info(f"RAW листы включены: будет создано {len(raw_chunks)} листа(ов) RAW", "ExcelFormatter", "_create_with_openpyxl")
        else:
            raw_chunks = []
            self.logger.info("RAW листы отключены (ENABLE_RAW_SHEETS=False), они не будут созданы", "ExcelFormatter", "_create_with_openpyxl")
        
        # Сначала сохраняем DataFrame в Excel через pandas
        from time import time as time_func
        save_start_time = time_func()
        last_log_time = save_start_time
        LOG_INTERVAL = 15  # Логируем прогресс каждые 15 секунд для большей видимости
        
        self.logger.info("Сохранение данных в Excel...")
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Сохраняем все чанки RAW (только если включены)
                total_raw_chunks = len(raw_chunks)
                for chunk_idx, (sheet_name, chunk_df) in enumerate(raw_chunks, 1):
                    if len(chunk_df) > 0:
                        # ВСЕГДА логируем перед началом сохранения каждого листа
                        chunk_rows = len(chunk_df)
                        chunk_cols = len(chunk_df.columns)
                        current_time = time_func()
                        elapsed = current_time - save_start_time
                        self.logger.info(
                            f"Начало сохранения листа '{sheet_name}' ({chunk_idx}/{total_raw_chunks}): "
                            f"{chunk_rows} строк × {chunk_cols} колонок (прошло {elapsed:.0f} сек)"
                        )
                        chunk_save_start = current_time
                        last_log_time = current_time
                        
                        # Запускаем поток для периодического логирования во время сохранения
                        import threading
                        save_logging_active = threading.Event()
                        save_logging_active.set()
                        
                        def log_save_progress():
                            """Периодически логирует прогресс сохранения"""
                            while save_logging_active.is_set():
                                threading.Event().wait(LOG_INTERVAL)
                                if save_logging_active.is_set():
                                    current_time = time_func()
                                    elapsed = current_time - save_start_time
                                    self.logger.info(
                                        f"Сохранение листа '{sheet_name}' ({chunk_idx}/{total_raw_chunks}) продолжается... "
                                        f"(прошло {elapsed:.0f} сек)"
                                    )
                        
                        progress_thread = threading.Thread(target=log_save_progress, daemon=True)
                        progress_thread.start()
                        
                        try:
                            chunk_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        except KeyboardInterrupt:
                            save_logging_active.clear()
                            self.logger.warning(f"Прерывание при сохранении листа '{sheet_name}'", "ExcelFormatter", "_create_with_openpyxl")
                            raise
                        finally:
                            save_logging_active.clear()
                        
                        # ВСЕГДА логируем после завершения сохранения
                        current_time = time_func()
                        elapsed = current_time - save_start_time
                        sheet_elapsed = current_time - chunk_save_start
                        self.logger.info(
                            f"Сохранен лист '{sheet_name}' ({chunk_idx}/{total_raw_chunks}) "
                            f"за {sheet_elapsed:.0f} сек (всего прошло {elapsed:.0f} сек)"
                        )
                        last_log_time = current_time
                
                # Сохраняем остальные листы
                other_sheets = [
                    ("Исходник", summary_df),
                    ("Расчет", calculated_df),
                    ("Нормализация", normalized_df),
                    ("Места и выбор", places_df),
                    ("Итог", final_df)
                ]
                if statistics_df is not None:
                    other_sheets.append(("Статистика", statistics_df))
                
                # Создаем детальные листы для табельных номеров из DEBUG_TAB_NUMBER
                self.logger.info(f"Проверка создания детальных листов: debug_tracker={debug_tracker is not None}", "ExcelFormatter", "_create_with_openpyxl")
                if debug_tracker:
                    tab_numbers = debug_tracker.get_all_tab_numbers()
                    self.logger.info(f"Табельные номера в трекере: {tab_numbers} (всего {len(tab_numbers)})", "ExcelFormatter", "_create_with_openpyxl")
                    if len(tab_numbers) > 0:
                        try:
                            self.logger.info("Вызов _create_debug_tab_sheets...", "ExcelFormatter", "_create_with_openpyxl")
                            self._create_debug_tab_sheets(debug_tracker, writer)
                            self.logger.info("_create_debug_tab_sheets завершен успешно", "ExcelFormatter", "_create_with_openpyxl")
                        except Exception as e:
                            self.logger.error(f"Ошибка при создании детальных листов: {str(e)}", "ExcelFormatter", "_create_with_openpyxl", exc_info=True)
                    else:
                        self.logger.warning("debug_tracker пуст, детальные листы не будут созданы", "ExcelFormatter", "_create_with_openpyxl")
                else:
                    self.logger.warning("debug_tracker не передан, детальные листы не будут созданы", "ExcelFormatter", "_create_with_openpyxl")
                
                for sheet_idx, (sheet_name, df) in enumerate(other_sheets, 1):
                    current_time = time_func()
                    if current_time - last_log_time >= LOG_INTERVAL:
                        elapsed = current_time - save_start_time
                        self.logger.info(f"Сохранение листа '{sheet_name}' ({sheet_idx}/{len(other_sheets)})... (прошло {elapsed:.0f} сек)")
                        last_log_time = current_time
                    try:
                        if sheet_name == "Статистика":
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                        else:
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    except KeyboardInterrupt:
                        self.logger.warning(f"Прерывание при сохранении листа '{sheet_name}'", "ExcelFormatter", "_create_with_openpyxl")
                        raise
                    current_time = time_func()
                    if current_time - last_log_time >= LOG_INTERVAL:
                        elapsed = current_time - save_start_time
                        self.logger.info(f"Сохранен лист '{sheet_name}' ({sheet_idx}/{len(other_sheets)}) (прошло {elapsed:.0f} сек)")
                        last_log_time = current_time
        except KeyboardInterrupt:
            self.logger.warning("Прерывание при сохранении данных в Excel", "ExcelFormatter", "_create_with_openpyxl")
            raise
        
        save_elapsed = time_func() - save_start_time
        self.logger.info(f"Данные сохранены в Excel за {save_elapsed:.0f} секунд")
        
        # Теперь форматируем файл
        self.logger.info("Начало форматирования Excel файла...")
        try:
            wb = load_workbook(output_path)
        except KeyboardInterrupt:
            self.logger.warning("Прерывание при загрузке Excel файла для форматирования", "ExcelFormatter", "_create_with_openpyxl")
            raise
        
        # Форматируем все листы
        # Собираем все листы RAW для форматирования (только если включены)
        sheet_data = {}
        if ENABLE_RAW_SHEETS:
            for sheet_name, chunk_df in raw_chunks:
                sheet_data[sheet_name] = chunk_df
        
        # Добавляем остальные листы
        sheet_data.update({
            "Исходник": summary_df,
            "Расчет": calculated_df,
            "Нормализация": normalized_df,
            "Места и выбор": places_df,
            "Итог": final_df
        })
        
        if statistics_df is not None:
            sheet_data["Статистика"] = statistics_df
        
        # Добавляем детальные листы для форматирования (если есть)
        debug_tab_sheets = []
        if debug_tracker and len(debug_tracker.get_all_tab_numbers()) > 0:
            for tab_number in debug_tracker.get_all_tab_numbers():
                sheet_name = f"Детально_{tab_number}"
                if len(sheet_name) > 31:
                    sheet_name = f"Дет_{tab_number[-8:]}"
                if sheet_name in wb.sheetnames:
                    debug_tab_sheets.append(sheet_name)
        
        total_sheets = len(sheet_data)
        from time import time
        format_start_time = time()
        last_progress_time = format_start_time
        PROGRESS_INTERVAL = 30  # Логируем прогресс каждые 30 секунд (максимум раз в минуту)
        
        try:
            for sheet_idx, (sheet_name, df) in enumerate(sheet_data.items(), 1):
                if sheet_name not in wb.sheetnames:
                    continue
                
                # ВСЕГДА логируем начало форматирования каждого листа
                current_time = time()
                elapsed = current_time - format_start_time
                total_rows = len(df)
                self.logger.info(f"Начало форматирования листа '{sheet_name}' ({sheet_idx}/{total_sheets}, {total_rows} строк)... (прошло {elapsed:.0f} сек)")
                last_progress_time = current_time
                
                try:
                    ws = wb[sheet_name]
                    if FORMATTING_MODE == "off":
                        # Форматирование выключено - форматируем только ТН и ИНН
                        self._format_sheet_minimal(ws, df, sheet_name)
                    elif sheet_name == "Статистика":
                        # Для листа статистики используем специальное форматирование
                        self._format_statistics_sheet_openpyxl(ws, df)
                    elif sheet_name.startswith("RAW"):
                        # Для всех листов RAW (RAW, RAW_2, RAW_3 и т.д.) используем стандартное форматирование
                        self._format_sheet_openpyxl(ws, df, sheet_name, sheet_idx, total_sheets)
                    else:
                        self._format_sheet_openpyxl(ws, df, sheet_name, sheet_idx, total_sheets)
                except KeyboardInterrupt:
                    self.logger.warning(f"Прерывание при форматировании листа '{sheet_name}'", "ExcelFormatter", "_create_with_openpyxl")
                    raise
                
                # ВСЕГДА логируем завершение форматирования каждого листа
                current_time = time()
                elapsed = current_time - format_start_time
                sheet_elapsed = current_time - last_progress_time
                self.logger.info(f"Завершено форматирование листа '{sheet_name}' ({sheet_idx}/{total_sheets}) за {sheet_elapsed:.0f} сек (всего прошло {elapsed:.0f} сек)")
                last_progress_time = current_time
            
            # Форматируем детальные листы
            for debug_sheet_name in debug_tab_sheets:
                if debug_sheet_name in wb.sheetnames:
                    try:
                        ws = wb[debug_sheet_name]
                        self._format_debug_tab_sheet(ws, debug_sheet_name)
                    except KeyboardInterrupt:
                        self.logger.warning(f"Прерывание при форматировании детального листа '{debug_sheet_name}'", "ExcelFormatter", "_create_with_openpyxl")
                        raise
            
            # Сохраняем файл
            format_elapsed = time() - format_start_time
            self.logger.info(f"Сохранение форматированного файла... (форматирование заняло {format_elapsed:.0f} сек)")
            try:
                wb.save(output_path)
            except KeyboardInterrupt:
                self.logger.warning("Прерывание при сохранении форматированного файла", "ExcelFormatter", "_create_with_openpyxl")
                raise
        except KeyboardInterrupt:
            self.logger.warning("Прерывание при форматировании Excel файла", "ExcelFormatter", "_create_with_openpyxl")
            raise
        self.logger.info(f"Файл {output_path} успешно создан с форматированием (openpyxl)")
    
    def _format_sheet_openpyxl(self, ws, df: pd.DataFrame, sheet_name: str = "", sheet_idx: int = 0, total_sheets: int = 0) -> None:
        """
        Форматирует лист Excel используя openpyxl (оптимизированная версия).
        
        Args:
            ws: Рабочий лист openpyxl
            df: DataFrame с данными
            sheet_name: Имя листа (для логирования)
            sheet_idx: Номер листа (для логирования)
            total_sheets: Всего листов (для логирования)
        """
        from time import time
        format_sheet_start_time = time()
        last_progress_time = format_sheet_start_time
        PROGRESS_INTERVAL = 15  # Логируем прогресс каждые 15 секунд для большей видимости
        
        total_rows = len(df)
        total_cols = len(df.columns)
        
        # Логируем начало форматирования листа
        mode_desc = {"full": "полное", "off": "выключено (только ТН и ИНН)", "simple": "упрощенное (ТН, ИНН, ФИО, ТБ, ГОСБ)"}.get(FORMATTING_MODE, FORMATTING_MODE)
        self.logger.info(f"Форматирование '{sheet_name}': {total_rows} строк, {total_cols} колонок (режим: {mode_desc})")
        
        # Фиксируем первую строку и 4 колонку (после ФИО)
        ws.freeze_panes = "E2"
        
        # Форматируем заголовки (первая строка)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        self.logger.debug(f"Заголовки отформатированы для '{sheet_name}'", "ExcelFormatter", "_format_sheet_openpyxl")
        
        # ОПТИМИЗАЦИЯ: Настраиваем ширину колонок
        self.logger.debug(f"Настройка ширины колонок для '{sheet_name}' ({total_cols} колонок)", "ExcelFormatter", "_format_sheet_openpyxl")
        for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            col_letter = get_column_letter(col_idx)
            
            # Вычисляем оптимальную ширину на основе содержимого
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Учитываем содержимое всех ячеек в колонке (первые 100 строк для производительности)
            for row in ws.iter_rows(min_row=2, max_row=min(102, ws.max_row), min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Применяем ограничения
            width = max(self.min_width, min(max_length + 2, self.max_width))
            ws.column_dimensions[col_letter].width = width
            
            # Логируем прогресс для больших листов каждые 15 секунд
            if total_cols > 20 and col_idx % 10 == 0:
                current_time = time()
                if current_time - last_progress_time >= PROGRESS_INTERVAL:
                    elapsed = current_time - format_sheet_start_time
                    self.logger.info(f"Форматирование '{sheet_name}': колонка {col_idx}/{total_cols} (прошло {elapsed:.0f} сек)")
                    last_progress_time = current_time
        
        self.logger.debug(f"Ширина колонок настроена для '{sheet_name}'", "ExcelFormatter", "_format_sheet_openpyxl")
        
        # ОПТИМИЗАЦИЯ: Настраиваем выравнивание и форматирование для всех ячеек (батчами)
        # Определяем базовые колонки (текстовые)
        base_columns = ["Табельный", "ТБ", "ФИО"]
        simple_format_columns = ["Табельный", "ТБ", "ФИО", "ИНН", "ГОСБ"]  # Колонки для упрощенного форматирования
        
        # Формат для чисел: разделитель разрядов и два знака после запятой
        number_format = "#,##0.00"
        # Формат для рангов: целое число с разделителем разрядов (без дробной части)
        rank_format = "#,##0"
        # Текстовый формат для сохранения лидирующих нулей
        text_format = "@"
        
        # Создаем объекты выравнивания один раз (переиспользуем)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center")
        
        # ОПТИМИЗАЦИЯ: Определяем типы колонок заранее (один раз)
        col_types = {}
        for col_idx in range(1, len(df.columns) + 1):
            col_name = ws.cell(row=1, column=col_idx).value
            if col_name == "Табельный":
                col_types[col_idx] = "tab"
            elif col_name == "ИНН":
                col_types[col_idx] = "inn"
            elif col_name in base_columns:
                col_types[col_idx] = "text"
            elif col_name == "Количество уникальных ИНН":
                # Целое число с разделителем разрядов без дробной части
                col_types[col_idx] = "inn_count"
            elif col_name and col_name.startswith("Score"):
                col_types[col_idx] = "score"
            elif col_name and "_norm" in col_name:
                col_types[col_idx] = "norm"
            elif col_name and col_name.startswith("Место"):
                col_types[col_idx] = "rank"
            elif col_name == "Лучший месяц":
                col_types[col_idx] = "text"
            else:
                col_types[col_idx] = "number"
        
        # ОПТИМИЗАЦИЯ: Для всех RAW листов используем упрощенное форматирование (только заголовки)
        # Для остальных листов - полное форматирование
        if sheet_name.startswith("RAW"):
            # Для всех RAW листов (RAW, RAW_2, RAW_3 и т.д.): форматируем только заголовки (без обработки каждой ячейки)
            # Это значительно ускоряет форматирование для больших листов (с 44 минут до ~1 минуты)
            self.logger.info(f"Форматирование листа '{sheet_name}': упрощенный режим (только заголовки, {total_rows} строк)")
            # Для RAW листов не форматируем ячейки - только заголовки уже отформатированы выше
        else:
            # Для остальных листов: полное форматирование
            if total_rows == 0:
                ws.auto_filter.ref = ws.dimensions
                return
            
            self.logger.debug(f"Начало форматирования ячеек для '{sheet_name}' ({total_rows} строк)", "ExcelFormatter", "_format_sheet_openpyxl")
            batch_size = 1000  # Обрабатываем по 1000 строк за раз
            processed_rows = 0
            
            for batch_start in range(2, ws.max_row + 1, batch_size):
                batch_end = min(batch_start + batch_size, ws.max_row + 1)
                
                for row_idx in range(batch_start, batch_end):
                    row = ws[row_idx]
                    for col_idx, cell in enumerate(row, start=1):
                        if col_idx not in col_types:
                            continue
                        
                        col_type = col_types.get(col_idx, "number")
                        col_name = ws.cell(row=1, column=col_idx).value
                        
                        # Определяем, нужно ли форматировать эту колонку в зависимости от режима
                        should_format = True
                        if FORMATTING_MODE == "simple":
                            # В упрощенном режиме форматируем только ТН, ИНН, ФИО, ТБ, ГОСБ
                            should_format = col_name in simple_format_columns
                        elif FORMATTING_MODE == "off":
                            # В режиме выключено форматируем только ТН и ИНН
                            should_format = col_name in ["Табельный", "ИНН"]
                        
                        # ТН и ИНН всегда форматируются (независимо от режима)
                        if col_type == "tab":
                            cell.number_format = text_format
                            cell.alignment = align_left
                        elif col_type == "inn":
                            cell.number_format = text_format
                            cell.alignment = align_left
                        elif FORMATTING_MODE == "off":
                            # В режиме выключено не форматируем остальные колонки
                            continue
                        elif not should_format:
                            # Не форматируем эту колонку (оставляем как есть)
                            continue
                        elif col_type == "text":
                            cell.alignment = align_left
                        elif col_type == "score" or col_type == "norm":
                            if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.number_format = number_format
                                cell.alignment = align_right
                            else:
                                cell.alignment = align_right
                        elif col_type == "rank" or col_type == "inn_count":
                            # Ранги и количество уникальных ИНН: целое число с разделителем разрядов
                            if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.number_format = rank_format
                                cell.alignment = align_right
                            else:
                                cell.alignment = align_right
                        else:  # number
                            if pd.notna(cell.value) and isinstance(cell.value, (int, float)):
                                cell.number_format = number_format
                                cell.alignment = align_right
                            else:
                                cell.alignment = align_left
                
                processed_rows = batch_end - 1
                # Логируем прогресс каждые 15 секунд
                current_time = time()
                if current_time - last_progress_time >= PROGRESS_INTERVAL:
                    elapsed = current_time - format_sheet_start_time
                    progress_pct = (processed_rows / total_rows) * 100 if total_rows > 0 else 0
                    self.logger.info(f"Форматирование '{sheet_name}': обработано {processed_rows}/{total_rows} строк ({progress_pct:.1f}%, прошло {elapsed:.0f} сек)")
                    last_progress_time = current_time
            
            self.logger.debug(f"Форматирование ячеек завершено для '{sheet_name}'", "ExcelFormatter", "_format_sheet_openpyxl")
        
        # Включаем автофильтр
        ws.auto_filter.ref = ws.dimensions
    
    def _format_debug_tab_sheet(self, ws, sheet_name: str) -> None:
        """
        Форматирует детальный лист для табельного номера.
        
        Args:
            ws: Рабочий лист openpyxl
            sheet_name: Имя листа
        """
        # Фиксируем первую строку
        ws.freeze_panes = "A2"
        
        # Форматируем заголовки таблиц (строки с названиями таблиц)
        section_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        section_font = Font(bold=True, size=14, color="000080")
        section_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Форматируем заголовки колонок (вторая строка после названия таблицы)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Обычный текст
        text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        number_alignment = Alignment(horizontal="right", vertical="center")
        
        current_row = 1
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            
            # Если это заголовок таблицы (первая колонка заполнена, остальные пустые или почти пустые)
            if cell_value and isinstance(cell_value, str) and len(cell_value) > 0:
                # Проверяем, является ли это заголовком таблицы
                is_table_header = True
                for col in range(2, min(10, ws.max_column + 1)):
                    other_cell = ws.cell(row=current_row, column=col).value
                    if other_cell and str(other_cell).strip():
                        is_table_header = False
                        break
                
                if is_table_header:
                    # Форматируем заголовок таблицы
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = section_font
                        cell.fill = section_fill
                        cell.alignment = section_alignment
                    
                    # Следующая строка - заголовки колонок
                    if current_row + 1 <= ws.max_row:
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=current_row + 1, column=col)
                            if cell.value:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                        current_row += 2
                        continue
            
            # Форматируем обычные строки
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=current_row, column=col)
                if cell.value is None:
                    continue
                
                # Проверяем, является ли значение числом
                try:
                    num_value = float(cell.value)
                    cell.alignment = number_alignment
                    cell.number_format = "#,##0.00"
                except (ValueError, TypeError):
                    cell.alignment = text_alignment
            
            current_row += 1
        
        # Настраиваем ширину колонок
        for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=min(100, ws.max_row)), start=1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            width = max(15, min(max_length + 2, 100))
            ws.column_dimensions[col_letter].width = width
        
        self.logger.debug(f"Детальный лист '{sheet_name}' отформатирован", "ExcelFormatter", "_format_debug_tab_sheet")
    
    def _format_sheet_minimal(self, ws, df: pd.DataFrame, sheet_name: str) -> None:
        """
        Минимальное форматирование листа: только ТН и ИНН (используется при FORMATTING_MODE="off").
        
        Args:
            ws: Рабочий лист openpyxl
            df: DataFrame с данными
            sheet_name: Имя листа (для логирования)
        """
        # Фиксируем первую строку и 4 колонку (после ФИО)
        ws.freeze_panes = "E2"
        
        # Форматируем заголовки (первая строка)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Настраиваем ширину колонок
        for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            col_letter = get_column_letter(col_idx)
            col_name = ws.cell(row=1, column=col_idx).value
            
            # Вычисляем оптимальную ширину
            max_length = len(str(col_name)) if col_name else 0
            for row in ws.iter_rows(min_row=2, max_row=min(102, ws.max_row), min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            width = max(self.min_width, min(max_length + 2, self.max_width))
            ws.column_dimensions[col_letter].width = width
        
        # Форматируем только ТН и ИНН
        text_format = "@"
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for col_idx in range(1, len(df.columns) + 1):
            col_name = ws.cell(row=1, column=col_idx).value
            if col_name in ["Табельный", "ИНН"]:
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = text_format
                        cell.alignment = align_left
        
        self.logger.debug(f"Минимальное форматирование применено к '{sheet_name}' (только ТН и ИНН)", "ExcelFormatter", "_format_sheet_minimal")
    
    def _format_statistics_sheet_openpyxl(self, ws, df: pd.DataFrame) -> None:
        """
        Форматирует лист статистики используя openpyxl.
        
        Args:
            ws: Рабочий лист openpyxl
            df: DataFrame со статистикой
        """
        # Фиксируем первую строку
        ws.freeze_panes = "A2"
        
        # Форматируем заголовки разделов (строки с одним значением в первой колонке и пустой второй)
        section_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        section_font = Font(bold=True, size=14, color="000080")
        section_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Форматируем заголовки таблиц (строки с двумя значениями, где второе пустое)
        table_header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        table_header_font = Font(bold=True, size=11)
        table_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Обычный текст
        text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        number_alignment = Alignment(horizontal="right", vertical="center")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
            for col_idx, cell in enumerate(row):
                if cell.value is None:
                    continue
                
                value = str(cell.value)
                # Проверяем, является ли это заголовком раздела (первая колонка заполнена, вторая пустая)
                if col_idx == 0 and len(row) > 1:
                    next_cell_value = row[1].value if len(row) > 1 else None
                    if next_cell_value is None or str(next_cell_value).strip() == "":
                        # Это заголовок раздела
                        cell.font = section_font
                        cell.fill = section_fill
                        cell.alignment = section_alignment
                    elif col_idx == 0 and row_idx == 1:
                        # Первая строка - заголовок таблицы
                        cell.font = table_header_font
                        cell.fill = table_header_fill
                        cell.alignment = table_header_alignment
                    else:
                        # Обычный текст
                        cell.alignment = text_alignment
                elif col_idx == 1 and row_idx == 1:
                    # Вторая колонка первой строки - заголовок таблицы
                    cell.font = table_header_font
                    cell.fill = table_header_fill
                    cell.alignment = table_header_alignment
                else:
                    # Проверяем, является ли значение числом
                    try:
                        num_value = float(value)
                        cell.alignment = number_alignment
                        cell.number_format = "#,##0"
                    except (ValueError, TypeError):
                        cell.alignment = text_alignment
        
        # Настраиваем ширину колонок
        for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=min(100, ws.max_row)), start=1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            width = max(15, min(max_length + 2, 100))
            ws.column_dimensions[col_letter].width = width
        
        self.logger.debug("Лист 'Статистика' отформатирован", "ExcelFormatter", "_format_statistics_sheet_openpyxl")
    
    # Методы xlsxwriter удалены - используется только openpyxl


# ============================================================================
# ОСНОВНАЯ ФУНКЦИЯ
# ============================================================================

def main():
    """Основная функция приложения."""
    # Инициализируем логгер
    logger = Logger(log_dir=LOG_DIR, level=LOG_LEVEL, theme=LOG_THEME)
    
    logger.info("=" * 80, "main", "main")
    logger.info("Запуск обработки месячных данных", "main", "main")
    logger.info("=" * 80, "main", "main")
    
    # Логируем все параметры конфигурации
    logger.info("", "main", "main")
    logger.info("ПАРАМЕТРЫ КОНФИГУРАЦИИ ПРИЛОЖЕНИЯ:", "main", "main")
    logger.info("-" * 80, "main", "main")
    
    # Пути к каталогам
    logger.info(f"INPUT_DIR = '{INPUT_DIR}' - Каталог с входными данными", "main", "main")
    logger.info(f"OUTPUT_DIR = '{OUTPUT_DIR}' - Каталог для выходных файлов", "main", "main")
    logger.info(f"LOG_DIR = '{LOG_DIR}' - Каталог для логов", "main", "main")
    
    # Параметры логирования
    logger.info(f"LOG_LEVEL = '{LOG_LEVEL}' - Уровень логирования: DEBUG (в файлы) - детальное, INFO (в консоль) - верхнеуровневое", "main", "main")
    logger.info(f"LOG_THEME = '{LOG_THEME}' - Тема логов (используется в имени файла)", "main", "main")
    
    # Параметры статистики
    logger.info(f"ENABLE_STATISTICS = {ENABLE_STATISTICS} - Сбор и вывод статистики: True - собирать статистику и создавать лист 'Статистика', False - не собирать", "main", "main")
    
    # Параметры оптимизации производительности
    logger.info(f"ENABLE_PARALLEL_LOADING = {ENABLE_PARALLEL_LOADING} - Параллельная загрузка файлов: True - параллельная загрузка, False - последовательная", "main", "main")
    logger.info(f"MAX_WORKERS = {MAX_WORKERS} - Количество потоков для параллельной загрузки (рекомендуется 8 по числу виртуальных ядер)", "main", "main")
    logger.info(f"ENABLE_CHUNKING = {ENABLE_CHUNKING} - Использование chunking для больших файлов: True - использовать chunking, False - загружать целиком (chunking медленный, отключен)", "main", "main")
    logger.info(f"CHUNK_SIZE = {CHUNK_SIZE} - Размер chunk для чтения больших файлов (строк)", "main", "main")
    logger.info(f"CHUNKING_THRESHOLD_MB = {CHUNKING_THRESHOLD_MB} - Порог размера файла для chunking (МБ) - если файл больше, используем chunking", "main", "main")
    
    # Параметры детального логирования
    debug_tab_str = str(DEBUG_TAB_NUMBER) if DEBUG_TAB_NUMBER else "None"
    logger.info(f"DEBUG_TAB_NUMBER = {debug_tab_str} - Список табельных номеров для детального логирования (например, ['12345678', '87654321'] или None для отключения)", "main", "main")
    if DEBUG_TAB_NUMBER and len(DEBUG_TAB_NUMBER) > 0:
        logger.info(f"  Детальное логирование включено для табельных номеров: {', '.join(DEBUG_TAB_NUMBER)}", "main", "main")
    else:
        logger.info(f"  Детальное логирование отключено", "main", "main")
    
    # Параметр выбора режима данных
    logger.info(f"DATA_MODE = '{DATA_MODE}' - Режим данных: 'TEST' - тестовые данные, 'PROM' - пром данные. Определяет, какие columns использовать из конфигурации (columns_test или columns_prom)", "main", "main")
    
    # Триггер для формирования RAW листов
    logger.info(f"ENABLE_RAW_SHEETS = {ENABLE_RAW_SHEETS} - Формирование RAW листов: True - формировать RAW листы, False - не формировать (по умолчанию выключено)", "main", "main")
    
    # Триггер для форматирования листов
    formatting_desc = {
        "full": "полное форматирование (как сейчас, по умолчанию)",
        "off": "форматирование выключено (листы формируются, но не переформатируются, кроме ТН и ИНН - их форматы всегда работают)",
        "simple": "упрощенное форматирование (только ТН, ИНН, ФИО, ТБ, ГОСБ и заголовок, не форматируем данные показателей и расчетов)"
    }
    logger.info(f"FORMATTING_MODE = '{FORMATTING_MODE}' - Режим форматирования: {formatting_desc.get(FORMATTING_MODE, 'неизвестный режим')}", "main", "main")
    
    # Информация о маппинге ТБ
    logger.info(f"TB_MAPPINGS - Маппинг территориальных банков: определено {len(TB_MAPPINGS)} банков", "main", "main")
    if len(TB_MAPPINGS) > 0:
        tb_list = ", ".join([f"{key} ({mapping.short_name})" for key, mapping in TB_MAPPINGS.items()])
        logger.info(f"  Банки: {tb_list}", "main", "main")
    
    # Информация о доступности openpyxl
    logger.info(f"OPENPYXL_AVAILABLE = {OPENPYXL_AVAILABLE} - Доступность openpyxl для форматирования Excel файлов", "main", "main")
    
    logger.info("-" * 80, "main", "main")
    logger.info("", "main", "main")
    
    try:
        # Создаем выходной каталог, если его нет
        output_path = Path(OUTPUT_DIR)
        output_path.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"Входной каталог: {INPUT_DIR}", "main", "main")
        logger.info(f"Выходной каталог: {OUTPUT_DIR}", "main", "main")
        
        # Инициализируем процессор файлов
        processor = FileProcessor(input_dir=INPUT_DIR, logger_instance=logger)
        
        # Загружаем все файлы
        logger.info("Этап 1: Загрузка файлов", "main", "main")
        processor.load_all_files()
        
        # Собираем уникальные табельные номера
        logger.info("Этап 2: Сбор уникальных табельных номеров", "main", "main")
        processor.collect_unique_tab_numbers()
        
        # Подготавливаем сводные данные
        logger.info("Этап 3: Подготовка сводных данных", "main", "main")
        summary_df = processor.prepare_summary_data()
        
        if summary_df.empty:
            logger.error("Сводные данные пусты, обработка завершена", "main", "main")
            return
        
        # Подготавливаем сырые данные для листа RAW
        logger.info("Этап 4: Подготовка сырых данных", "main", "main")
        raw_df = processor.prepare_raw_data()
        
        # Подготавливаем расчетные данные
        logger.info("Этап 5: Подготовка расчетных данных", "main", "main")
        calculated_df = processor.prepare_calculated_data(summary_df)
        
        # Создаем менеджер конфигурации
        config_manager = ConfigManager()
        
        # Нормализуем показатели (вариант 3)
        logger.info("Этап 6: Нормализация показателей", "main", "main")
        normalized_df = processor._normalize_indicators(calculated_df, config_manager)
        
        # Рассчитываем лучший месяц (вариант 3)
        logger.info("Этап 7: Расчет лучшего месяца", "main", "main")
        places_df, final_df = processor._calculate_best_month_variant3(calculated_df, normalized_df, config_manager, raw_df)
        
        # Подготавливаем лист статистики (если включен)
        statistics_df = None
        if ENABLE_STATISTICS:
            logger.info("Этап 8: Подготовка статистики", "main", "main")
            statistics_df = processor.prepare_statistics_sheet()
        
        # Формируем имя выходного файла с датой и временем
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_path / f"Сводные_данные_{timestamp}.xlsx"
        
        # Создаем форматтер
        formatter = ExcelFormatter(logger_instance=logger)
        
        # Сохраняем данные в Excel с форматированием (6 основных листов + статистика, если включена)
        logger.info(f"Этап 9: Сохранение результата в {output_file}", "main", "main")
        try:
            formatter.create_formatted_excel(raw_df, summary_df, calculated_df, normalized_df, places_df, final_df, str(output_file), statistics_df, processor.debug_tracker)
        except KeyboardInterrupt:
            # Прерывание при сохранении/форматировании Excel - пробрасываем дальше
            raise
        
        if ENABLE_STATISTICS and statistics_df is not None:
            logger.info("Лист 'Статистика' добавлен в файл", "main", "main")
        
        logger.info("=" * 80, "main", "main")
        logger.info(f"Обработка завершена успешно. Результат сохранен в: {output_file}", "main", "main")
        logger.info(f"Обработано табельных номеров: {len(summary_df)}", "main", "main")
        logger.info(f"Колонок в результате: {len(summary_df.columns)}", "main", "main")
        logger.info("=" * 80, "main", "main")
        
    except KeyboardInterrupt:
        # Обработка прерывания пользователем (Ctrl+C)
        logger.warning("=" * 80, "main", "main")
        logger.warning("Обработка прервана пользователем (Ctrl+C)", "main", "main")
        logger.warning("=" * 80, "main", "main")
        sys.exit(0)
        
    except Exception as e:
        logger.error(f"Критическая ошибка при выполнении: {str(e)}", "main", "main")
        import traceback
        logger.error(f"Трассировка: {traceback.format_exc()}", "main", "main")
        sys.exit(1)


if __name__ == "__main__":
    main()

